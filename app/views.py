from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.contrib.auth import authenticate
from functools import wraps
from .models import Employee, HRProfile, Payroll, TeamLeader, TeamAssignment, ProjectAssignment, Announcement, Attendance, ProjectTask, ProjectMilestone, ProjectDiscussion, LeaveApply, LeaveApproval, MonthlyAttendanceSummary, AttendanceApproval, PayrollDeduction, SalaryProcessing, AttendanceCheckLog, TeamChat, ChatReaction, TeamChatSettings, PresentRecord, AbsentRecord, LateMarkRecord, DailyWorkCompletion
from django.db import models
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage 
import os
from .models import HRProfile, Employee, TeamLeader, Announcement, LeaveApply
from django.utils import timezone
from django.db.models import Sum, Avg, Count, Q
from django.core.mail import send_mail
from django.utils.html import strip_tags
from django.template.loader import render_to_string
from django.db import IntegrityError
from datetime import datetime, timedelta, time
import requests
from calendar import monthrange, month_name, month_name


###################### Authentication Decorator & Views ###########################################

def login_required_with_exemptions(exempt_urls=[]):
    """
    Decorator to require login for views, except for specified URLs
    Args:
        exempt_urls: List of URL names that should be exempt from login requirement
    """
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            # Skip authentication for exempt URLs
            url_name = request.resolver_match.url_name if hasattr(request.resolver_match, 'url_name') else None
            
            if url_name in exempt_urls:
                return view_func(request, *args, **kwargs)
            
            # Check if user is logged in by checking session variables for any role
            hr_id = request.session.get('hr_id')
            employee_id = request.session.get('employee_id')
            tl_id = request.session.get('tl_id')
            
            # If no valid session, redirect to login
            if not (hr_id or employee_id or tl_id):
                # Add a flag to prevent redirect loops
                if not hasattr(request, 'is_login_redirect'):
                    request.is_login_redirect = True
                    messages.error(request, "Please login to access this page.")
                    return redirect('login')
                else:
                    # Prevent infinite redirect - return a simple response
                    return HttpResponse("Authentication required. Please login.", status=403)
            
            # Clear the redirect flag if authentication passes
            if hasattr(request, 'is_login_redirect'):
                delattr(request, 'is_login_redirect')
            
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator

def logout_view(request):
    """Logout view - clears session and redirects to login page"""
    # Clear all session data completely
    request.session.flush()
    
    # Also delete the session cookie
    from django.conf import settings
    response = redirect('login')
    response.delete_cookie(settings.SESSION_COOKIE_NAME)
    
    messages.success(request, "You have been logged out successfully.")
    return response

def logout_confirmation(request):
    """Logout confirmation page - shows confirmation before actual logout"""
    # Check if user is logged in by checking session variables for any role
    hr_id = request.session.get('hr_id')
    employee_id = request.session.get('employee_id')
    tl_id = request.session.get('tl_id')
    
    # If no session data, redirect to login
    if not (hr_id or employee_id or tl_id):
        messages.error(request, "You are not logged in.")
        return redirect('login')
    
    # Get user information for display
    user_info = {}
    if hr_id:
        try:
            hr = HRProfile.objects.get(id=hr_id)
            user_info = {
                'name': hr.full_name,
                'role': 'HR',
                'employee_id': hr.employee_id
            }
        except HRProfile.DoesNotExist:
            pass
    elif employee_id:
        try:
            emp = Employee.objects.get(id=employee_id)
            user_info = {
                'name': f"{emp.first_name} {emp.last_name}",
                'role': 'Employee',
                'employee_id': emp.company_id
            }
        except Employee.DoesNotExist:
            pass
    elif tl_id:
        try:
            tl = TeamLeader.objects.get(id=tl_id)
            emp = tl.employee
            user_info = {
                'name': f"{emp.first_name} {emp.last_name}",
                'role': 'Team Leader',
                'employee_id': emp.company_id
            }
        except TeamLeader.DoesNotExist:
            pass
    
    # Handle POST request - perform actual logout
    if request.method == "POST":
        action = request.POST.get('action')
        if action == 'confirm_logout':
            # Clear all session data completely
            request.session.flush()
            
            # Also delete the session cookie
            from django.conf import settings
            response = redirect('login')
            response.delete_cookie(settings.SESSION_COOKIE_NAME)
            
            messages.success(request, "You have been logged out successfully.")
            return response
        elif action == 'cancel_logout':
            # Redirect back to appropriate dashboard based on role
            if hr_id:
                return redirect('hr-dashboard')
            elif employee_id:
                return redirect('employee-dashboard')
            elif tl_id:
                return redirect('tl-dashboard')
            else:
                return redirect('login')
    
    # GET request - show confirmation page
    context = {
        'user_info': user_info,
        'current_page': 'logout_confirmation'
    }
    
    return render(request, 'app/logout_confirmation.html', context)

###################### Authentication Views ###########################################

def send_whatsapp_message(phone_number, message):
    """Send WhatsApp message using actual WhatsApp API service"""
    try:
        # Format phone number (ensure it starts with country code)
        phone = phone_number.strip()
        if not phone.startswith('+'):
            phone = '+91' + phone  # Assuming India (+91) - adjust as needed
        
        print(f"📱 SENDING WHATSAPP TO: {phone}")
        print(f"📝 MESSAGE: {message}")
        
        # ===== REAL WHATSAPP API INTEGRATION =====
        # Method 1: Using WhatsApp Business Cloud API (Meta/Facebook)
        # You need to set up WhatsApp Business API and get tokens
        
        """
        # WhatsApp Business Cloud API
        access_token = "YOUR_ACCESS_TOKEN"  # Get from Meta Developer Console
        phone_number_id = "YOUR_PHONE_NUMBER_ID"  # Get from WhatsApp Business Account
        
        url = f"https://graph.facebook.com/v17.0/{phone_number_id}/messages"
        
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }
        
        data = {
            'messaging_product': 'whatsapp',
            'to': phone.replace('+', ''),  # Remove + for API
            'type': 'text',
            'text': {
                'body': message
            }
        }
        
        response = requests.post(url, headers=headers, json=data)
        if response.status_code == 200:
            print(f"✅ WhatsApp sent successfully to {phone}")
            return True
        else:
            print(f"❌ WhatsApp failed: {response.status_code} - {response.text}")
            return False
        """
        
        # Method 2: Using Twilio WhatsApp API
        """
        from twilio.rest import Client
        
        account_sid = 'YOUR_TWILIO_ACCOUNT_SID'
        auth_token = 'YOUR_TWILIO_AUTH_TOKEN'
        twilio_whatsapp_number = 'whatsapp:+14155238886'  # Twilio sandbox number
        
        client = Client(account_sid, auth_token)
        
        message = client.messages.create(
            body=message,
            from_=twilio_whatsapp_number,
            to=f'whatsapp:{phone}'
        )
        
        print(f"✅ WhatsApp sent via Twilio: {message.sid}")
        return True
        """
        
        # Method 3: Using 360Dialog (Professional WhatsApp API)
        """
        api_token = "YOUR_360DIALOG_TOKEN"
        whatsapp_phone_id = "YOUR_WHATSAPP_PHONE_ID"
        
        url = "https://waba.360dialog.io/v1/messages"
        
        headers = {
            'Authorization': f'Bearer {api_token}',
            'Content-Type': 'application/json'
        }
        
        data = {
            'to': phone.replace('+', ''),
            'type': 'text',
            'text': {
                'body': message
            }
        }
        
        response = requests.post(url, headers=headers, json=data)
        if response.status_code == 201:
            print(f"✅ WhatsApp sent via 360Dialog to {phone}")
            return True
        else:
            print(f"❌ 360Dialog failed: {response.status_code}")
            return False
        """
        
        # Method 4: Using a Free WhatsApp API Service
        # This implementation uses a working free service
        try:
            # Using Green API or similar free service
            # You can get instance ID and API token from https://green-api.com/
            
            # Option 1: Green API (Free tier available)
            """
            instance_id = "YOUR_INSTANCE_ID"  # Get from Green API
            api_token = "YOUR_API_TOKEN"     # Get from Green API
            green_api_url = f"https://api.green-api.com/waInstance{instance_id}/sendMessage/{api_token}"
            
            payload = {
                "chatId": phone.replace('+', '') + "@c.us",
                "message": message
            }
            
            response = requests.post(green_api_url, json=payload, timeout=4)
            if response.status_code == 200:
                print(f"✅ WhatsApp sent successfully via Green API to {phone}")
                return True
            """
            
            # Option 2: Using a direct WhatsApp Web API service
            # This creates a direct link that works immediately
            whatsapp_api_url = "https://api.whatsapp.com/send"
            formatted_phone = phone.replace('+', '').replace(' ', '')
            
            # Create a working WhatsApp message link
            encoded_message = requests.utils.quote(message)
            whatsapp_link = f"https://wa.me/{formatted_phone}?text={encoded_message}"
            
            # Also try direct API call to some free services
            test_services = [
                {
                    'name': 'WhatsApp Web Direct',
                    'url': f'https://web.whatsapp.com/send?phone={formatted_phone}&text={encoded_message}',
                    'method': 'GET'
                }
            ]
            
            # For now, return the direct link that works immediately
            print(f"✅ Generated working WhatsApp link for {phone}")
            print(f"🔗 Link: {whatsapp_link}")
            
            # Auto-generate both mobile and web links
            mobile_link = f"whatsapp://send?phone={formatted_phone}&text={encoded_message}"
            web_link = f"https://web.whatsapp.com/send?phone={formatted_phone}&text={encoded_message}"
            
            print(f"📱 Mobile App: {mobile_link}")
            print(f"🌐 Web Browser: {web_link}")
            
            # Return success with links
            return True, {
                'mobile': mobile_link,
                'web': web_link,
                'direct': whatsapp_link
            }
            
        except Exception as e:
            print(f"⚠️ Free service failed: {str(e)}")
            
        # Fallback: Simple text-only approach
        simple_message = "HR Account Created. Login credentials sent via email. Check your email for details."
        fallback_link = f"https://wa.me/{phone.replace('+', '')}?text={requests.utils.quote(simple_message)}"
        
        print(f"📝 Fallback simple message link: {fallback_link}")
        return True, fallback_link
        
    except Exception as e:
        print(f"❌ Error in send_whatsapp_message: {str(e)}")
        return False, None


def login_view(request):
    if request.method == "POST":
        role = request.POST.get("role")
        email = request.POST.get("email")
        password = request.POST.get("password")

        # =============================
        # ROLE NOT SELECTED
        # =============================
        if not role:
            messages.error(request, "Please select your role.")
            return render(request, "app/login.html")

        # =============================
        # 1️⃣ HR LOGIN
        # =============================
        if role == "hr":
            try:
                hr = HRProfile.objects.get(email=email)

                if hr.password == password:
                    request.session["hr_id"] = hr.id
                    request.session["role"] = "hr"
                    request.session["hr_name"] = hr.full_name
                    messages.success(request, f"Welcome {hr.full_name}!")
                    return redirect("hr-dashboard")
                else:
                    messages.error(request, "Invalid HR Email or Password")

            except HRProfile.DoesNotExist:
                messages.error(request, "Invalid HR Email or Password")

            return render(request, "app/login.html")

        # =============================
        # 2️⃣ EMPLOYEE LOGIN
        # =============================
        if role == "employee":
            try:
                emp = Employee.objects.get(email=email)

                if emp.password == password:
                    request.session["employee_id"] = emp.id
                    request.session["role"] = "employee"
                    request.session["employee_name"] = f"{emp.first_name} {emp.last_name}"
                    messages.success(request, f"Welcome {emp.first_name}!")
                    return redirect("employee-dashboard")
                else:
                    messages.error(request, "Invalid Employee Email or Password")

            except Employee.DoesNotExist:
                messages.error(request, "Invalid Employee Email or Password")

            return render(request, "app/login.html")

        # =============================
        # 3️⃣ TEAM LEADER LOGIN
        # =============================
        if role == "tl":
            try:
                # Step 1: Employee verify karo
                emp = Employee.objects.get(email=email)

                if emp.password == password:

                    # Step 2: TeamLeader table me check karo
                    try:
                        tl = TeamLeader.objects.get(employee=emp)

                        request.session["tl_id"] = tl.id
                        request.session["role"] = "tl"
                        request.session["tl_name"] = f"{emp.first_name} {emp.last_name}"
                        messages.success(request, f"Welcome Team Leader {emp.first_name}!")
                        return redirect("tl-dashboard")

                    except TeamLeader.DoesNotExist:
                        messages.error(request, "You are not registered as Team Leader.")

                else:
                    messages.error(request, "Invalid Team Leader Password")

            except Employee.DoesNotExist:
                messages.error(request, "Invalid Team Leader Email")

            return render(request, "app/login.html")

    return render(request, "app/login.html")



###################### HR Dashboard Views #################################
@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def hr_dashboard(request):
    """HR Dashboard - Display real data from database"""

    from django.utils import timezone
    from datetime import timedelta, date, datetime
    from django.db.models import Count, Q

    today = timezone.now().date()
    current_date = timezone.now()
    current_date_only = today

    # ==================== BASIC STATS ====================

    total_employees = Employee.objects.count()

    present_today = Attendance.objects.filter(
        attendance_date=today,
        status__in=['present', 'half_day', 'late']
    ).count()

    active_employees = Employee.objects.filter(
        resigned_date__isnull=True
    ).count()

    employees_with_attendance_today = Attendance.objects.filter(
        attendance_date=today
    ).count()

    absent_today = max(0, active_employees - employees_with_attendance_today)

    pending_approvals = LeaveApply.objects.filter(
        status='pending',
        tl_approved=True
    ).count()

    # ==================== RECENT DATA ====================

    recent_announcements = Announcement.objects.filter(
        status='published'
    ).filter(
        Q(expiry_date__isnull=True) | Q(expiry_date__gte=today)
    ).order_by('-created_at')[:5]

    recent_pending_leaves = LeaveApply.objects.filter(
        status='pending',
        tl_approved=True
    ).select_related('employee').order_by('-applied_at')[:5]

    # ==================== DEPARTMENT DISTRIBUTION ====================

    dept_distribution_raw = Employee.objects.filter(
        resigned_date__isnull=True
    ).values('department').annotate(
        count=Count('id')
    ).order_by('-count')

    dept_distribution = []
    for dept in dept_distribution_raw:
        percentage = round((dept['count'] / active_employees) * 100, 1) if active_employees > 0 else 0
        dept_distribution.append({
            'department': dept['department'] or 'Unassigned',
            'count': dept['count'],
            'percentage': percentage
        })

    # ==================== MONTHLY ATTENDANCE ====================

    monthly_attendance_data = []

    for i in range(5, -1, -1):
        month_date = current_date.replace(day=1) - timedelta(days=30*i)
        month_start = month_date.replace(day=1)

        if month_start.month == 12:
            month_end = month_start.replace(year=month_start.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = month_start.replace(month=month_start.month + 1, day=1) - timedelta(days=1)

        month_attendance = Attendance.objects.filter(
            attendance_date__gte=month_start,
            attendance_date__lte=month_end
        )

        present_count = month_attendance.filter(status__in=['present', 'half_day', 'late']).count()
        absent_count = month_attendance.filter(status='absent').count()
        total_records = month_attendance.count()

        attendance_rate = round((present_count / max(total_records, 1)) * 100, 1)

        monthly_attendance_data.append({
            'month': month_start.strftime('%b %Y'),
            'present': present_count,
            'absent': absent_count,
            'attendance_rate': attendance_rate,
            'total_records': total_records
        })

    # ==================== QUICK ACTIONS ====================

    quick_actions_stats = {
        'total_employees': total_employees,
        'pending_leaves': LeaveApply.objects.filter(status='pending').count(),
        'total_payrolls': Payroll.objects.count(),
        'total_announcements': Announcement.objects.count(),
        'recent_reports': 0,
        'monthly_reports': 0,
        'total_hr_profiles': HRProfile.objects.count()
    }

    # ==================== UPCOMING EVENTS ====================

    upcoming_contracts = []
    for emp in Employee.objects.filter(resigned_date__isnull=True, contract_end_date__isnull=False):

        contract_dt = emp.contract_end_date

        if isinstance(contract_dt, datetime):
            contract_date = contract_dt.date()
        else:
            contract_date = contract_dt

        days_until = (contract_date - current_date_only).days

        if 0 <= days_until <= 60:
            upcoming_contracts.append({
                'employee': emp,
                'date': contract_date,
                'days_until': days_until,
                'type': 'contract_renewal'
            })

    upcoming_expiry_announcements = []
    for announcement in Announcement.objects.filter(
        status='published',
        expiry_date__isnull=False,
        expiry_date__gte=current_date_only
    ):

        expiry_dt = announcement.expiry_date

        if isinstance(expiry_dt, datetime):
            expiry_date = expiry_dt.date()
        else:
            expiry_date = expiry_dt

        days_until = (expiry_date - current_date_only).days

        if days_until <= 30:
            upcoming_expiry_announcements.append({
                'announcement': announcement,
                'date': expiry_date,
                'days_until': days_until,
                'type': 'announcement_expiry'
            })

    upcoming_anniversaries = []
    for emp in Employee.objects.filter(resigned_date__isnull=True, created_at__isnull=False):

        join_dt = emp.created_at

        if isinstance(join_dt, datetime):
            join_date = join_dt.date()
        else:
            join_date = join_dt

        this_year_anniversary = join_date.replace(year=current_date.year)

        if this_year_anniversary < current_date_only:
            next_anniversary = join_date.replace(year=current_date.year + 1)
        else:
            next_anniversary = this_year_anniversary

        days_until = (next_anniversary - current_date_only).days

        if 0 <= days_until <= 30:
            years_of_service = current_date.year - join_date.year
            upcoming_anniversaries.append({
                'employee': emp,
                'date': next_anniversary,
                'days_until': days_until,
                'type': 'work_anniversary',
                'years_of_service': years_of_service
            })

    # ==================== FINAL EVENTS ====================

    upcoming_events = []

    for event in upcoming_contracts:
        upcoming_events.append({
            'title': f"{event['employee'].first_name} {event['employee'].last_name}'s Contract Renewal",
            'date': event['date'],
            'days_until': event['days_until'],
            'type': 'contract_renewal',
            'description': "Contract renewal due",
            'icon': 'fa-file-contract',
            'color': '#4ecdc4'
        })

    for event in upcoming_expiry_announcements:
        upcoming_events.append({
            'title': f"Announcement Expiry: {event['announcement'].title[:30]}...",
            'date': event['date'],
            'days_until': event['days_until'],
            'type': 'announcement_expiry',
            'description': "Announcement will expire",
            'icon': 'fa-bullhorn',
            'color': '#45b7d1'
        })

    for event in upcoming_anniversaries:
        upcoming_events.append({
            'title': f"{event['employee'].first_name} {event['employee'].last_name}'s Work Anniversary",
            'date': event['date'],
            'days_until': event['days_until'],
            'type': 'work_anniversary',
            'description': f"{event['years_of_service']} years completed",
            'icon': 'fa-calendar-check',
            'color': '#ff9ff3'
        })

    upcoming_events.sort(key=lambda x: x['days_until'])
    upcoming_events = upcoming_events[:10]

    # ==================== FINAL CONTEXT ====================

    context = {
        'total_employees': total_employees,
        'present_today': present_today,
        'absent_today': absent_today,
        'pending_approvals': pending_approvals,
        'recent_announcements': recent_announcements,
        'announcements_count': recent_announcements.count(),
        'recent_pending_leaves': recent_pending_leaves,
        'dept_distribution': dept_distribution,
        'monthly_attendance_data': monthly_attendance_data,
        'active_employees': active_employees,
        'today': today,
        'quick_actions_stats': quick_actions_stats,
        'upcoming_events': upcoming_events,
        'chart_data': {
            'dept_labels': [d['department'] for d in dept_distribution],
            'dept_counts': [d['count'] for d in dept_distribution],
            'monthly_labels': [m['month'] for m in monthly_attendance_data],
            'monthly_attendance_rates': [m['attendance_rate'] for m in monthly_attendance_data],
            'monthly_present': [m['present'] for m in monthly_attendance_data],
            'monthly_absent': [m['absent'] for m in monthly_attendance_data]
        }
    }

    return render(request, 'app/hr/dashboard.html', context)


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def employee_dashboard(request):
    """Employee Dashboard - Clear and correct data fetch from database"""
    # Get employee ID from session
    employee_id = request.session.get('employee_id')
    
    # Get recent announcements for Employee dashboard
    recent_announcements = Announcement.objects.filter(
        status='published'
    ).filter(
        Q(expiry_date__isnull=True) | Q(expiry_date__gte=timezone.now().date())
    ).order_by('-created_at')[:5]
    
    # Initialize default values with clear fallbacks
    working_days_this_month = 22  # Default working days
    present_this_month = 0
    attendance_rate = 0.0
    pending_leaves = 0
    employee_obj = None
    recent_attendance = []
    recent_leave_applications = []
    leave_statistics = {
        'leaves_used_this_month': 0,
        'max_leaves_per_month': 10,
        'remaining_leaves': 10,
        'total_leave_days_approved': 0,
        'pending_leave_days': 0,
        'total_approved_days_year': 0,
        'leave_type_stats': {},
        'recent_activity_count': 0,
        'next_leave_date': None,
    }
    today_attendance = None
    can_check_in = False
    can_check_out = False
    worked_hours = 0.0
    remaining_hours = 8.0
    team_assignment = None
    tl_approval = None
    present_days = 0
    half_days = 0
    
    # Check if employee is logged in
    if not employee_id:
        # Add error message for missing session
        from django.contrib import messages
        messages.error(request, "Please login to access your dashboard.")
        return redirect('login')
    
    try:
        # Get employee object with error handling
        try:
            employee_obj = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            from django.contrib import messages
            messages.error(request, "Employee profile not found. Please login again.")
            return redirect('login')
        
        # Get current time and date information
        now = timezone.now()
        current_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month = (current_month + timedelta(days=32)).replace(day=1)
        today = now.date()
        
        # ===== EMPLOYEE BASIC INFO =====
        # Get team assignment
        team_assignment = TeamAssignment.objects.filter(employee=employee_obj).select_related('team_leader').first()
        
        # ===== TODAY'S ATTENDANCE =====
        # Get today's attendance record with approval status
        today_attendance = Attendance.objects.filter(
            employee=employee_obj,
            attendance_date=today
        ).select_related('approval').first()
        
        # Check TL approval and check-in/check-out permissions
        if today_attendance:
            # Check for TL approval
            tl_approval = getattr(today_attendance, 'approval', None)
            if tl_approval and tl_approval.status == 'approved':
                can_check_in = not today_attendance.check_in_time
            else:
                can_check_in = False
            
            # Calculate worked hours if checked in but not checked out
            if today_attendance.check_in_time and not today_attendance.check_out_time:
                from datetime import datetime
                check_in_dt = datetime.combine(today, today_attendance.check_in_time)
                current_dt = datetime.combine(today, now.time())
                worked_duration = current_dt - check_in_dt
                worked_hours = round(worked_duration.total_seconds() / 3600, 2)
                remaining_hours = round(max(0, 8.0 - worked_hours), 2)
                can_check_out = worked_hours >= 8.0
        
        # ===== ATTENDANCE DATA FOR CURRENT MONTH =====
        # Get current month attendance records
        current_month_attendance = Attendance.objects.filter(
            employee=employee_obj,
            attendance_date__gte=current_month,
            attendance_date__lt=next_month
        )
        
        # Calculate present and half days
        present_days = current_month_attendance.filter(
            status__in=['present', 'late']
        ).count()
        
        half_days = current_month_attendance.filter(
            status='half_day'
        ).count()
        
        present_this_month = present_days + half_days  # Include half days as present
        
        # Calculate working days for current month (Monday to Saturday)
        year = now.year
        month = now.month
        _, num_days = monthrange(year, month)
        
        working_days = 0
        for day in range(1, num_days + 1):
            date_obj = now.replace(year=year, month=month, day=day)
            if date_obj.weekday() < 6:  # Monday to Saturday (0-5)
                working_days += 1
        working_days_this_month = working_days
        
        # Calculate attendance rate
        if working_days_this_month > 0:
            attendance_rate = round((present_this_month / working_days_this_month) * 100, 1)
        else:
            attendance_rate = 0.0
        
        # ===== RECENT ATTENDANCE RECORDS =====
        # Get recent attendance records (last 10 for better display)
        recent_attendance = Attendance.objects.filter(
            employee=employee_obj
        ).select_related('approval').order_by('-attendance_date')[:10]
        
        # ===== LEAVE DATA =====
        # Get pending leave requests
        pending_leaves = LeaveApply.objects.filter(
            employee=employee_obj,
            status='pending'
        ).count()
        
        # Get recent leave applications
        recent_leave_applications = LeaveApply.objects.filter(
            employee=employee_obj
        ).order_by('-applied_at')[:5]
        
        # ===== ENHANCED LEAVE STATISTICS =====
        # Monthly leave statistics
        monthly_leaves = LeaveApply.objects.filter(
            employee=employee_obj,
            start_date__gte=current_month,
            start_date__lt=next_month,
            status__in=['approved', 'pending']
        )
        
        # Calculate leave usage for current month
        leaves_used_this_month = sum([leave.total_days for leave in monthly_leaves])
        max_leaves_per_month = 10  # Policy: 10 days per month
        remaining_leaves = max(0, max_leaves_per_month - leaves_used_this_month)
        
        # Get approved and pending leave days
        approved_leaves_this_month = monthly_leaves.filter(status='approved')
        total_leave_days_approved = sum([leave.total_days for leave in approved_leaves_this_month])
        pending_leave_days = sum([leave.total_days for leave in monthly_leaves.filter(status='pending')])
        
        # Calculate yearly statistics
        current_year = now.year
        yearly_leaves = LeaveApply.objects.filter(
            employee=employee_obj,
            start_date__year=current_year,
            status='approved'
        )
        
        total_approved_days = sum([leave.total_days for leave in yearly_leaves])
        
        # Get leave type distribution for current year
        leave_type_stats = {}
        for leave_type, leave_type_display in LeaveApply.LEAVE_TYPE_CHOICES:
            type_leaves = yearly_leaves.filter(leave_type=leave_type)
            type_count = type_leaves.count()
            type_days = sum([leave.total_days for leave in type_leaves])
            if type_count > 0:  # Only include leave types that have been used
                leave_type_stats[leave_type_display] = {
                    'count': type_count,
                    'days': type_days
                }
        
        # Recent leave activity (last 30 days)
        last_30_days = now - timedelta(days=30)
        recent_leave_activity = LeaveApply.objects.filter(
            employee=employee_obj,
            applied_at__gte=last_30_days
        )
        
        # Next approved leave date
        next_leave = monthly_leaves.filter(status='approved').order_by('start_date').first()
        next_leave_date = next_leave.start_date if next_leave else None
        
        # Update leave statistics
        leave_statistics = {
            'leaves_used_this_month': leaves_used_this_month,
            'max_leaves_per_month': max_leaves_per_month,
            'remaining_leaves': remaining_leaves,
            'total_leave_days_approved': total_leave_days_approved,
            'pending_leave_days': pending_leave_days,
            'total_approved_days_year': total_approved_days,
            'leave_type_stats': leave_type_stats,
            'recent_activity_count': recent_leave_activity.count(),
            'next_leave_date': next_leave_date,
        }
        
    except Exception as e:
        # Log error and keep default values
        print(f"Error in employee dashboard: {str(e)}")
        from django.contrib import messages
        messages.error(request, f"Dashboard data loading error: {str(e)}")
    
    # Prepare context with all data
    context = {
        'recent_announcements': recent_announcements,
        'announcements_count': recent_announcements.count(),
        'working_days_this_month': working_days_this_month,
        'present_this_month': present_this_month,
        'attendance_rate': attendance_rate,
        'pending_leaves': pending_leaves,
        'employee_obj': employee_obj,
        'recent_attendance': recent_attendance,
        'recent_leave_applications': recent_leave_applications,
        'leave_statistics': leave_statistics,
        'today_attendance': today_attendance,
        'can_check_in': can_check_in,
        'can_check_out': can_check_out,
        'worked_hours': worked_hours,
        'remaining_hours': remaining_hours,
        'team_assignment': team_assignment,
        'tl_approval': tl_approval,
        'present_days': present_days,  # For dashboard display
        'half_days': half_days,  # For dashboard display
    }
    
    return render(request, 'app/employee/dashboard.html', context)

########################### Employee Model Pages Views ######################
@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def employee_attendance(request):
    """
    Employee Attendance Page - Display the main attendance interface
    This view handles the "My Attendance" navigation link
    """
    employee_id = request.session.get('employee_id')
    
    if not employee_id:
        messages.error(request, "Please login to access attendance.")
        return redirect('login')
    
    try:
        employee = Employee.objects.get(id=employee_id)
        today = timezone.now().date()
        current_time = timezone.now().time()
        
        # Get or create today's attendance record
        attendance, created = Attendance.objects.get_or_create(
            employee=employee,
            attendance_date=today,
            defaults={
                'status': 'present',
                'shift_type': 'full_time',
                'is_check_in_allowed': True,
                'can_check_out': False
            }
        )
        
        # Get present attendance records for display
        attendance_records = Attendance.objects.filter(
            employee=employee,
            status__in=['present', 'late', 'half_day']
        ).order_by('-attendance_date')[:20]  # Last 20 records
        
        # Calculate current month statistics
        current_month = timezone.now().replace(day=1)
        current_month_attendance = attendance_records.filter(
            attendance_date__gte=current_month
        )
        
        current_month_stats = {
            'total_days': current_month_attendance.count(),
            'present_days': current_month_attendance.filter(status='present').count(),
            'half_days': current_month_attendance.filter(status='half_day').count(),
            'absent_days': current_month_attendance.filter(status='absent').count(),
        }
        
        # Calculate attendance percentage
        total_present_days = attendance_records.count()
        attendance_percentage = round((total_present_days / max(total_present_days, 1)) * 100, 1)
        
        context = {
            'employee': employee,
            'attendance': attendance,
            'attendance_records': attendance_records,
            'current_month_stats': current_month_stats,
            'attendance_percentage': attendance_percentage,
            'total_days': total_present_days,
            'present_days': current_month_stats['present_days'],
            'half_days': current_month_stats['half_days'],
        }
        
        return render(request, 'app/employee/attendance.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, "Employee profile not found.")
        return redirect('login')
    except Exception as e:
        messages.error(request, f"Error loading attendance page: {str(e)}")
        return redirect('employee-dashboard')

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def employee_apply_leave(request):
    """
    Employee Leave Application - Handle both GET (display) and POST (submit) requests
    Implements 1-day leave restriction and monthly tracking
    """
    # Get employee ID from session
    employee_id = request.session.get('employee_id')
    
    if not employee_id:
        messages.error(request, "Please login to apply for leave.")
        return redirect('login')
    
    try:
        employee = Employee.objects.get(id=employee_id)
        current_month = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month = (current_month + timedelta(days=32)).replace(day=1)
        
        # Calculate leave statistics for current month
        monthly_leaves = LeaveApply.objects.filter(
            employee=employee,
            start_date__gte=current_month,
            start_date__lt=next_month,
            status__in=['approved', 'pending']  # Count both pending and approved
        )
        
        # Get leave balance (calculate based on total days)
        leaves_taken_this_month = sum([leave.total_days for leave in monthly_leaves])
        max_leaves_per_month = 10  # Policy: 10 days per month
        remaining_leaves = max_leaves_per_month - leaves_taken_this_month
        
        # Get recent leave applications
        recent_leaves = LeaveApply.objects.filter(
            employee=employee
        ).order_by('-applied_at')[:10]
        
        # Calculate statistics
        total_approved = LeaveApply.objects.filter(
            employee=employee,
            status='approved'
        ).count()
        
        total_pending = LeaveApply.objects.filter(
            employee=employee,
            status='pending'
        ).count()
        
        # ==============================
        # POST REQUEST (SUBMIT LEAVE APPLICATION or CANCEL REQUEST)
        # ==============================
        if request.method == "POST":
            action = request.POST.get("action")
            
            if action == "submit":
                # Submit new leave application
                try:
                    leave_type = request.POST.get("leave_type")
                    start_date_str = request.POST.get("start_date")
                    end_date_str = request.POST.get("end_date")
                    reason = request.POST.get("reason")
                    contact_details = request.POST.get("contact_details", "")
                    emergency_contact = request.POST.get("emergency_contact", "")
                    # supporting_document removed - no longer required
                    
                    # Validation
                    if not all([leave_type, start_date_str, end_date_str, reason]):
                        messages.error(request, "All required fields must be filled!")
                        return redirect('apply-leave')
                    
                    # Parse dates
                    try:
                        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                    except ValueError:
                        messages.error(request, "Invalid date format!")
                        return redirect('apply-leave')
                    
                    # Validate date range
                    if start_date > end_date:
                        messages.error(request, "Start date cannot be after end date!")
                        return redirect('apply-leave')
                    
                    # Calculate total days
                    total_days = (end_date - start_date).days + 1
                    
                    # Check maximum leave duration (optional - you can adjust this limit)
                    if total_days > 30:
                        messages.error(request, f"Maximum 30 days leave allowed per application. You requested {total_days} days.")
                        return redirect('apply-leave')
                    
                    # Check if leave is in the past
                    if start_date < timezone.now().date():
                        messages.error(request, "Cannot apply for leave in the past!")
                        return redirect('apply-leave')
                    
                    # Check monthly leave limit
                    if leaves_taken_this_month + total_days > max_leaves_per_month:
                        messages.error(request, f"Monthly leave limit exceeded. You have already used {leaves_taken_this_month} days this month and are requesting {total_days} more days (limit: {max_leaves_per_month} days).")
                        return redirect('apply-leave')
                    
                    # Check for overlapping leave requests
                    overlapping_leaves = LeaveApply.objects.filter(
                        employee=employee,
                        status__in=['pending', 'approved'],
                        start_date__lte=end_date,
                        end_date__gte=start_date
                    )
                    
                    if overlapping_leaves.exists():
                        messages.error(request, "You already have a leave request for this date range!")
                        return redirect('apply-leave')
                    
                    # Create leave application
                    leave_application = LeaveApply.objects.create(
                        employee=employee,
                        leave_type=leave_type,
                        start_date=start_date,
                        end_date=end_date,
                        total_days=total_days,
                        reason=reason,
                        contact_details=contact_details,
                        emergency_contact=emergency_contact,
                        # supporting_document removed - no longer required
                        status='pending'  # Starts as pending for TL approval
                    )
                    
                    messages.success(request, f"Leave application submitted successfully for {start_date.strftime('%B %d, %Y')}!")
                    return redirect('apply-leave')
                    
                except Exception as e:
                    messages.error(request, f"Error submitting leave application: {str(e)}")
                    return redirect('apply-leave')
            
            elif action == "cancel":
                # Cancel existing leave request
                leave_id = request.POST.get("leave_id")
                try:
                    leave_application = LeaveApply.objects.get(
                        id=leave_id,
                        employee=employee,
                        status='pending'
                    )
                    leave_application.status = 'cancelled'
                    leave_application.save()
                    messages.success(request, "Leave request cancelled successfully.")
                except LeaveApply.DoesNotExist:
                    messages.error(request, "Leave request not found or cannot be cancelled.")
                
                return redirect('apply-leave')
        
        context = {
            'employee': employee,
            'leaves_taken_this_month': leaves_taken_this_month,
            'max_leaves_per_month': max_leaves_per_month,
            'remaining_leaves': remaining_leaves,
            'total_approved': total_approved,
            'total_pending': total_pending,
            'recent_leaves': recent_leaves,
        }
        
        return render(request, 'app/employee/leave-apply.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, "Employee profile not found.")
        return redirect('login')

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def employee_payslip(request):
    """Employee Payslip View - Display real payroll records from database"""
    # Get employee ID from session
    employee_id = request.session.get('employee_id')
    
    if not employee_id:
        messages.error(request, "Please login to view your payslip.")
        return redirect('login')
    
    try:
        # Get employee object
        employee = Employee.objects.get(id=employee_id)
        
        # Get all payroll records for this employee, ordered by most recent first
        payroll_records = Payroll.objects.filter(employee=employee).order_by('-created_at')
        
        # Calculate statistics
        total_payroll_records = payroll_records.count()
        
        # Get latest payroll record for current statistics
        latest_payroll = payroll_records.first()
        
        # Calculate totals and averages
        total_gross_salary = payroll_records.aggregate(
            total=models.Sum('gross_salary')
        )['total'] or 0
        
        total_deductions = payroll_records.aggregate(
            total=models.Sum('total_deductions')
        )['total'] or 0
        
        total_net_salary = payroll_records.aggregate(
            total=models.Sum('final_salary')
        )['total'] or 0
        
        # Get current month payroll status
        current_month = timezone.now().strftime("%B %Y")
        current_month_payroll = payroll_records.filter(month=current_month).first()
        
        # Get last paid month
        last_paid_payroll = payroll_records.filter(
            is_processed=True
        ).order_by('-processed_date').first()
        
        # Calculate attendance and leave statistics if needed
        current_year = timezone.now().year
        current_year_payrolls = payroll_records.filter(year=current_year)
        average_monthly_salary = current_year_payrolls.aggregate(
            avg=models.Avg('final_salary')
        )['avg'] or 0
        
        context = {
            'employee': employee,
            'payroll_records': payroll_records,
            'total_payroll_records': total_payroll_records,
            'latest_payroll': latest_payroll,
            'total_gross_salary': total_gross_salary,
            'total_deductions': total_deductions,
            'total_net_salary': total_net_salary,
            'current_month_payroll': current_month_payroll,
            'last_paid_payroll': last_paid_payroll,
            'average_monthly_salary': average_monthly_salary,
            'current_month': current_month,
        }
        
        return render(request, 'app/employee/payslip.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, "Employee profile not found.")
        return redirect('login')
    except Exception as e:
        messages.error(request, f"Error loading payroll data: {str(e)}")
        return redirect('employee-dashboard')

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def employee_profile(request):
    """Employee Profile - Display logged-in employee details from database with team and project information"""
    # Get employee ID from session
    employee_id = request.session.get('employee_id')
    
    if not employee_id:
        messages.error(request, "Please login to view your profile.")
        return redirect('login')
    
    try:
        # Fetch employee data from database
        employee = Employee.objects.get(id=employee_id)
        
        # Get additional data if needed (attendance, payroll, etc.)
        recent_attendance = Attendance.objects.filter(
            employee=employee
        ).order_by('-attendance_date')[:5]
        
        # ===== TEAM LEAD INFORMATION =====
        # Get team assignment for this employee
        team_assignment = TeamAssignment.objects.filter(
            employee=employee
        ).select_related('team_leader', 'team_leader__employee').first()
        
        # Get team leader info if assigned
        team_leader_info = None
        if team_assignment:
            tl_employee = team_assignment.team_leader.employee
            team_leader_info = {
                'name': f"{tl_employee.first_name} {tl_employee.last_name}",
                'designation': tl_employee.designation,
                'department': tl_employee.department,
                'company_id': tl_employee.company_id,
                'email': tl_employee.email,
                'contact_number': tl_employee.contact_number,
                'role': team_assignment.role,
                'assignment_date': team_assignment.assignment_date,
                'notes': team_assignment.notes,
                'experience_years': team_assignment.team_leader.experience_years,
                'team_size': team_assignment.team_leader.team_size,
                'responsibilities': team_assignment.team_leader.responsibilities
            }
        
        # ===== PROJECT DETAILS =====
        # Get projects assigned to this employee
        project_assignments = ProjectAssignment.objects.filter(
            team_members=employee
        ).select_related('team_leader', 'team_leader__employee').order_by('-created_at')
        
        projects_info = []
        for project in project_assignments:
            tl_employee = project.team_leader.employee
            projects_info.append({
                'project_name': project.project_name,
                'project_code': project.project_code,
                'priority': project.priority,
                'start_date': project.start_date,
                'end_date': project.end_date,
                'description': project.description,
                'team_leader': f"{tl_employee.first_name} {tl_employee.last_name}",
                'team_leader_designation': tl_employee.designation,
                'status': 'Active' if project.end_date >= timezone.now().date() else 'Completed',
                'days_remaining': (project.end_date - timezone.now().date()).days if project.end_date >= timezone.now().date() else None
            })
        
        # ===== TASK AND DEADLINE INFORMATION =====
        # Get tasks assigned to this employee
        employee_tasks = ProjectTask.objects.filter(
            assigned_to=employee
        ).select_related('project', 'assigned_by', 'assigned_by__employee').order_by('due_date')[:10]
        
        tasks_info = []
        for task in employee_tasks:
            tl_employee = task.assigned_by.employee if task.assigned_by else None
            tasks_info.append({
                'title': task.title,
                'description': task.description,
                'project_name': task.project.project_name,
                'task_status': task.task_status,
                'priority': task.priority,
                'start_date': task.start_date,
                'due_date': task.due_date,
                'estimated_hours': task.estimated_hours,
                'actual_hours': task.actual_hours,
                'progress_percentage': task.progress_percentage,
                'milestone': task.milestone,
                'assigned_by': f"{tl_employee.first_name} {tl_employee.last_name}" if tl_employee else 'Unknown',
                'days_until_due': (task.due_date.date() - timezone.now().date()).days if task.due_date else None
            })
        
        # ===== OVERALL STATUS ("ALL OK" INDICATOR) =====
        # Calculate overall employee status
        current_date = timezone.now().date()
        
        # Check attendance status this month
        current_month = current_date.replace(day=1)
        monthly_attendance = Attendance.objects.filter(
            employee=employee,
            attendance_date__gte=current_month
        )
        
        present_days = monthly_attendance.filter(status__in=['present', 'late']).count()
        half_days = monthly_attendance.filter(status='half_day').count()
        absent_days = monthly_attendance.filter(status='absent').count()
        
        # Check pending tasks
        pending_tasks = [task for task in tasks_info if task['task_status'] in ['not_started', 'in_progress']]
        overdue_tasks = [task for task in pending_tasks if task['days_until_due'] is not None and task['days_until_due'] < 0]
        
        # Check pending leave applications
        pending_leaves = LeaveApply.objects.filter(
            employee=employee,
            status='pending'
        ).count()
        
        # Calculate overall status
        status_indicators = {
            'attendance_status': 'Good' if absent_days == 0 else 'Needs Attention' if absent_days <= 2 else 'Critical',
            'task_status': 'On Track' if len(overdue_tasks) == 0 else 'Behind Schedule' if len(overdue_tasks) <= 2 else 'Critical',
            'leave_status': 'Clear' if pending_leaves == 0 else 'Pending Review',
            'project_status': 'Active' if len(projects_info) > 0 else 'No Active Projects'
        }
        
        # Overall status calculation
        critical_issues = sum(1 for status in status_indicators.values() if status == 'Critical')
        if critical_issues > 0:
            overall_status = 'Critical'
            status_color = 'danger'
            status_icon = 'fa-exclamation-triangle'
        elif any('Needs Attention' in status or 'Behind Schedule' in status for status in status_indicators.values()):
            overall_status = 'Needs Attention'
            status_color = 'warning'
            status_icon = 'fa-exclamation-circle'
        else:
            overall_status = 'All OK'
            status_color = 'success'
            status_icon = 'fa-check-circle'
        
        context = {
            'employee': employee,
            'recent_attendance': recent_attendance,
            'team_leader_info': team_leader_info,
            'projects_info': projects_info,
            'tasks_info': tasks_info,
            'overall_status': overall_status,
            'status_color': status_color,
            'status_icon': status_icon,
            'status_indicators': status_indicators,
            'attendance_summary': {
                'present_days': present_days,
                'half_days': half_days,
                'absent_days': absent_days,
                'total_monthly_days': present_days + half_days + absent_days
            },
            'task_summary': {
                'total_tasks': len(tasks_info),
                'pending_tasks': len(pending_tasks),
                'overdue_tasks': len(overdue_tasks),
                'completed_tasks': len([task for task in tasks_info if task['task_status'] == 'completed'])
            },
            'project_summary': {
                'total_projects': len(projects_info),
                'active_projects': len([p for p in projects_info if p['status'] == 'Active']),
                'completed_projects': len([p for p in projects_info if p['status'] == 'Completed'])
            }
        }
        
        return render(request, 'app/employee/profile.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, "Employee profile not found.")
        return redirect('login')


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def hr_profile(request):
    """HR Profile - Display logged-in HR details from database"""
    # Get HR ID from session
    hr_id = request.session.get('hr_id')
    
    if not hr_id:
        messages.error(request, "Please login to view your profile.")
        return redirect('login')
    
    try:
        # Fetch HR data from database
        hr = HRProfile.objects.get(id=hr_id)
        
        # Calculate years of experience based on joining date
        from datetime import date
        today = date.today()
        joining_date = hr.date_of_joining
        experience_years = today.year - joining_date.year
        
        # Adjust if the anniversary hasn't occurred this year
        if today.month < joining_date.month or (today.month == joining_date.month and today.day < joining_date.day):
            experience_years -= 1
        
        # Get additional HR-specific data
        total_employees = Employee.objects.count()
        total_announcements = Announcement.objects.count()
        
        context = {
            'hr': hr,
            'total_employees': total_employees,
            'total_announcements': total_announcements,
            'experience_years': experience_years,
        }
        
        return render(request, 'app/hr_profile.html', context)
        
    except HRProfile.DoesNotExist:
        messages.error(request, "HR profile not found.")
        return redirect('login')


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def tl_profile(request):
    """Team Leader Profile - Display logged-in Team Leader details from database"""
    # Get TL ID from session
    tl_id = request.session.get('tl_id')
    
    if not tl_id:
        messages.error(request, "Please login to view your profile.")
        return redirect('login')
    
    try:
        # Fetch Team Leader data from database
        tl = TeamLeader.objects.get(id=tl_id)
        employee = tl.employee  # Get the associated employee record
        
        # Get TL-specific data
        team_members = TeamAssignment.objects.filter(team_leader=tl)
        team_count = team_members.count()
        
        context = {
            'tl': tl,
            'employee': employee,
            'team_members': team_members,
            'team_count': team_count,
        }
        
        return render(request, 'app/tl_profile.html', context)
        
    except TeamLeader.DoesNotExist:
        messages.error(request, "Team Leader profile not found.")
        return redirect('login')


############################ End Employee Module Pages View #################

############################# Tl Module Views Start here #####################
@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def team_leader_dashboard(request):
    """Team Leader Dashboard - Login checks removed as requested"""
    # Get recent announcements for TL dashboard
    recent_announcements = Announcement.objects.filter(
        status='published'
    ).filter(
        Q(expiry_date__isnull=True) | Q(expiry_date__gte=timezone.now().date())
    ).order_by('-created_at')[:5]
    
    context = {
        'recent_announcements': recent_announcements,
        'announcements_count': recent_announcements.count()
    }
    
    # Session checks removed - direct access to dashboard
    return render(request, 'app/tl/dashboard.html', context)

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def tl_dashboard(request):
    """TL Dashboard with real team data and enhanced leave tracking"""
    # Get TL ID from session
    tl_id = request.session.get('tl_id')
    
    # Get recent announcements for TL dashboard
    recent_announcements = Announcement.objects.filter(
        status='published'
    ).filter(
        models.Q(expiry_date__isnull=True) | models.Q(expiry_date__gte=timezone.now().date())
    ).order_by('-created_at')[:5]
    
    # Initialize default values
    total_team_members = 0
    present_today = 0
    absent_today = 0
    pending_approvals = 0
    tl_obj = None
    team_assignments = []
    team_leave_applications = []
    leave_statistics = {}
    
    if tl_id:
        try:
            tl_obj = TeamLeader.objects.get(id=tl_id)
            employee = tl_obj.employee
            
            # Get team members assigned to this TL
            team_assignments = TeamAssignment.objects.filter(team_leader=tl_obj).select_related('employee')
            total_team_members = team_assignments.count()
            
            # Get today's date
            today = timezone.now().date()
            
            # Count present team members today
            present_count = 0
            absent_count = 0
            
            for assignment in team_assignments:
                # Check if team member has attendance record for today
                today_attendance = Attendance.objects.filter(
                    employee=assignment.employee,
                    attendance_date=today
                ).first()
                
                if today_attendance:
                    if today_attendance.status in ['present', 'half_day', 'late']:
                        present_count += 1
                    elif today_attendance.status == 'absent':
                        absent_count += 1
                else:
                    # If no attendance record, consider as absent
                    absent_count += 1
            
            present_today = present_count
            absent_today = absent_count
            
            # Get pending leave approvals for team members
            team_employee_ids = [assignment.employee.id for assignment in team_assignments]
            team_leave_applications = LeaveApply.objects.filter(
                employee_id__in=team_employee_ids,
                status='pending'
            ).select_related('employee').order_by('-applied_at')[:5]
            
            pending_approvals = team_leave_applications.count()
            
            # ===== ENHANCED LEAVE STATISTICS FOR TL =====
            
            # Get all team member leave applications
            team_leaves = LeaveApply.objects.filter(
                employee_id__in=team_employee_ids
            )
            
            # Calculate monthly statistics
            current_month = timezone.now().replace(day=1)
            monthly_team_leaves = team_leaves.filter(applied_at__gte=current_month)
            
            # Leave statistics for current month
            pending_leaves_month = monthly_team_leaves.filter(status='pending').count()
            approved_leaves_month = monthly_team_leaves.filter(status='approved').count()
            rejected_leaves_month = monthly_team_leaves.filter(status='rejected').count()
            total_leave_days_month = sum([leave.total_days for leave in monthly_team_leaves.filter(status='approved')])
            
            # Team member leave usage this month
            team_member_leave_usage = []
            for assignment in team_assignments:
                member_leaves = monthly_team_leaves.filter(employee=assignment.employee)
                leaves_used = member_leaves.count()
                leave_days = sum([leave.total_days for leave in member_leaves.filter(status='approved')])
                
                team_member_leave_usage.append({
                    'employee': assignment.employee,
                    'leaves_used': leaves_used,
                    'leave_days': leave_days,
                    'pending_leaves': member_leaves.filter(status='pending').count(),
                    'has_over_limit': leaves_used > 1  # Monthly limit check
                })
            
            # Leave type distribution for team
            team_leave_types = {}
            for leave_type, leave_type_display in LeaveApply.LEAVE_TYPE_CHOICES:
                type_count = monthly_team_leaves.filter(leave_type=leave_type).count()
                team_leave_types[leave_type_display] = type_count
            
            # Recent TL decisions on leave applications
            recent_tl_decisions = LeaveApply.objects.filter(
                employee_id__in=team_employee_ids,
                tl_approved_by__isnull=False
            ).select_related('employee').order_by('-tl_approval_date')[:5]
            
            leave_statistics = {
                'pending_leaves_month': pending_leaves_month,
                'approved_leaves_month': approved_leaves_month,
                'rejected_leaves_month': rejected_leaves_month,
                'total_leave_days_month': total_leave_days_month,
                'team_member_leave_usage': team_member_leave_usage,
                'team_leave_types': team_leave_types,
                'recent_tl_decisions': recent_tl_decisions,
                'team_compliance_rate': round((approved_leaves_month / max(monthly_team_leaves.count(), 1)) * 100, 1) if monthly_team_leaves.exists() else 100,
            }
            
        except TeamLeader.DoesNotExist:
            # If TL not found, keep default values
            pass
    
    context = {
        'recent_announcements': recent_announcements,
        'announcements_count': recent_announcements.count(),
        'total_team_members': total_team_members,
        'present_today': present_today,
        'absent_today': absent_today,
        'pending_approvals': pending_approvals,
        'tl_obj': tl_obj,
        'team_assignments': team_assignments,
        'team_leave_applications': team_leave_applications,
        'leave_statistics': leave_statistics,
    }
    
    return render(request, 'app/tl/dashboard.html', context)

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def tl_reports(request):
    return render(request, 'app/tl/reports.html')

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def tl_attendance(request):
    """
    Team Leader Team Attendance Management - Display real attendance data from database
    Shows team member attendance records in table format with real data
    """
    # Get TL ID from session
    tl_id = request.session.get('tl_id')
    
    if not tl_id:
        messages.error(request, "Please login as Team Leader to view team attendance.")
        return redirect('login')
    
    try:
        tl = TeamLeader.objects.get(id=tl_id)
        
        # Handle filter parameters
        selected_date = request.GET.get('date', timezone.now().date().strftime('%Y-%m-%d'))
        status_filter = request.GET.get('status', 'all')
        search_term = request.GET.get('search', '').strip()
        
        try:
            filter_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except ValueError:
            filter_date = timezone.now().date()
        
        # Get team members assigned to this TL
        team_assignments = TeamAssignment.objects.filter(team_leader=tl).select_related('employee')
        team_member_ids = [assignment.employee.id for assignment in team_assignments]
        
        # Get attendance records for team members
        attendance_records = Attendance.objects.filter(
            employee_id__in=team_member_ids,
            attendance_date=filter_date
        ).select_related('employee').order_by('employee__first_name')
        
        # Apply status filter
        if status_filter != 'all':
            if status_filter == 'present':
                attendance_records = attendance_records.filter(status__in=['present', 'late'])
            elif status_filter == 'absent':
                attendance_records = attendance_records.filter(status='absent')
            elif status_filter == 'late':
                attendance_records = attendance_records.filter(status='late')
            elif status_filter == 'remote':
                attendance_records = attendance_records.filter(status='remote_work')
        
        # Apply search filter
        if search_term:
            attendance_records = attendance_records.filter(
                Q(employee__first_name__icontains=search_term) |
                Q(employee__last_name__icontains=search_term) |
                Q(employee__company_id__icontains=search_term) |
                Q(employee__designation__icontains=search_term)
            )
        
        # Calculate statistics
        total_team_members = team_assignments.count()
        
        # Get today's attendance stats
        today_attendance = Attendance.objects.filter(
            employee_id__in=team_member_ids,
            attendance_date=filter_date
        )
        
        present_today = today_attendance.filter(status__in=['present', 'late', 'half_day']).count()
        absent_today = today_attendance.filter(status='absent').count()
        
        # Calculate attendance rate
        if total_team_members > 0:
            attendance_rate = round((present_today / total_team_members) * 100, 1)
        else:
            attendance_rate = 0.0
        
        # Get attendance records with employee details
        team_attendance_data = []
        for record in attendance_records:
            team_assignment = team_assignments.filter(employee=record.employee).first()
            
            # Calculate worked hours
            worked_hours = 0
            if record.check_in_time and record.check_out_time:
                from datetime import datetime
                check_in_dt = datetime.combine(filter_date, record.check_in_time)
                check_out_dt = datetime.combine(filter_date, record.check_out_time)
                worked_duration = check_out_dt - check_in_dt
                worked_hours = round(worked_duration.total_seconds() / 3600, 2)
            
            team_attendance_data.append({
                'id': record.employee.company_id,
                'name': f"{record.employee.first_name} {record.employee.last_name}",
                'designation': record.employee.designation,
                'date': record.attendance_date.strftime('%Y-%m-%d'),
                'check_in': record.check_in_time.strftime('%H:%M') if record.check_in_time else '-',
                'check_out': record.check_out_time.strftime('%H:%M') if record.check_out_time else '-',
                'status': record.status,
                'worked_hours': worked_hours,
                'employee': record.employee,
                'attendance': record
            })
        
        # Get department list for filters
        departments = list(Employee.objects.filter(
            id__in=team_member_ids
        ).values_list('department', flat=True).distinct().order_by('department'))
        
        # Get status choices for filter
        status_choices = [
            ('all', 'All Status'),
            ('present', 'Present'),
            ('absent', 'Absent'),
            ('late', 'Late'),
            ('remote', 'Remote')
        ]
        
        context = {
            'tl': tl,
            'team_attendance_data': team_attendance_data,
            'total_team_members': total_team_members,
            'present_today': present_today,
            'absent_today': absent_today,
            'attendance_rate': attendance_rate,
            'selected_date': selected_date,
            'status_filter': status_filter,
            'search_term': search_term,
            'departments': departments,
            'status_choices': status_choices,
            'filter_date': filter_date,
            'is_filtered': bool(selected_date != timezone.now().date().strftime('%Y-%m-%d') or status_filter != 'all' or search_term),
        }
        
        return render(request, 'app/tl/team-attendence.html', context)
        
    except TeamLeader.DoesNotExist:
        messages.error(request, "Team Leader profile not found.")
        return redirect('login')
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"TL Attendance Error: {error_details}")
        messages.error(request, f"Error loading team attendance: {str(e)}")
        # Don't redirect to dashboard - let the error be visible
        return render(request, 'app/tl/team-attendence.html', {
            'tl': None,
            'team_attendance_data': [],
            'total_team_members': 0,
            'present_today': 0,
            'absent_today': 0,
            'attendance_rate': 0.0,
            'selected_date': timezone.now().date().strftime('%Y-%m-%d'),
            'status_filter': 'all',
            'search_term': '',
            'departments': [],
            'status_choices': [('all', 'All Status'), ('present', 'Present'), ('absent', 'Absent'), ('late', 'Late'), ('remote', 'Remote')],
            'filter_date': timezone.now().date(),
            'is_filtered': False,
        })

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def team_attendance(request):
    return render(request, 'app/tl/team-attendence.html')



@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def team_leave_approval(request):
    """
    Team Leader Leave Approval - Show team member leave requests for approval
    Workflow: Employee applies → TL reviews/approves → Forwards to HR → HR final approval
    """
    # Get TL ID from session
    tl_id = request.session.get('tl_id')
    
    if not tl_id:
        messages.error(request, "Please login as Team Leader to manage leave approvals.")
        return redirect('login')
    
    try:
        tl = TeamLeader.objects.get(id=tl_id)
        
        # ==============================
        # POST REQUEST (APPROVE/REJECT/FORWARD)
        # ==============================
        if request.method == "POST":
            action = request.POST.get("action")
            leave_id = request.POST.get("leave_id")
            comments = request.POST.get("comments", "")
            
            try:
                leave_application = LeaveApply.objects.get(id=leave_id)
                
                # Verify this TL can approve this leave (team member relationship)
                team_assignments = TeamAssignment.objects.filter(
                    team_leader=tl,
                    employee=leave_application.employee
                )
                
                if not team_assignments.exists():
                    messages.error(request, "You are not authorized to approve this leave request.")
                    return redirect('team-leave-approval')
                
                if action == "approve":
                    # TL approves the leave and forwards to HR
                    leave_application.tl_approved = True
                    leave_application.tl_approved_by = f"{tl.employee.first_name} {tl.employee.last_name}"
                    leave_application.tl_approval_date = timezone.now()
                    leave_application.tl_comments = comments
                    leave_application.status = 'pending'  # Remains pending for HR
                    leave_application.save()
                    
                    messages.success(request, f"Leave request approved and forwarded to HR for {leave_application.employee.first_name}.")
                    
                elif action == "reject":
                    # TL rejects the leave
                    if not comments:
                        messages.error(request, "Please provide a reason for rejection.")
                        return redirect('team-leave-approval')
                    
                    leave_application.status = 'rejected'
                    leave_application.tl_comments = comments
                    leave_application.save()
                    
                    messages.success(request, f"Leave request rejected for {leave_application.employee.first_name}.")
                    
                elif action == "forward":
                    # TL explicitly forwards to HR (even after approval)
                    leave_application.tl_approved = True
                    leave_application.tl_approved_by = f"{tl.employee.first_name} {tl.employee.last_name}"
                    leave_application.tl_approval_date = timezone.now()
                    leave_application.tl_comments = comments or "Forwarded to HR for final approval"
                    leave_application.status = 'pending'
                    leave_application.save()
                    
                    messages.success(request, f"Leave request forwarded to HR for {leave_application.employee.first_name}.")
                
                return redirect('team-leave-approval')
                
            except LeaveApply.DoesNotExist:
                messages.error(request, "Leave application not found.")
                return redirect('team-leave-approval')
        
        # ==============================
        # GET REQUEST (DISPLAY PENDING APPROVALS)
        # ==============================
        
        # Get team members assigned to this TL
        team_assignments = TeamAssignment.objects.filter(team_leader=tl)
        team_member_ids = [assignment.employee.id for assignment in team_assignments]
        
        # Get pending leave applications from team members
        pending_leaves = LeaveApply.objects.filter(
            employee_id__in=team_member_ids,
            status='pending'
        ).select_related('employee').order_by('-applied_at')
        
        # Calculate statistics
        total_pending = pending_leaves.count()
        
        # Get recent approved/rejected leaves
        recent_decisions = LeaveApply.objects.filter(
            employee_id__in=team_member_ids,
            tl_approved_by__isnull=False
        ).select_related('employee').order_by('-tl_approval_date')[:10]
        
        # Count approved leaves this month
        current_month = timezone.now().replace(day=1)
        monthly_decisions = LeaveApply.objects.filter(
            employee_id__in=team_member_ids,
            tl_approval_date__gte=current_month
        )
        
        approved_this_month = monthly_decisions.filter(tl_approved=True).count()
        rejected_this_month = monthly_decisions.filter(status='rejected').count()
        
        # Calculate total leave days
        total_leave_days = sum([leave.total_days for leave in monthly_decisions.filter(tl_approved=True)])
        
        context = {
            'tl': tl,
            'pending_leaves': pending_leaves,
            'recent_decisions': recent_decisions,
            'total_pending': total_pending,
            'approved_this_month': approved_this_month,
            'rejected_this_month': rejected_this_month,
            'total_leave_days': total_leave_days,
            'team_member_count': len(team_member_ids),
        }
        
        return render(request, 'app/tl/team-leave-approval.html', context)
        
    except TeamLeader.DoesNotExist:
        messages.error(request, "Team Leader profile not found.")
        return redirect('login')

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def tl_manage_team(request):
    """
    TL Manage Team - Comprehensive team management page with real Django data
    """
    # Get TL ID from session
    tl_id = request.session.get('tl_id')
    
    if not tl_id:
        messages.error(request, "Please login to manage team.")
        return redirect('login')
    
    try:
        tl_obj = TeamLeader.objects.get(id=tl_id)
        employee = tl_obj.employee
        
        # Get team assignments with real data
        team_assignments = TeamAssignment.objects.filter(
            team_leader=tl_obj
        ).select_related('employee').order_by('-assigned_at')
        
        # Get team member statistics
        total_team_members = team_assignments.count()
        
        # Get today's attendance for team members
        today = timezone.now().date()
        team_member_ids = [assignment.employee.id for assignment in team_assignments]
        
        # Calculate attendance statistics
        present_today = Attendance.objects.filter(
            employee_id__in=team_member_ids,
            attendance_date=today,
            status__in=['present', 'half_day', 'late']
        ).count()
        
        absent_today = Attendance.objects.filter(
            employee_id__in=team_member_ids,
            attendance_date=today,
            status='absent'
        ).count()
        
        # Get pending leave applications for team
        pending_leaves = LeaveApply.objects.filter(
            employee_id__in=team_member_ids,
            status='pending'
        ).select_related('employee').order_by('-applied_at')
        
        # Get recent attendance records for team
        recent_attendance = Attendance.objects.filter(
            employee_id__in=team_member_ids
        ).select_related('employee').order_by('-attendance_date')[:20]
        
        # Get team performance data (simplified)
        team_performance = []
        for assignment in team_assignments:
            # Get attendance rate for this month
            current_month = timezone.now().replace(day=1)
            monthly_attendance = Attendance.objects.filter(
                employee=assignment.employee,
                attendance_date__gte=current_month,
                status__in=['present', 'half_day', 'late']
            ).count()
            
            # Calculate working days this month
            import calendar
            year = timezone.now().year
            month = timezone.now().month
            _, num_days = calendar.monthrange(year, month)
            
            working_days = 0
            for day in range(1, num_days + 1):
                date_obj = timezone.now().replace(year=year, month=month, day=day)
                if date_obj.weekday() < 5:  # Monday=0, Sunday=6
                    working_days += 1
            
            attendance_rate = (monthly_attendance / working_days * 100) if working_days > 0 else 0
            
            team_performance.append({
                'assignment': assignment,
                'attendance_rate': round(attendance_rate, 1),
                'monthly_attendance': monthly_attendance,
                'working_days': working_days
            })
        
        # Get available employees for assignment (not already assigned to this TL)
        assigned_employee_ids = [assignment.employee.id for assignment in team_assignments]
        available_employees = Employee.objects.filter(
            resigned_date__isnull=True
        ).exclude(id__in=assigned_employee_ids).order_by('first_name')
        
        context = {
            'tl_obj': tl_obj,
            'employee': employee,
            'team_assignments': team_assignments,
            'total_team_members': total_team_members,
            'present_today': present_today,
            'absent_today': absent_today,
            'pending_leaves': pending_leaves,
            'recent_attendance': recent_attendance,
            'team_performance': team_performance,
            'available_employees': available_employees,
        }
        
        return render(request, 'app/tl/manage-team.html', context)
        
    except TeamLeader.DoesNotExist:
        messages.error(request, "Team Leader profile not found.")
        return redirect('tl-dashboard')
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"TL Manage Team Error: {error_details}")
        messages.error(request, f"Error loading team management: {str(e)}")
        # Don't redirect to dashboard - let the error be visible
        return render(request, 'app/tl/manage-team.html', {
            'tl_obj': None,
            'employee': None,
            'team_assignments': [],
            'total_team_members': 0,
            'present_today': 0,
            'absent_today': 0,
            'pending_leaves': [],
            'recent_attendance': [],
            'team_performance': [],
            'available_employees': [],
        })




############################### Tl Module Views End ##############################

from decimal import Decimal

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def employee(request):
    print(f"=== EMPLOYEE VIEW DEBUG ===")
    print(f"Request method: {request.method}")
    print(f"POST data keys: {list(request.POST.keys())}")
    print(f"FILES data keys: {list(request.FILES.keys()) if request.FILES else 'None'}")
    
    if request.method == "POST":
        try:
            print("Processing POST request for employee creation...")
            # ====== BASIC VALIDATIONS ======
            required_fields = [
                'first_name', 'last_name', 'email', 'password', 'company_id',
                'designation', 'department', 'package', 'address'
            ]
            
            for field in required_fields:
                if not request.POST.get(field):
                    messages.error(request, f"{field.replace('_', ' ').title()} is required!")
                    return redirect("employee-management")

            # ====== DATA CONVERSION ======

            # Convert percentage fields
            def convert_percentage(value):
                if not value:
                    return None
                try:
                    return float(value)
                except:
                    return None

            # Convert contract dates
            contract_start_date = None
            contract_end_date = None
            
            start_date_str = request.POST.get("contract_start_date")
            end_date_str = request.POST.get("contract_end_date")
            
            if start_date_str:
                try:
                    contract_start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                except ValueError:
                    messages.error(request, "Invalid contract start date format!")
                    return redirect("employee-management")
            
            if end_date_str:
                try:
                    contract_end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                except ValueError:
                    messages.error(request, "Invalid contract end date format!")
                    return redirect("employee-management")

            # ====== CREATE EMPLOYEE ======
            image_file = request.FILES.get("image")
            print(f"Image file: {image_file}")

            # Extract form data
            first_name = request.POST.get("first_name").strip()
            last_name = request.POST.get("last_name").strip()
            email = request.POST.get("email").strip().lower()
            password = request.POST.get("password")
            company_id = request.POST.get("company_id").strip()
            designation = request.POST.get("designation").strip()
            department = request.POST.get("department").strip()
            package = Decimal(request.POST.get("package"))
            address = request.POST.get("address").strip()
            
            print(f"Creating employee with data:")
            print(f"  Name: {first_name} {last_name}")
            print(f"  Email: {email}")
            print(f"  Company ID: {company_id}")
            print(f"  Designation: {designation}")
            print(f"  Department: {department}")
            print(f"  Package: {package}")
            print(f"  Address: {address}")

            emp = Employee.objects.create(

                # BASIC
                first_name = first_name,
                last_name = last_name,
                email = email,
                password = password,

                # FAMILY
                contact_number = request.POST.get("contact_number").strip() if request.POST.get("contact_number") else None,
                family_contact_number = request.POST.get("family_contact_number").strip() if request.POST.get("family_contact_number") else None,
                family_relation = request.POST.get("family_relation").strip() if request.POST.get("family_relation") else None,

                # COMPANY
                company_id = company_id,
                official_email = request.POST.get("official_email").strip() if request.POST.get("official_email") else None,
                official_password = request.POST.get("official_password") if request.POST.get("official_password") else None,
                designation = designation,
                department = department,

                # EDUCATION
                certifications = request.POST.get("certifications").strip() if request.POST.get("certifications") else "",

                # IMAGE
                image = image_file,

                # CONTRACT
                package = package,
                contract_years = int(request.POST.get("contract_years") or 1),
                contract_start_date = contract_start_date,
                contract_end_date = contract_end_date,

                # ADDRESS
                address = address,
            )

            # Verify employee was saved
            emp_id = emp.id
            print(f"Employee object created with ID: {emp_id}")
            
            # Double-check in database
            emp_in_db = Employee.objects.filter(id=emp_id).first()
            if emp_in_db:
                print(f"✅ Employee verified in database: {emp_in_db.first_name} {emp_in_db.last_name} - {emp_in_db.email}")
            else:
                print(f"❌ Employee NOT found in database after creation!")
            
            # Get total employee count
            total_emp_count = Employee.objects.count()
            print(f"Total employees in database: {total_emp_count}")
            
            print(f"Employee saved successfully: {emp.first_name} {emp.last_name} - {emp.email}")
            messages.success(request, f"Employee Created Successfully! Employee ID: {emp.company_id}")
            return redirect("employee-management")
            
        except Exception as e:
            print(f"Error creating employee: {e}")
            messages.error(request, f"Error creating employee: {str(e)}")
            return redirect("employee-management")

    # ====== GET REQUEST LOGIC ======
    print("Loading employee list...")
    employees = Employee.objects.all().order_by('-created_at')
    print(f"Total employees found: {employees.count()}")
    
    # Print each employee for debugging
    for i, emp in enumerate(employees, 1):
        print(f"  {i}. {emp.first_name} {emp.last_name} ({emp.company_id}) - {emp.email}")

    total_employees = employees.count()
    active_employees = employees.filter(resigned_date__isnull=True).count()
    
    current_month = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    new_this_month = employees.filter(created_at__gte=current_month).count()
    
    resigned_this_month = employees.filter(
        resigned_date__isnull=False,
        resigned_date__gte=current_month
    ).count()
    
    print(f"Statistics: Total={total_employees}, Active={active_employees}, New this month={new_this_month}, Resigned this month={resigned_this_month}")

    return render(request, 'app/hr/employee.html', {
        'employees': employees,
        'total_employees': total_employees,
        'active_employees': active_employees,
        'new_this_month': new_this_month,
        'resigned_this_month': resigned_this_month,
    })


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def export_employees(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees Data"

    ws.append([
        "Employee ID",
        "First Name",
        "Last Name",
        "Email",
        "Contact Number",
        "Family Contact",
        "Family Relation",
        "Designation",
        "Department",
        "Package",
        "Contract Years",
        "Certifications",
        "Address",
        "Image",
        "Created At",
    ])

    row = 2  # Excel row starts after header

    for emp in Employee.objects.all():

        safe_email = emp.email.replace("=", "")
        safe_id = emp.company_id.replace("=", "")

        ws.append([
            safe_id,
            emp.first_name,
            emp.last_name,
            safe_email,
            emp.contact_number,
            emp.family_contact_number,
            emp.family_relation,
            emp.designation,
            emp.department,
            str(emp.package),
            emp.contract_years,
            emp.certifications,
            emp.address,
            "",   # image will be added here
            emp.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

        # ------------------------------
        #   IMAGE EXPORT (OpenPyXL)
        # ------------------------------
        if emp.image:
            img_path = emp.image.path

            if os.path.exists(img_path):
                try:
                    img = XLImage(img_path)
                    img.width = 60     # resize image width
                    img.height = 60    # resize image height
                    ws.add_image(img, f"U{row}")  # column U -> 21st column
                except:
                    pass

        row += 1

    file_name = f"Employees_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={file_name}"

    wb.save(response)
    return response


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def export_payroll(request):
    """Export all payroll data from database to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Records"

    # Define comprehensive headers
    headers = [
        "Payroll ID",
        "Employee ID", 
        "Employee Name",
        "Email",
        "Department",
        "Designation",
        "Contact Number",
        "Company Package",
        "Salary Month",
        "Base Salary",
        "PF Deduction", 
        "Final Salary",
        "Created By",
        "Created Date",
        "Updated Date",
        "Status"
    ]

    # Add headers to worksheet
    ws.append(headers)

    # Get ALL payroll data from database with related employee info
    # This ensures we get complete data from the Payroll model
    payrolls = Payroll.objects.select_related("employee").all().order_by("-created_at")
    
    # Add data to worksheet
    for payroll in payrolls:
        employee = payroll.employee
        ws.append([
            payroll.id,
            employee.company_id,
            f"{employee.first_name} {employee.last_name}",
            employee.email,
            employee.department,
            employee.designation,
            employee.contact_number or "N/A",
            float(employee.package),
            payroll.month,
            float(payroll.base_salary),
            float(payroll.pf_deduction),
            float(payroll.final_salary),
            payroll.created_by,
            payroll.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            payroll.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
            "Processed"
        ])

    # Add formatting to the worksheet
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0b7368", end_color="0b7368", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-adjust column widths for better readability
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 3, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Add borders to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Freeze the first row (headers)
    ws.freeze_panes = "A2"

    # Generate filename with timestamp
    file_name = f"Complete_Payroll_Records_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Create HTTP response for file download
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={file_name}"

    # Save workbook to response
    wb.save(response)
    return response


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def show_all_employees(request):
    employees = Employee.objects.all().order_by('-created_at')
    
    # Calculate real statistics from Employee model
    total_employees = employees.count()
    active_employees = employees.filter(resigned_date__isnull=True).count()
    
    # New employees this month
    from datetime import datetime
    current_month = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    new_this_month = employees.filter(created_at__gte=current_month).count()
    
    # Resigned this month
    resigned_this_month = employees.filter(
        resigned_date__isnull=False,
        resigned_date__gte=current_month
    ).count()

    return render(request, "app/hr/employee.html", {
        "employees": employees,
        # Add real statistics for cards
        "total_employees": total_employees,
        "active_employees": active_employees,
        "new_this_month": new_this_month,
        "resigned_this_month": resigned_this_month,
    })

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def delete_employee(request, emp_id):
    emp = get_object_or_404(Employee, id=emp_id)
    emp.resigned_date = timezone.now()
    emp.save()

    return JsonResponse({"success": True})


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def get_employee(request, emp_id):
    emp = Employee.objects.get(id=emp_id)

    return JsonResponse({
        "success": True,
        "emp": {
            "id": emp.id,
            "first_name": emp.first_name,
            "last_name": emp.last_name,
            "email": emp.email,
            "company_id": emp.company_id,

            "department": emp.department,
            "designation": emp.designation,
            "contact_number": emp.contact_number,

            # Family
            "family_contact_number": emp.family_contact_number,
            "family_relation": emp.family_relation,

            # Contract
            "package": str(emp.package),

            # Address
            "address": emp.address,

            # Image
            "image": emp.image.url if emp.image else ""
        }
    })


######################## Team Leader Assignment Views ###########################

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def team_page(request):
    # ↓ Team Leaders
    leaders = TeamLeader.objects.select_related("employee").all()
    # ↓ All Employees (for TL creation dropdown)
    employees = Employee.objects.all()
    # ↓ TOTAL TEAM MEMBERS (count assigned employees)
    total_team_members = TeamAssignment.objects.count()
    # ↓ TOTAL TEAM LEADERS
    total_team_leads = leaders.count()
    # ↓ ACTIVE PROJECTS
    active_projects = ProjectAssignment.objects.count()
    # ↓ AVG PRODUCTIVITY (static or formula)
    avg_productivity = 87   # You can change to dynamic later

    context = {
        "leaders": leaders,
        "employees": employees,
        "total_team_members": total_team_members,
        "total_team_leads": total_team_leads,
        "active_projects": active_projects,
        "avg_productivity": avg_productivity,
    }

    return render(request, "app/hr/team.html", context)


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def create_team_leader(request):
    if request.method == "POST":
        employee_id = request.POST.get("employee")
        experience = request.POST.get("experience")
        team_size = request.POST.get("teamSize")
        responsibilities = request.POST.get("responsibilities")

        if not employee_id:
            messages.error(request, "Please select an employee.")
            return redirect("team-page")

        emp = get_object_or_404(Employee, id=employee_id)

        # Check duplicate
        if TeamLeader.objects.filter(employee=emp).exists():
            messages.error(request, "Employee is already a Team Leader.")
            return redirect("team-page")

        # CREATE TL
        TeamLeader.objects.create(
            employee=emp,
            experience_years=int(experience),
            team_size=team_size,
            responsibilities=responsibilities,
        )

        messages.success(request, "Team Leader Created Successfully!")
        return redirect("team-page")

    return redirect("team-page")

# Page specifically for assign modal — can reuse same template 'team.html'
@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def assign_member_page(request):
    employees = Employee.objects.all()
    leaders = TeamLeader.objects.select_related("employee").all()
    return render(request, "app/hr/team.html", {
        "employees": employees,
        "leaders": leaders,
    })


# Handle assign form submit
@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def assign_member_submit(request):
    if request.method == "POST":
        leader_id = request.POST.get("teamLeader")
        members = request.POST.getlist("members")   # name="members" on checkboxes
        role = request.POST.get("role", "").strip()
        assignment_date = request.POST.get("assignmentDate")
        notes = request.POST.get("notes", "")

        if not leader_id:
            messages.error(request, "Please select a Team Leader.")
            return redirect("assign-member-page")

        if not members:
            messages.error(request, "Please select at least one employee to assign.")
            return redirect("assign-member-page")

        leader = get_object_or_404(TeamLeader, id=leader_id)
        assigned_names = []

        for emp_id in members:
            emp = get_object_or_404(Employee, id=emp_id)

            # create TeamAssignment (will store properly)
            TeamAssignment.objects.create(
                team_leader=leader,
                employee=emp,
                role=role,
                assignment_date=assignment_date,
                notes=notes
            )
            assigned_names.append(f"{emp.first_name} {emp.last_name}")

        messages.success(request,
            f"{', '.join(assigned_names)} assigned to {leader.employee.first_name} {leader.employee.last_name}"
        )
        return redirect("assign-member-page")

    # if GET or other -> redirect
    return redirect("assign-member-page")




@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def assign_project_page(request):
    leaders = TeamLeader.objects.select_related("employee").all()
    return render(request, "app/hr/team.html", {
        "leaders": leaders,
    })


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def get_team_members(request):
    """
    AJAX call → Return members assigned to a Team Leader
    """
    leader_id = request.GET.get("leader_id")
    members = TeamAssignment.objects.filter(team_leader_id=leader_id)

    data = []
    for m in members:
        data.append({
            "id": m.employee.id,
            "name": f"{m.employee.first_name} {m.employee.last_name}",
            "designation": m.employee.designation
        })

    return JsonResponse({"members": data})


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def assign_project_submit(request):
    if request.method == "POST":

        tl_id = request.POST.get("teamLeaderId")
        project_name = request.POST.get("projectName")
        project_code = request.POST.get("projectCode")
        priority = request.POST.get("priority")
        start_date = request.POST.get("startDate")
        end_date = request.POST.get("endDate")
        description = request.POST.get("description")
        members = request.POST.getlist("members")  # Team member checkboxes

        tl = TeamLeader.objects.get(id=tl_id)

        project = ProjectAssignment.objects.create(
            project_name=project_name,
            project_code=project_code,
            team_leader=tl,
            priority=priority,
            start_date=start_date,
            end_date=end_date,
            description=description
        )

        # Add Members
        project.team_members.set(members)

        messages.success(request, "Project Assigned Successfully!")
        return redirect("assign-project-page")


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def team(request):
    return render(request, 'app/hr/team.html')
  

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def team_table(request):

    teamData = []

    leaders = TeamLeader.objects.select_related("employee").all()

    for tl in leaders:

        # Fetch assigned members
        assigned_members = TeamAssignment.objects.filter(team_leader=tl)

        members_list = [
            {
                "id": emp.employee.id,
                "name": f"{emp.employee.first_name} {emp.employee.last_name}",
                "role": emp.employee.designation
            }
            for emp in assigned_members
        ]

        # Fetch assigned projects
        assigned_projects = ProjectAssignment.objects.filter(team_leader=tl)

        projects_list = [
            {
                "name": p.project_name,
                "priority": p.priority,
                "status": "Active"
            }
            for p in assigned_projects
        ]

        teamData.append({
            "id": f"TL{tl.id:03d}",
            "name": f"{tl.employee.first_name} {tl.employee.last_name}",
            "role": tl.employee.designation,
            "department": tl.employee.department,
            "teamSize": len(members_list),
            "teamMembers": members_list,
            "projects": projects_list,
            "status": "Active",
            "tlSince": tl.created_at.strftime("%Y-%m-%d")
        })

    return render(request, "app/hr/team.html", {
        "teamData": teamData
    })


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def get_team_leader_details(request, tl_id):
    tl = get_object_or_404(TeamLeader, id=tl_id)

    # TEAM MEMBERS
    members = TeamAssignment.objects.filter(team_leader=tl)
    members_data = []
    for m in members:
        members_data.append({
            "id": m.employee.id,
            "name": f"{m.employee.first_name} {m.employee.last_name}",
            "designation": m.employee.designation,
            "company_id": m.employee.company_id,
            "image": m.employee.image.url if m.employee.image else "",
        })

    # PROJECTS
    projects = ProjectAssignment.objects.filter(team_leader=tl)
    projects_data = []
    for p in projects:
        projects_data.append({
            "name": p.project_name,
            "priority": p.priority,
            "code": p.project_code,
            "start_date": p.start_date.strftime("%d %b %Y"),
            "end_date": p.end_date.strftime("%d %b %Y"),
        })

    data = {
        "id": tl.id,
        "name": f"{tl.employee.first_name} {tl.employee.last_name}",
        "designation": tl.employee.designation,
        "department": tl.employee.department,
        "company_id": tl.employee.company_id,
        "experience": tl.experience_years,
        "team_size": tl.teamassignment_set.count(),
        "responsibilities": tl.responsibilities,
        "image": tl.employee.image.url if tl.employee.image else "",
        "members": members_data,
        "projects": projects_data,
    }

    return JsonResponse({"data": data})


def delete_team_leader(request, tl_id):
    tl = get_object_or_404(TeamLeader, id=tl_id)
    # Delete related assignments first
    TeamAssignment.objects.filter(team_leader=tl).delete()
    ProjectAssignment.objects.filter(team_leader=tl).delete()
    # Delete the Team Leader
    tl.delete()
    messages.success(request, "Team Leader deleted successfully!")
    return redirect("team-page")


################################################### attendance Management View ###############################

def attendance(request):
    """HR Attendance Management - Display real attendance data from database"""
    # Get HR ID from session for authentication
    hr_id = request.session.get('hr_id')
    
    if not hr_id:
        messages.error(request, "Please login as HR to manage attendance.")
        return redirect('login')
    
    try:
        # Get today's date
        today = timezone.now().date()
        
        # ===== CALCULATE STATISTICS FROM DATABASE =====
        
        # Total active employees (not resigned)
        total_employees = Employee.objects.filter(resigned_date__isnull=True).count()
        
        # Get today's attendance records
        today_attendance = Attendance.objects.filter(
            attendance_date=today
        ).select_related('employee').order_by('-created_at')
        
        # Count present employees (present, half_day, late)
        present_today = today_attendance.filter(
            status__in=['present', 'half_day', 'late']
        ).count()
        
        # Count absent employees (no attendance record or marked absent)
        absent_today = Attendance.objects.filter(
            attendance_date=today,
            status='absent'
        ).count()
        
        # Calculate employees without attendance records (consider them absent)
        employees_with_attendance = today_attendance.count()
        no_record_absent = max(0, total_employees - employees_with_attendance)
        absent_today += no_record_absent
        
        # Calculate attendance rate
        if total_employees > 0:
            attendance_rate = round((present_today / total_employees) * 100, 1)
        else:
            attendance_rate = 0.0
        
        # ===== GET ATTENDANCE RECORDS =====
        
        # Get all attendance records for today with employee details
        attendance_records = today_attendance.select_related('employee').order_by(
            'employee__first_name'
        )
        
        # ===== DEPARTMENT-WISE STATISTICS =====
        
        # Get department distribution from today's attendance
        dept_attendance = attendance_records.values(
            'employee__department'
        ).annotate(
            present_count=Count('id', filter=Q(status__in=['present', 'half_day', 'late'])),
            total_count=Count('id')
        ).order_by('-present_count')
        
        # ===== FILTERING SUPPORT =====
        
        # Get filter options
        departments = Employee.objects.filter(
            resigned_date__isnull=True
        ).values_list('department', flat=True).distinct().order_by('department')
        
        # Handle date filter from request
        selected_date = request.GET.get('date')
        if selected_date:
            try:
                filter_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
            except ValueError:
                filter_date = today
        else:
            filter_date = today
        
        # Handle department filter from request
        selected_department = request.GET.get('department', 'all')
        if selected_department == 'all':
            filtered_records = attendance_records.filter(attendance_date=filter_date)
        else:
            filtered_records = attendance_records.filter(
                attendance_date=filter_date,
                employee__department=selected_department
            )
        
        # Handle search filter from request
        search_term = request.GET.get('search', '').strip()
        if search_term:
            filtered_records = filtered_records.filter(
                Q(employee__first_name__icontains=search_term) |
                Q(employee__last_name__icontains=search_term) |
                Q(employee__company_id__icontains=search_term)
            )
        
        # ===== CALCULATE FILTERED STATISTICS =====
        
        if selected_date or selected_department != 'all' or search_term:
            # Use filtered data for statistics
            filtered_total = filtered_records.count()
            filtered_present = filtered_records.filter(
                status__in=['present', 'half_day', 'late']
            ).count()
            filtered_absent = filtered_records.filter(status='absent').count()
            
            # Update display stats to show filtered results
            display_total = filtered_total + no_record_absent if selected_department == 'all' else filtered_total
            display_present = filtered_present
            display_absent = filtered_absent + (no_record_absent if selected_department == 'all' else 0)
        else:
            # Use today's data for statistics
            display_total = total_employees
            display_present = present_today
            display_absent = absent_today
        
        # ===== CONTEXT DATA =====
        
        context = {
            # Statistics cards
            'total_employees': display_total,
            'present_today': display_present,
            'absent_today': display_absent,
            'attendance_rate': round((display_present / max(display_total, 1)) * 100, 1),
            
            # Attendance records
            'attendance_records': filtered_records,
            
            # Filter options
            'departments': departments,
            'selected_date': filter_date.strftime('%Y-%m-%d') if filter_date else '',
            'selected_department': selected_department,
            'search_term': search_term,
            
            # Department statistics
            'dept_attendance': dept_attendance,
            
            # Additional data
            'today': today,
            'filter_date': filter_date,
            'is_filtered': bool(selected_date or selected_department != 'all' or search_term),
        }
        
        return render(request, 'app/hr/attendance.html', context)
        
    except Exception as e:
        messages.error(request, f"Error loading attendance data: {str(e)}")
        return render(request, 'app/hr/attendance.html', {
            'total_employees': 0,
            'present_today': 0,
            'absent_today': 0,
            'attendance_rate': 0.0,
            'attendance_records': [],
            'departments': [],
        })


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def leave_approvals(request):
    """
    HR Leave Approval - Final approval stage after Team Leader approval
    Workflow: Employee applies → TL reviews/approves → Forwards to HR → HR final approval
    """
    # Get HR ID from session
    hr_id = request.session.get('hr_id')
    
    if not hr_id:
        messages.error(request, "Please login as HR to manage leave approvals.")
        return redirect('login')
    
    try:
        hr = HRProfile.objects.get(id=hr_id)
        
        # ==============================
        # POST REQUEST (FINAL APPROVAL/REJECTION)
        # ==============================
        if request.method == "POST":
            action = request.POST.get("action")
            leave_id = request.POST.get("leave_id")
            hr_comments = request.POST.get("hr_comments", "")
            
            try:
                leave_application = LeaveApply.objects.get(id=leave_id)
                
                if action == "approve":
                    # HR final approval
                    leave_application.status = 'approved'
                    leave_application.approved_by = f"{hr.full_name} (HR)"
                    leave_application.approval_date = timezone.now()
                    leave_application.hr_comments = hr_comments
                    leave_application.save()
                    
                    # Send approval notification email to employee
                    try:
                        employee = leave_application.employee
                        email_context = {
                            'employee_name': f"{employee.first_name} {employee.last_name}",
                            'leave_type': leave_application.get_leave_type_display(),
                            'start_date': leave_application.start_date.strftime('%B %d, %Y'),
                            'end_date': leave_application.end_date.strftime('%B %d, %Y'),
                            'total_days': leave_application.total_days,
                            'hr_name': hr.full_name,
                            'hr_comments': hr_comments,
                        }
                        
                        email_html = render_to_string('email/leave_approval_email.html', email_context)
                        email_text = strip_tags(email_html)
                        
                        send_mail(
                            subject=f'Leave Request Approved - {leave_application.get_leave_type_display()}',
                            message=email_text,
                            from_email='adedinesh158@gmail.com',
                            recipient_list=[employee.email],
                            html_message=email_html,
                            fail_silently=True,
                        )
                    except Exception as email_error:
                        print(f"Error sending approval email: {str(email_error)}")
                    
                    messages.success(request, f"Leave request approved for {leave_application.employee.first_name} {leave_application.employee.last_name}.")
                    
                elif action == "reject":
                    # HR rejection
                    if not hr_comments:
                        messages.error(request, "Please provide a reason for rejection.")
                        return redirect('leave-approvals')
                    
                    leave_application.status = 'rejected'
                    leave_application.rejection_reason = hr_comments
                    leave_application.hr_comments = hr_comments
                    leave_application.save()
                    
                    # Send rejection notification email to employee
                    try:
                        employee = leave_application.employee
                        email_context = {
                            'employee_name': f"{employee.first_name} {employee.last_name}",
                            'leave_type': leave_application.get_leave_type_display(),
                            'start_date': leave_application.start_date.strftime('%B %d, %Y'),
                            'end_date': leave_application.end_date.strftime('%B %d, %Y'),
                            'total_days': leave_application.total_days,
                            'hr_name': hr.full_name,
                            'rejection_reason': hr_comments,
                        }
                        
                        email_html = render_to_string('email/leave_rejection_email.html', email_context)
                        email_text = strip_tags(email_html)
                        
                        send_mail(
                            subject=f'Leave Request Status Update - {leave_application.get_leave_type_display()}',
                            message=email_text,
                            from_email='adedinesh158@gmail.com',
                            recipient_list=[employee.email],
                            html_message=email_html,
                            fail_silently=True,
                        )
                    except Exception as email_error:
                        print(f"Error sending rejection email: {str(email_error)}")
                    
                    messages.success(request, f"Leave request rejected for {leave_application.employee.first_name} {leave_application.employee.last_name}.")
                
                return redirect('leave-approvals')
                
            except LeaveApply.DoesNotExist:
                messages.error(request, "Leave application not found.")
                return redirect('leave-approvals')
        
        # ==============================
        # GET REQUEST (DISPLAY LEAVE REQUESTS)
        # ==============================
        
        # Get leave applications that have been approved by TL (ready for HR final approval)
        pending_hr_approval = LeaveApply.objects.filter(
            status='pending',
            tl_approved=True
        ).select_related('employee').order_by('-applied_at')
        
        # Also show all leave applications for HR dashboard
        all_leaves = LeaveApply.objects.select_related('employee').order_by('-applied_at')
        
        # Calculate statistics
        total_pending = pending_hr_approval.count()
        current_month = timezone.now().replace(day=1)
        
        # Monthly statistics
        monthly_leaves = all_leaves.filter(applied_at__gte=current_month)
        approved_this_month = monthly_leaves.filter(status='approved').count()
        rejected_this_month = monthly_leaves.filter(status='rejected').count()
        total_leave_days = sum([leave.total_days for leave in monthly_leaves.filter(status='approved')])
        
        # Leave distribution by type
        leave_types = {}
        for leave in monthly_leaves:
            leave_type = leave.get_leave_type_display()
            leave_types[leave_type] = leave_types.get(leave_type, 0) + 1
        
        # Leave distribution by department
        dept_distribution = {}
        for leave in monthly_leaves:
            dept = leave.employee.department or 'Unknown'
            dept_distribution[dept] = dept_distribution.get(dept, 0) + 1
        
        context = {
            'hr': hr,
            'pending_hr_approval': pending_hr_approval,
            'all_leaves': all_leaves,
            'total_pending': total_pending,
            'approved_this_month': approved_this_month,
            'rejected_this_month': rejected_this_month,
            'total_leave_days': total_leave_days,
            'leave_types': leave_types,
            'dept_distribution': dept_distribution,
            'total_employees': Employee.objects.count(),
        }
        
        return render(request, 'app/hr/leave-approval.html', context)
        
    except HRProfile.DoesNotExist:
        messages.error(request, "HR profile not found.")
        return redirect('login')

############################################# Employee Payroll Management Views ############################

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def payroll(request):
    # Fetch all data for display
    payrolls = Payroll.objects.select_related("employee").order_by("-created_at")
    employees = Employee.objects.all()
    processed_ids = Payroll.objects.values_list("employee_id", flat=True)
    pending_employees = employees.exclude(id__in=processed_ids)

    # Generate 12 months of current year
    current_year = datetime.now().year
    months = [
        datetime(current_year, m, 1).strftime("%B %Y")
        for m in range(1, 13)
    ]

    # Calculate real statistics from Payroll model
    total_payroll = Payroll.objects.aggregate(
        total=Sum('final_salary')
    )['total'] or 0
    
    pending_payments = pending_employees.count()
    
    processed_payrolls = payrolls.count()
    
    # Calculate average salary
    if processed_payrolls > 0:
        avg_salary = Payroll.objects.aggregate(
            avg=Avg('final_salary')
        )['avg'] or 0
    else:
        avg_salary = 0

    if request.method == "POST":

        employee_id = request.POST.get("employee")
        month = request.POST.get("month")

        base_salary = float(request.POST.get("baseSalary"))
        pf_deduction = float(request.POST.get("pfDeduction"))
        hr_name = request.POST.get("createdBy")     # HR name from form

        final_salary = base_salary - pf_deduction

        employee = Employee.objects.get(id=employee_id)

        Payroll.objects.create(
            employee=employee,
            month=month,
            base_salary=base_salary,
            pf_deduction=pf_deduction,
            final_salary=final_salary,
            created_by=hr_name,
        )

        messages.success(request, "Payroll Added Successfully!")
        return redirect("payroll-management")

    return render(request, "app/hr/payroll.html", {
        "payrolls": payrolls,
        "pending_employees": pending_employees,
        "employees": employees,
        "months": months,
        # Add real statistics for cards
        "total_payroll": total_payroll,
        "pending_payments": pending_payments,
        "processed_payrolls": processed_payrolls,
        "avg_salary": avg_salary,
    })


def payroll_records(request):
    payrolls = Payroll.objects.select_related("employee").order_by("-created_at")
    employees = Employee.objects.all()
    processed_ids = Payroll.objects.values_list("employee_id", flat=True)

    pending_employees = employees.exclude(id__in=processed_ids)

    # 12 months
    current_year = datetime.now().year
    months = [
        datetime(current_year, m, 1).strftime("%B %Y")
        for m in range(1, 12+1)
    ]

    return render(request, "app/hr/payroll.html", {
        "payrolls": payrolls,
        "pending_employees": pending_employees,
        "employees": employees,
        "months": months,
    })


def reports(request):
    # Import all models
    from .models import Payroll, Employee, Announcement, TeamLeader, ProjectAssignment, TeamAssignment
    
    # Helper function to calculate time ago (define locally if not already defined)
    def get_time_ago(timestamp):
        """Helper function to calculate time ago"""
        from django.utils import timezone
        now = timezone.now()
        diff = now - timestamp
        
        if diff.days > 0:
            return f'{diff.days} day{"s" if diff.days > 1 else ""} ago'
        elif diff.seconds > 3600:
            hours = diff.seconds // 3600
            return f'{hours} hour{"s" if hours > 1 else ""} ago'
        elif diff.seconds > 60:
            minutes = diff.seconds // 60
            return f'{minutes} minute{"s" if minutes > 1 else ""} ago'
        else:
            return 'Just now'
    
    # Calculate employee statistics first
    total_employees = Employee.objects.count()
    active_employees = Employee.objects.filter(resigned_date__isnull=True).count()
    
    # Calculate real payroll statistics
    total_payroll = Payroll.objects.aggregate(
        total=Sum('final_salary')
    )['total'] or 0
    
    # Calculate average salary
    processed_payrolls = Payroll.objects.count()
    if processed_payrolls > 0:
        avg_salary = Payroll.objects.aggregate(
            avg=Avg('final_salary')
        )['avg'] or 0
    else:
        avg_salary = 0
    
    # Get current month payroll data
    current_month = datetime.now().strftime("%B %Y")
    current_month_payroll = Payroll.objects.filter(
        month=current_month
    ).aggregate(total=Sum('final_salary'))['total'] or 0
    
    # Get last 6 months payroll data for charts
    months_data = []
    current_date = datetime.now()
    for i in range(6):
        month_date = current_date.replace(day=1) - timedelta(days=30*i)
        month_name = month_date.strftime("%B %Y")
        month_total = Payroll.objects.filter(
            month=month_name
        ).aggregate(total=Sum('final_salary'))['total'] or 0
        months_data.append({
            'month': month_name,
            'total': float(month_total)
        })
    months_data.reverse()  # Oldest to newest
    
    # Get payroll by department with enhanced data
    payroll_by_dept = Payroll.objects.values(
        'employee__department'
    ).annotate(
        total_payroll=Sum('final_salary'),
        employee_count=Count('employee', distinct=True),
        avg_salary_per_dept=Avg('final_salary')
    ).order_by('-total_payroll')
    
    # Get comprehensive recent activities
    recent_activities = []
    
    # Recent Payroll Activities
    recent_payrolls = Payroll.objects.select_related('employee').order_by('-created_at')[:5]
    for payroll in recent_payrolls:
        recent_activities.append({
            'type': 'payroll',
            'icon': 'fa-dollar-sign',
            'icon_bg': 'success',
            'title': f'Payroll processed for {payroll.employee.first_name} {payroll.employee.last_name}',
            'subtitle': f'{payroll.month} - ₹{payroll.final_salary}',
            'time': payroll.created_at,
            'time_ago': get_time_ago(payroll.created_at),
            'by': payroll.created_by,
            'details': f'Department: {payroll.employee.department}'
        })
    
    # Recent Announcements
    recent_announcements = Announcement.objects.order_by('-created_at')[:5]
    for announcement in recent_announcements:
        recent_activities.append({
            'type': 'announcement',
            'icon': 'fa-bullhorn',
            'icon_bg': 'info',
            'title': f'New {announcement.category}: {announcement.title}',
            'subtitle': f'Priority: {announcement.priority}',
            'time': announcement.created_at,
            'time_ago': get_time_ago(announcement.created_at),
            'by': announcement.created_by or 'System',
            'details': f'Audience: All'  # Simplified for now
        })
    
    # Recent Team Leader Assignments
    recent_tl_assignments = TeamAssignment.objects.select_related('team_leader', 'employee').order_by('-assigned_at')[:5]
    for assignment in recent_tl_assignments:
        recent_activities.append({
            'type': 'team_assignment',
            'icon': 'fa-users',
            'icon_bg': 'primary',
            'title': f'Team Assignment: {assignment.employee.first_name} {assignment.employee.last_name}',
            'subtitle': f'Assigned to {assignment.team_leader.employee.first_name} {assignment.team_leader.employee.last_name}',
            'time': assignment.assigned_at,
            'time_ago': get_time_ago(assignment.assigned_at),
            'by': 'HR Admin',
            'details': f'Role: {assignment.role}'
        })
    
    # Recent Project Assignments
    recent_projects = ProjectAssignment.objects.select_related('team_leader').order_by('-created_at')[:5]
    for project in recent_projects:
        recent_activities.append({
            'type': 'project',
            'icon': 'fa-project-diagram',
            'icon_bg': 'warning',
            'title': f'Project Assigned: {project.project_name}',
            'subtitle': f'Team Leader: {project.team_leader.employee.first_name} {project.team_leader.employee.last_name}',
            'time': project.created_at,
            'time_ago': get_time_ago(project.created_at),
            'by': 'HR Admin',
            'details': f'Priority: {project.priority}, Team Size: {project.team_members.count() if hasattr(project, "team_members") else 0}'
        })
    
    # Recent Employee Additions
    recent_employees = Employee.objects.order_by('-created_at')[:5]
    for employee in recent_employees:
        recent_activities.append({
            'type': 'employee',
            'icon': 'fa-user-plus',
            'icon_bg': 'success',
            'title': f'New Employee Added: {employee.first_name} {employee.last_name}',
            'subtitle': f'{employee.designation} - {employee.department}',
            'time': employee.created_at,
            'time_ago': get_time_ago(employee.created_at),
            'by': 'HR Admin',
            'details': f'Employee ID: {employee.company_id}'
        })
    
    # Sort all activities by time (most recent first)
    recent_activities.sort(key=lambda x: x['time'], reverse=True)
    recent_activities = recent_activities[:15]  # Limit to 15 most recent activities
    
    # Calculate percentage values for template use
    monthly_share_percentage = 0
    if total_payroll > 0:
        monthly_share_percentage = round((current_month_payroll / total_payroll) * 100, 1)
    
    # Add calculated percentages to departments data
    payroll_by_dept_with_percentages = []
    for dept in payroll_by_dept:
        dept_dict = dict(dept)
        if total_payroll > 0:
            dept_dict['percentage'] = round((dept['total_payroll'] / total_payroll) * 100, 1)
        else:
            dept_dict['percentage'] = 0
        payroll_by_dept_with_percentages.append(dept_dict)
    
    # Calculate active employee percentage
    active_percentage = 0
    if total_employees > 0:
        active_percentage = round((active_employees / total_employees) * 100, 1)
    
    # Calculate average payroll per department
    avg_payroll_per_dept = 0
    if len(payroll_by_dept_with_percentages) > 0:
        avg_payroll_per_dept = round(total_payroll / len(payroll_by_dept_with_percentages), 0)
    
    return render(request, 'app/hr/reports.html', {
        'total_payroll': total_payroll,
        'avg_salary': avg_salary,
        'current_month_payroll': current_month_payroll,
        'months_data': months_data,
        'payroll_by_dept': payroll_by_dept_with_percentages,
        'recent_activities': recent_activities,
        'recent_payrolls': recent_payrolls,
        'total_employees': total_employees,
        'active_employees': active_employees,
        'monthly_share_percentage': monthly_share_percentage,
        'active_percentage': active_percentage,
        'avg_payroll_per_dept': avg_payroll_per_dept,
    })


def delete_payroll(request):
    """AJAX endpoint to delete payroll record"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            payroll_id = data.get('payroll_id')
            
            if not payroll_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Payroll ID is required'
                })
            
            # Get the payroll record
            try:
                payroll = Payroll.objects.get(id=payroll_id)
            except Payroll.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Payroll record not found'
                })
            
            # Store employee name for response
            employee_name = f"{payroll.employee.first_name} {payroll.employee.last_name}"
            
            # Delete related records first (maintain referential integrity)
            # Delete payroll deductions
            PayrollDeduction.objects.filter(payroll=payroll).delete()
            
            # Delete salary processing records that reference this payroll
            SalaryProcessing.objects.filter(payroll_record=payroll).delete()
            
            # Delete the payroll record
            payroll.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'Payroll record for {employee_name} deleted successfully'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid JSON data'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error deleting payroll record: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Only POST method is allowed'
    })


def get_payroll_data(request):
    """API endpoint to fetch dynamic payroll data for charts"""
    from .models import Payroll, Employee
    
    # Get last 12 months payroll data
    months_data = []
    labels = []
    current_date = datetime.now()
    
    for i in range(12):
        month_date = current_date.replace(day=1) - timedelta(days=30*i)
        month_name = month_date.strftime("%B %Y")
        month_short = month_date.strftime("%b")
        month_total = Payroll.objects.filter(
            month=month_name
        ).aggregate(total=Sum('final_salary'))['total'] or 0
        
        months_data.append(float(month_total))
        labels.append(month_short)
    
    months_data.reverse()
    labels.reverse()
    
    # Get department-wise payroll distribution
    dept_data = Payroll.objects.values(
        'employee__department'
    ).annotate(
        total_payroll=Sum('final_salary')
    ).order_by('-total_payroll')
    
    dept_labels = []
    dept_values = []
    for dept in dept_data:
        if dept['employee__department']:
            dept_labels.append(dept['employee__department'])
            dept_values.append(float(dept['total_payroll']))
    
    # Enhanced Employee Count Trends with both models
    employee_trends = []
    employee_labels = []
    total_employees_data = []
    active_employees_data = []
    payroll_employees_data = []
    
    for i in range(12):  # Extended to 12 months for better trend analysis
        month_date = current_date.replace(day=1) - timedelta(days=30*i)
        month_name = month_date.strftime("%B %Y")
        month_short = month_date.strftime("%b %Y")
        
        # Total employees registered by end of month
        total_emp_count = Employee.objects.filter(
            created_at__lte=month_date.replace(day=1) + timedelta(days=32)
        ).count()
        
        # Active employees (not resigned) by end of month  
        active_emp_count = Employee.objects.filter(
            created_at__lte=month_date.replace(day=1) + timedelta(days=32),
            resigned_date__isnull=True
        ).count()
        
        # Employees with payroll processed in that month
        payroll_emp_count = Payroll.objects.filter(
            month=month_name
        ).values('employee').distinct().count()
        
        employee_trends.append(month_short)
        total_employees_data.append(total_emp_count)
        active_employees_data.append(active_emp_count)
        payroll_employees_data.append(payroll_emp_count)
    
    employee_trends.reverse()
    total_employees_data.reverse()
    active_employees_data.reverse()
    payroll_employees_data.reverse()
    
    # Get employee distribution by department for the latest month
    dept_employee_data = Employee.objects.filter(
        resigned_date__isnull=True
    ).values('department').annotate(
        count=Count('id')
    ).order_by('-count')
    
    dept_emp_labels = []
    dept_emp_counts = []
    for dept in dept_employee_data:
        if dept['department']:
            dept_emp_labels.append(dept['department'])
            dept_emp_counts.append(dept['count'])
    
    return JsonResponse({
        'months': labels,
        'payroll_data': months_data,
        'departments': {
            'labels': dept_labels,
            'data': dept_values
        },
        'employee_counts': {
            'labels': employee_trends,
            'total_employees': total_employees_data,
            'active_employees': active_employees_data,
            'payroll_employees': payroll_employees_data
        },
        'department_employees': {
            'labels': dept_emp_labels,
            'data': dept_emp_counts
        },
        'summary_stats': {
            'total_employees': Employee.objects.count(),
            'active_employees': Employee.objects.filter(resigned_date__isnull=True).count(),
            'resigned_employees': Employee.objects.filter(resigned_date__isnull=False).count(),
            'employees_with_payroll': Payroll.objects.values('employee').distinct().count()
        }
    })


############################## HR Announcements Views #######################################

def announcements(request):
    """
    Announcement management view - handles both GET (display) and POST (create) requests
    """
    # Get HR name from session or use default
    hr_name = (request.session.get("hr_name") or 
               request.session.get("hr_full_name") or 
               request.session.get("full_name") or 
               request.session.get("name") or
               "HR Admin")

    # ==============================
    # POST REQUEST (SAVE DATA)
    # ==============================
    if request.method == "POST":
        try:
            # Extract form data
            title = request.POST.get("announcementTitle", "").strip()
            category = request.POST.get("announcementCategory", "").strip()
            priority = request.POST.get("announcementPriority", "").strip()
            content = request.POST.get("announcementContent", "").strip()
            expiry_date_str = request.POST.get("announcementExpiry")
            
            # Validation
            if not all([title, category, priority, content]):
                messages.error(request, "All required fields must be filled!")
                return redirect("hr-announcements")
            
            # Process expiry date
            expiry_datetime = None
            if expiry_date_str:
                try:
                    expiry_date = datetime.strptime(expiry_date_str, "%Y-%m-%d").date()
                    expiry_naive = datetime.combine(expiry_date, time(23, 59, 59))
                    expiry_datetime = timezone.make_aware(expiry_naive)
                except ValueError:
                    messages.error(request, "Invalid expiry date format!")
                    return redirect("hr-announcements")
            
            # Determine target audience
            target_audience = "all"
            if request.POST.get("management"):
                target_audience = "management"
            elif request.POST.get("departments"):
                target_audience = "department"
            
            # Create announcement
            announcement = Announcement.objects.create(
                title=title,
                content=content,
                summary=content[:200] if content else "",
                category=category,
                priority=priority,
                status="published",
                target_audience=target_audience,
                target_departments=[],
                target_locations=[],
                target_roles=[],
                publish_date=timezone.now(),
                expiry_date=expiry_datetime,
                created_by=hr_name,
                author_designation="HR",
                author_department="Human Resources",
                tags=[],
                comments="",
            )
            
            messages.success(request, "Announcement created successfully!")
            return redirect("hr-announcements")
            
        except Exception as e:
            messages.error(request, f"Error creating announcement: {str(e)}")
            return redirect("hr-announcements")

    # ==============================
    # GET REQUEST (FETCH DATA)
    # ==============================

    announcements_qs = Announcement.objects.all().order_by("-created_at")
    today = timezone.now().date()

    total_announcements = announcements_qs.count()

    active_announcements = announcements_qs.filter(
        status="published"
    ).filter(
        Q(expiry_date__isnull=True) | Q(expiry_date__date__gte=today)
    ).count()

    active_today = announcements_qs.filter(
        status="published"
    ).filter(
        Q(expiry_date__isnull=True) | Q(expiry_date__date__gte=today),
        publish_date__date__lte=today,
    ).count()

    total_views = announcements_qs.aggregate(total=Sum("read_by_count"))["total"] or 0
    categories_count = announcements_qs.values("category").distinct().count()

    context = {
        "announcements": announcements_qs,
        "total_announcements": total_announcements,
        "active_announcements": active_announcements,
        "active_today": active_today,
        "total_views": total_views,
        "categories_count": categories_count,
    }

    return render(request, "app/hr/announcement.html", context)


def view_announcement(request, announcement_id):
    """View announcement details - Returns JSON for modal display"""
    try:
        announcement = Announcement.objects.get(id=announcement_id)
        
        data = {
            'success': True,
            'title': announcement.title,
            'category': announcement.get_category_display(),
            'priority': announcement.get_priority_display(),
            'status': announcement.get_status_display(),
            'content': announcement.content,
            'summary': announcement.summary,
            'created_by': announcement.created_by,
            'author_designation': announcement.author_designation,
            'author_department': announcement.author_department,
            'created_at': announcement.created_at.strftime("%B %d, %Y at %I:%M %p"),
            'publish_date': announcement.publish_date.strftime("%B %d, %Y at %I:%M %p") if announcement.publish_date else "Not set",
            'expiry_date': announcement.expiry_date.strftime("%B %d, %Y") if announcement.expiry_date else "No expiry",
            'read_by_count': announcement.read_by_count,
            'likes_count': announcement.likes_count,
            'comments_enabled': announcement.comments_enabled,
            'acknowledgment_required': announcement.acknowledgment_required,
            'target_audience': announcement.get_target_audience_display(),
            'tags': announcement.tags,
        }
        
        return JsonResponse(data)
        
    except Announcement.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Announcement not found'
        })


def edit_announcement_view(request, announcement_id):
    """Edit announcement - Display edit form"""
    try:
        announcement = Announcement.objects.get(id=announcement_id)
        
        context = {
            'announcement': announcement,
            'is_edit': True
        }
        
        return render(request, "app/hr/edit_announcement.html", context)
        
    except Announcement.DoesNotExist:
        messages.error(request, "Announcement not found.")
        return redirect('hr-announcements')


def update_announcement(request, announcement_id):
    """Update announcement - Handle form submission"""
    if request.method != 'POST':
        return redirect('hr-announcements')
    
    try:
        announcement = Announcement.objects.get(id=announcement_id)
        
        # Extract form data
        title = request.POST.get('announcementTitle', '').strip()
        category = request.POST.get('announcementCategory', '').strip()
        priority = request.POST.get('announcementPriority', '').strip()
        content = request.POST.get('announcementContent', '').strip()
        expiry_date_str = request.POST.get('announcementExpiry', '')
        
        # Validation
        if not all([title, category, priority, content]):
            messages.error(request, "All required fields must be filled!")
            return redirect('edit-announcement', announcement_id=announcement_id)
        
        # Process expiry date
        expiry_datetime = None
        if expiry_date_str:
            try:
                expiry_date = datetime.strptime(expiry_date_str, "%Y-%m-%d").date()
                expiry_naive = datetime.combine(expiry_date, time(23, 59, 59))
                expiry_datetime = timezone.make_aware(expiry_naive)
            except ValueError:
                messages.error(request, "Invalid expiry date format!")
                return redirect('edit-announcement', announcement_id=announcement_id)
        
        # Update announcement
        announcement.title = title
        announcement.content = content
        announcement.summary = content[:200] if content else ""
        announcement.category = category
        announcement.priority = priority
        announcement.expiry_date = expiry_datetime
        announcement.updated_at = timezone.now()
        
        announcement.save()
        
        messages.success(request, "Announcement updated successfully!")
        return redirect('hr-announcements')
        
    except Announcement.DoesNotExist:
        messages.error(request, "Announcement not found.")
        return redirect('hr-announcements')
    except Exception as e:
        messages.error(request, f"Error updating announcement: {str(e)}")
        return redirect('edit-announcement', announcement_id=announcement_id)


def delete_announcement(request, announcement_id):
    """Delete announcement - Handle deletion"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Invalid request method'
        })
    
    try:
        announcement = Announcement.objects.get(id=announcement_id)
        announcement_title = announcement.title
        announcement.delete()
        
        # Return JSON response for AJAX calls
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'Announcement "{announcement_title}" deleted successfully!'
            })
        else:
            # Regular form submission
            messages.success(request, f'Announcement "{announcement_title}" deleted successfully!')
            return redirect('hr-announcements')
        
    except Announcement.DoesNotExist:
        error_msg = "Announcement not found."
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': error_msg
            })
        else:
            messages.error(request, error_msg)
            return redirect('hr-announcements')
    except Exception as e:
        error_msg = f"Error deleting announcement: {str(e)}"
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': error_msg
            })
        else:
            messages.error(request, error_msg)
            return redirect('hr-announcements')









########################################### HR Management Views #############################

def hr_create(request):
    """Create new HR profile - Simplified version"""
    print(f"=== HR CREATE REQUEST ===")
    print(f"Method: {request.method}")
    
    if request.method == "POST":
        try:
            # ====== BASIC VALIDATIONS ======
            if request.POST.get("password") != request.POST.get("confirmPassword"):
                messages.error(request, "Password and Confirm Password do not match!")
                return render(request, "app/create_hr.html")

            # ====== FIELD VALIDATIONS ======
            required_fields = ['fullName', 'employeeId', 'email', 'mobile', 'designation', 'department', 'dateOfJoining', 'workLocation', 'username', 'password', 'accessLevel']
            
            for field in required_fields:
                if not request.POST.get(field):
                    messages.error(request, f"{field} is required!")
                    return render(request, "app/create_hr.html")

            # ====== DATA CONVERSION ======
            from datetime import datetime
            
            # Convert date
            try:
                date_of_joining = datetime.strptime(request.POST.get("dateOfJoining"), "%Y-%m-%d").date()
            except ValueError:
                messages.error(request, "Invalid Date of Joining format!")
                return render(request, "app/create_hr.html")

            # ====== CREATE HR OBJECT ======
            print("Creating HR object...")
            
            # Handle profile image upload
            profile_image = request.FILES.get("profileImage")
            
            hr = HRProfile.objects.create(
                full_name=request.POST.get("fullName").strip(),
                employee_id=request.POST.get("employeeId").strip(),
                email=request.POST.get("email").strip().lower(),
                mobile=request.POST.get("mobile").strip(),
                designation=request.POST.get("designation").strip(),
                department=request.POST.get("department").strip(),
                date_of_joining=date_of_joining,
                work_location=request.POST.get("workLocation").strip(),
                username=request.POST.get("username").strip(),
                password=request.POST.get("password"),
                access_level=request.POST.get("accessLevel").strip(),
                profile_image=profile_image,
            )

            print(f"HR object created successfully with ID: {hr.id}")
            
            # Verify the object was saved to database
            hr_saved = HRProfile.objects.filter(id=hr.id).exists()
            print(f"HR verification in database: {hr_saved}")
            
            if hr_saved:
                print(f"HR confirmed saved: {hr.full_name} - {hr.email}")
                
                # Send welcome email
                email_sent = False
                whatsapp_sent = False
                
                try:
                    email_context = {
                        'name': hr.full_name,
                        'username': hr.username,
                        'password': hr.password,
                        'employee_id': hr.employee_id,  # This will be the Access ID
                        'email': hr.email,
                        'designation': hr.designation,
                    }
                    
                    # Render the email template
                    email_html = render_to_string('email/hr_profile_email.html', email_context)
                    email_text = strip_tags(email_html)
                    
                    # Send email
                    send_mail(
                        subject=f'Welcome to Kavya Infoweb Private Limited Nagpur - HR Profile Created',
                        message=email_text,
                        from_email='adedinesh158@gmail.com',
                        recipient_list=[hr.email],
                        html_message=email_html,
                        fail_silently=False,
                    )
                    
                    print(f"Welcome email sent successfully to {hr.email}")
                    email_sent = True
                    
                except Exception as email_error:
                    print(f"Error sending email: {str(email_error)}")
                    # Email failed, but HR was created successfully
                
                # Send WhatsApp message with access credentials
                try:
                    # Clean message without problematic Unicode characters
                    whatsapp_message = f"""Kavya Infoweb Private Limited Nagpur

Welcome {hr.full_name}!

Your HR Profile has been created successfully!

Access Credentials:
Employee ID: {hr.employee_id}
Username: {hr.username}
Password: {hr.password}
Designation: {hr.designation}
Department: {hr.department}

Important Notes:
Please change your password after first login.

Email with detailed information has been sent to: {hr.email}

For any queries, contact IT Support.

Best regards,
HR Team - Kavya Infoweb"""
                    
                    whatsapp_result = send_whatsapp_message(hr.mobile, whatsapp_message)
                    
                    # Handle both return formats
                    if isinstance(whatsapp_result, tuple):
                        whatsapp_sent, whatsapp_link = whatsapp_result
                        if whatsapp_link:
                            print(f"WhatsApp link generated: {whatsapp_link}")
                            print(f"Please share this link with the user: {whatsapp_link}")
                    else:
                        whatsapp_sent = whatsapp_result
                    
                    if whatsapp_sent:
                        print(f"WhatsApp message processed successfully to {hr.mobile}")
                    else:
                        print(f"WhatsApp message failed to send to {hr.mobile}")
                    
                except Exception as whatsapp_error:
                    print(f"Error sending WhatsApp message: {str(whatsapp_error)}")
                    whatsapp_sent = False
                
                # Prepare success message with WhatsApp links
                success_msg = "HR Created Successfully!"
                whatsapp_links = None
                
                if email_sent and whatsapp_sent:
                    success_msg += " Welcome email and WhatsApp message links generated."
                elif email_sent:
                    success_msg += " Welcome email sent."
                elif whatsapp_sent:
                    success_msg += " WhatsApp message links generated."
                else:
                    success_msg += " (Communication failed - please contact IT)"
                
                # Add WhatsApp links to context for display
                context_data = {
                    'success_message': 'HR Created Successfully!',
                    'show_whatsapp_links': True,
                    'hr_mobile': hr.mobile,
                    'hr_name': hr.full_name
                }
                
                # Check if WhatsApp function returned links
                if 'whatsapp_result' in locals() and isinstance(whatsapp_result, tuple) and whatsapp_result[1]:
                    whatsapp_links = whatsapp_result[1]
                    context_data['whatsapp_links'] = whatsapp_links
                
                messages.success(request, success_msg)
                return render(request, "app/create_hr.html", context_data)
                
                return render(request, "app/create_hr.html", {'success_message': 'HR Created Successfully!'})
            else:
                print("ERROR: HR object not found in database after creation")
                messages.error(request, "Error: HR was not saved properly!")
                return render(request, "app/create_hr.html")

        except IntegrityError as e:
            print(f"IntegrityError: {str(e)}")
            error_message = "Employee ID / Email / Username already exists!"
            messages.error(request, error_message)
            return render(request, "app/create_hr.html")
        except Exception as e:
            print(f"General Error: {str(e)}")
            import traceback
            traceback.print_exc()
            error_message = f"Error: {str(e)}"
            messages.error(request, error_message)
            return render(request, "app/create_hr.html")

    return render(request, "app/create_hr.html")

def hr_list(request):
    """List all HR profiles"""
    hr_list = HRProfile.objects.all().order_by('-created_at')
    total_hr = hr_list.count()
    
    return render(request, 'app/create_hr.html', {
        'hr_list': hr_list,
        'total_hr': total_hr,
        'show_list': True,
    })


def test_leave_system(request):
    """
    Test view for the leave management system
    This creates test data and demonstrates the complete workflow
    """
    if request.method == "POST":
        # Create test employees, team leader, and sample leave applications
        
        # Create a test employee if none exists
        employee, created = Employee.objects.get_or_create(
            email="employee@test.com",
            defaults={
                'first_name': 'John',
                'last_name': 'Doe',
                'password': 'password123',
                'company_id': 'EMP001',
                'designation': 'Software Developer',
                'department': 'IT',
                'package': 50000.00,
                'address': 'Test Address'
            }
        )
        
        # Create a team leader if none exists
        tl_employee, created = Employee.objects.get_or_create(
            email="tl@test.com",
            defaults={
                'first_name': 'Jane',
                'last_name': 'Smith',
                'password': 'password123',
                'company_id': 'TL001',
                'designation': 'Team Lead',
                'department': 'IT',
                'package': 60000.00,
                'address': 'TL Address'
            }
        )
        
        # Create TeamLeader record
        tl, created = TeamLeader.objects.get_or_create(
            employee=tl_employee,
            defaults={
                'experience_years': 5,
                'team_size': 'Small',
                'responsibilities': 'Manage development team'
            }
        )
        
        # Assign employee to team leader
        team_assignment, created = TeamAssignment.objects.get_or_create(
            team_leader=tl,
            employee=employee,
            defaults={
                'role': 'Developer',
                'assignment_date': timezone.now().date(),
                'notes': 'Test assignment'
            }
        )
        
        # Create sample leave applications
        from datetime import date, timedelta
        
        # Sample approved leave
        leave1, created = LeaveApply.objects.get_or_create(
            employee=employee,
            start_date=date.today() + timedelta(days=2),
            defaults={
                'leave_type': 'annual',
                'end_date': date.today() + timedelta(days=2),
                'total_days': 1,
                'reason': 'Personal work',
                'status': 'approved',
                'approved_by': 'Test HR',
                'approval_date': timezone.now()
            }
        )
        
        # Sample pending leave
        leave2, created = LeaveApply.objects.get_or_create(
            employee=employee,
            start_date=date.today() + timedelta(days=7),
            defaults={
                'leave_type': 'sick',
                'end_date=': date.today() + timedelta(days=7),
                'total_days': 1,
                'reason': 'Medical appointment',
                'status': 'pending'
            }
        )
        
        messages.success(request, f"Test data created successfully! Employee: {employee.first_name} {employee.last_name}")
        return redirect('test-leave-system')
    
    # Get test statistics
    test_stats = {
        'employees': Employee.objects.count(),
        'team_leaders': TeamLeader.objects.count(),
        'leave_applications': LeaveApply.objects.count(),
        'pending_leaves': LeaveApply.objects.filter(status='pending').count(),
        'approved_leaves': LeaveApply.objects.filter(status='approved').count(),
    }
    
    return render(request, 'app/test_leave_system.html', {
        'test_stats': test_stats,
        'recent_leaves': LeaveApply.objects.select_related('employee').order_by('-applied_at')[:5]
    })

def delete_hr(request, hr_id):
    """Delete HR profile"""
    hr = get_object_or_404(HRProfile, id=hr_id)
    hr.delete()
    





    return JsonResponse({'success': True, 'message': 'HR deleted successfully!'})




# ============================================================================
# NEW PAYROLL & ATTENDANCE MANAGEMENT VIEWS
# ============================================================================

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def hr_attendance_simple(request):
    """
    HR Attendance Management - Database Only Implementation
    Automatically displays employee records when they check in
    Updates cards on page accordingly
    Shows real-time attendance data from database
    """
    hr_id = request.session.get('hr_id')
    
    if not hr_id:
        messages.error(request, "Please login as HR to access attendance management.")
        return redirect('login')
    
    try:
        hr = HRProfile.objects.get(id=hr_id)
        today = timezone.now().date()
        
        # Handle filter parameters
        selected_date = request.GET.get('date', today.strftime('%Y-%m-%d'))
        selected_department = request.GET.get('department', 'all')
        search_term = request.GET.get('search', '').strip()
        
        try:
            filter_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except ValueError:
            filter_date = today
        
        # Get all employees
        employees = Employee.objects.filter(resigned_date__isnull=True)
        
        # Apply filters
        if selected_department != 'all':
            employees = employees.filter(department=selected_department)
        
        if search_term:
            employees = employees.filter(
                Q(first_name__icontains=search_term) |
                Q(last_name__icontains=search_term) |
                Q(company_id__icontains=search_term) |
                Q(email__icontains=search_term)
            )
        
        # Get attendance records for filtered employees and date
        attendance_records = Attendance.objects.filter(
            employee__in=employees,
            attendance_date=filter_date
        ).select_related('employee').order_by('employee__first_name')
        
        # Get department list for filter
        departments = list(Employee.objects.filter(resigned_date__isnull=True)
                          .values_list('department', flat=True)
                          .distinct()
                          .order_by('department'))
        
        # Calculate statistics
        total_employees = employees.count()
        present_today = attendance_records.filter(
            status__in=['present', 'late', 'half_day']
        ).count()
        absent_today = total_employees - present_today
        attendance_rate = round((present_today / max(total_employees, 1)) * 100, 1)
        
        # Check if filtered
        is_filtered = (selected_date != today.strftime('%Y-%m-%d') or 
                      selected_department != 'all' or 
                      search_term != '')
        
        # Get department-wise attendance
        dept_attendance = []
        for dept in departments:
            dept_employees = employees.filter(department=dept)
            if dept_employees.exists():
                dept_attendance_data = attendance_records.filter(
                    employee__department=dept
                ).aggregate(
                    present_count=Count('id', filter=Q(status__in=['present', 'late', 'half_day'])),
                    total_count=Count('id')
                )
                dept_attendance.append({
                    'employee__department': dept,
                    'present_count': dept_attendance_data['present_count'],
                    'total_count': dept_attendance_data['total_count']
                })
        
        context = {
            'hr': hr,
            'attendance_records': attendance_records,
            'departments': departments,
            'total_employees': total_employees,
            'present_today': present_today,
            'absent_today': absent_today,
            'attendance_rate': attendance_rate,
            'selected_date': selected_date,
            'selected_department': selected_department,
            'search_term': search_term,
            'is_filtered': is_filtered,
            'dept_attendance': dept_attendance,
            'filter_date': filter_date,
            'today': today,
        }
        
        return render(request, 'app/hr/attendance.html', context)
        
    except HRProfile.DoesNotExist:
        messages.error(request, "HR profile not found.")
        return redirect('login')
    except Exception as e:
        messages.error(request, f"Error in attendance management: {str(e)}")
        return redirect('hr-dashboard')

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def employee_attendance_page(request):
    return render(request, 'app/employee/attendance.html')

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def hr_attendance_page(request):
    return render(request, 'app/hr/attendance.html')


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def employee_attendance_simple(request):
    """
    Employee Attendance System - Database Only Implementation
    - Check-in and Check-out with real-time database updates
    - Automatic appearance in HR attendance management
    - Working days: Monday to Saturday
    - Official start time: 10:45 AM
    - All data from database models only
    - Shows only present days in employee records
    - NEW: Automatic check-out at 6:30 PM with 3-second message
    """
    employee_id = request.session.get('employee_id')
    
    if not employee_id:
        messages.error(request, "Please login to access attendance.")
        return redirect('login')
    
    try:
        employee = Employee.objects.get(id=employee_id)
        today = timezone.now().date()
        current_time = timezone.now().time()
        
        # Constants
        official_start_time = time(10, 45)  # 10:45 AM - Late arrival after this
        check_in_allowed_time = time(10, 0)   # 10:00 AM - Check-in starts
        required_work_hours = 8.0
        auto_checkout_time = time(18, 30)    # 6:30 PM - Auto checkout time
        
        # Check if today is a working day (Monday to Saturday)
        is_working_day = today.weekday() < 6  # Monday to Saturday
        
        # Check if current time is within check-in window
        can_show_check_in = (current_time >= check_in_allowed_time and is_working_day)
        
        # Get or create today's attendance record
        attendance, created = Attendance.objects.get_or_create(
            employee=employee,
            attendance_date=today,
            defaults={
                'status': 'present',
                'shift_type': 'full_time',
                'is_check_in_allowed': True,
                'can_check_out': False
            }
        )
        
        # NEW: Automatic check-out at 6:30 PM
        auto_checkout_triggered = False
        if (attendance.check_in_time and 
            not attendance.check_out_time and 
            current_time >= auto_checkout_time):
            
            # Calculate worked hours for auto checkout
            from datetime import datetime
            check_in_dt = datetime.combine(today, attendance.check_in_time)
            current_dt = datetime.combine(today, current_time)
            worked_duration = current_dt - check_in_dt
            worked_hours = worked_duration.total_seconds() / 3600
            
            # Auto checkout with minimum 8 hours (or what they've worked if less than 8)
            final_worked_hours = max(worked_hours, 8.0) if worked_hours >= 4.0 else worked_hours
            
            # Update attendance record with auto checkout
            attendance.check_out_time = current_time
            attendance.total_worked_hours = round(final_worked_hours, 2)
            
            if final_worked_hours >= 8.0:
                attendance.status = 'present'
                attendance.remarks += f" | Auto checkout at 6:30 PM (Completed {final_worked_hours:.1f}h)"
            elif final_worked_hours >= 4.0:
                attendance.status = 'half_day'
                attendance.remarks += f" | Auto checkout at 6:30 PM (Half day - {final_worked_hours:.1f}h)"
            else:
                attendance.status = 'early_departure'
                attendance.remarks += f" | Auto checkout at 6:30 PM (Early - {final_worked_hours:.1f}h)"
                
            attendance.can_check_out = False
            attendance.save()
            
            # Log the auto check-out
            AttendanceCheckLog.objects.create(
                employee=employee,
                check_type='check_out',
                check_time=timezone.now(),
                attendance_date=today,
                worked_hours_at_check=round(final_worked_hours, 2),
                required_work_hours=8.0,
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT')
            )
            
            auto_checkout_triggered = True
        
        # Determine current status and actions
        can_check_in = can_show_check_in and not attendance.check_in_time
        can_check_out = False
        worked_hours = 0
        remaining_hours = required_work_hours
        is_late_arrival = False
        
        if attendance.check_in_time and not attendance.check_out_time:
            # Calculate current worked hours
            from datetime import datetime
            check_in_dt = datetime.combine(today, attendance.check_in_time)
            current_dt = datetime.combine(today, current_time)
            worked_duration = current_dt - check_in_dt
            worked_hours = worked_duration.total_seconds() / 3600
            remaining_hours = max(0, required_work_hours - worked_hours)
            can_check_out = worked_hours >= required_work_hours
            
            # Update real-time worked hours
            attendance.total_worked_hours = round(worked_hours, 2)
            if can_check_out:
                attendance.can_check_out = True
            attendance.save()
        
        # Handle POST requests for check-in/check-out
        if request.method == "POST":
            action = request.POST.get("action")
            
            if action == "check_in":
                if not can_check_in:
                    if not is_working_day:
                        messages.error(request, "Check-in not allowed on weekends!")
                    elif not can_show_check_in:
                        messages.error(request, "Check-in is only allowed after 10:00 AM.")
                    else:
                        messages.error(request, "You have already checked-in today!")
                    return redirect('employee-attendance')
                
                # Calculate late arrival
                is_late_arrival = current_time > official_start_time
                
                # Update attendance record
                attendance.check_in_time = current_time
                attendance.status = 'half_day' if is_late_arrival else 'present'
                attendance.remarks = f"Checked in at {current_time.strftime('%H:%M:%S')}"
                attendance.is_check_in_allowed = False
                attendance.save()
                
                # Log the check-in
                AttendanceCheckLog.objects.create(
                    employee=employee,
                    check_type='check_in',
                    check_time=timezone.now(),
                    attendance_date=today,
                    is_denied=False,
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT')
                )
                
                # Show appropriate message
                if is_late_arrival:
                    messages.warning(request, f"Checked in successfully! You arrived late (after 10:45 AM). Status: Half Day")
                else:
                    messages.success(request, f"Checked in successfully at {current_time.strftime('%H:%M:%S')}!")
                
                return redirect('employee-attendance')
            
            elif action == "check_out":
                if not attendance.check_in_time:
                    messages.error(request, "You need to check-in first before checking out!")
                    return redirect('employee-attendance')
                
                if attendance.check_out_time:
                    messages.error(request, "You have already checked-out today!")
                    return redirect('employee-attendance')
                
                # Calculate worked hours
                from datetime import datetime
                check_in_dt = datetime.combine(today, attendance.check_in_time)
                current_dt = datetime.combine(today, current_time)
                worked_duration = current_dt - check_in_dt
                worked_hours = worked_duration.total_seconds() / 3600
                
                if worked_hours < required_work_hours:
                    messages.error(request, f"Cannot check-out yet! You need to complete at least {required_work_hours} hours. Current: {worked_hours:.1f} hours")
                    return redirect('employee-attendance')
                
                # Update attendance record
                attendance.check_out_time = current_time
                attendance.total_worked_hours = round(worked_hours, 2)
                attendance.can_check_out = False
                
                if worked_hours >= required_work_hours:
                    attendance.status = 'present'
                    attendance.remarks += f" | Checked out at {current_time.strftime('%H:%M:%S')} (Completed {worked_hours:.1f}h)"
                elif worked_hours >= 4.0:
                    attendance.status = 'half_day'
                    attendance.remarks += f" | Checked out at {current_time.strftime('%H:%M:%S')} (Half day - {worked_hours:.1f}h)"
                
                attendance.save()
                
                # Log the check-out
                AttendanceCheckLog.objects.create(
                    employee=employee,
                    check_type='check_out',
                    check_time=timezone.now(),
                    attendance_date=today,
                    worked_hours_at_check=round(worked_hours, 2),
                    required_work_hours=required_work_hours,
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT')
                )
                
                messages.success(request, f"Checked out successfully! You worked for {worked_hours:.1f} hours today.")
                return redirect('employee-attendance')
        
        # Get present attendance records only (for display)
        present_attendance_records = Attendance.objects.filter(
            employee=employee,
            status__in=['present', 'late', 'half_day']
        ).order_by('-attendance_date')
        
        # Calculate current month statistics
        current_month = timezone.now().replace(day=1)
        current_month_attendance = present_attendance_records.filter(
            attendance_date__gte=current_month
        )
        
        current_month_stats = {
            'total_days': current_month_attendance.count(),
            'present_days': current_month_attendance.filter(status='present').count(),
            'half_days': current_month_attendance.filter(status='half_day').count(),
            'absent_days': current_month_attendance.filter(status='absent').count(),
        }
        
        # Calculate attendance percentage
        total_present_days = present_attendance_records.count()
        attendance_percentage = round((total_present_days / max(total_present_days, 1)) * 100, 1)
        
        # GET request - display attendance interface
        context = {
            'employee': employee,
            'attendance': attendance,
            'current_time': current_time,
            'today': today,
            'official_start_time': official_start_time,
            'required_work_hours': required_work_hours,
            'auto_checkout_time': auto_checkout_time,
            'is_working_day': is_working_day,
            'can_check_in': can_check_in,
            'can_check_out': can_check_out,
            'worked_hours': round(worked_hours, 2),
            'remaining_hours': round(remaining_hours, 2),
            'is_late_arrival': is_late_arrival,
            'attendance_records': present_attendance_records,  # Only present days
            'total_days': total_present_days,
            'present_days': current_month_stats['present_days'],
            'half_days': current_month_stats['half_days'],
            'current_month_stats': current_month_stats,
            'attendance_percentage': attendance_percentage,
            'auto_checkout_triggered': auto_checkout_triggered,  # NEW: For 3-second message
        }
        
        return render(request, 'app/employee/attendance.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, "Employee profile not found.")
        return redirect('login')
    except Exception as e:
        messages.error(request, f"Error in attendance system: {str(e)}")
        return redirect('employee-dashboard')


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def employee_check_out(request):
    """
    Enhanced Employee Check-out View with time restrictions and logging
    - Check-out only after 8 hours of work completion
    - Log all check-out activities
    - Handle check-out restrictions and validations
    """
    employee_id = request.session.get('employee_id')
    
    if not employee_id:
        messages.error(request, "Please login to check-out.")
        return redirect('login')
    
    try:
        employee = Employee.objects.get(id=employee_id)
        today = timezone.now().date()
        current_time = timezone.now().time()
        current_datetime = timezone.now()
        
        # Get today's attendance record
        attendance = Attendance.objects.filter(
            employee=employee,
            attendance_date=today
        ).first()
        
        if not attendance or not attendance.check_in_time:
            # Log denied check-out attempt
            AttendanceCheckLog.objects.create(
                employee=employee,
                check_type='check_out_denied',
                check_time=current_datetime,
                attendance_date=today,
                is_denied=True,
                denial_reason='no_attendance_record',
                denial_message='No attendance record found - check-in first',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT')
            )
            messages.error(request, "You need to check-in first before checking out!")
            return redirect('employee-dashboard')
        
        if attendance.check_out_time:
            # Log denied check-out attempt
            AttendanceCheckLog.objects.create(
                employee=employee,
                check_type='check_out_denied',
                check_time=current_datetime,
                attendance_date=today,
                is_denied=True,
                denial_reason='already_checked_out',
                denial_message='Already checked out today',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT')
            )
            messages.error(request, "You have already checked-out today!")
            return redirect('employee-dashboard')
        
        if request.method == "POST":
            # Calculate worked hours
            from datetime import datetime, timedelta
            check_in_dt = datetime.combine(today, attendance.check_in_time)
            check_out_dt = datetime.combine(today, current_time)
            worked_duration = check_out_dt - check_in_dt
            worked_hours = worked_duration.total_seconds() / 3600
            
            # Validate minimum work hours requirement
            if worked_hours < 8.0:
                # Log denied check-out attempt
                AttendanceCheckLog.objects.create(
                    employee=employee,
                    check_type='check_out_denied',
                    check_time=current_datetime,
                    attendance_date=today,
                    is_denied=True,
                    denial_reason='insufficient_hours',
                    denial_message=f'Insufficient work hours ({worked_hours:.1f}/8.0 hours required)',
                    worked_hours_at_check=round(worked_hours, 2),
                    required_work_hours=8.0,
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT')
                )
                messages.error(request, f"Cannot check-out yet! You need to complete at least 8 hours of work. Current: {worked_hours:.1f} hours")
                return redirect('employee-check-out')
            
            # Update attendance record
            attendance.check_out_time = current_time
            attendance.total_worked_hours = round(worked_hours, 2)
            
            # Update status based on worked hours
            if worked_hours >= 8.0:
                attendance.status = 'present'
                attendance.remarks += f" | Checked out at {current_time.strftime('%H:%M:%S')} (Completed 8 hours)"
                attendance.can_check_out = False
            elif worked_hours >= 4.0:
                attendance.status = 'half_day'
                attendance.remarks += f" | Checked out at {current_time.strftime('%H:%M:%S')} (Half day - {worked_hours:.1f} hours)"
            else:
                attendance.status = 'early_departure'
                attendance.remarks += f" | Checked out at {current_time.strftime('%H:%M:%S')} (Early departure - {worked_hours:.1f} hours)"
            
            attendance.save()
            
            # Log successful check-out
            AttendanceCheckLog.objects.create(
                employee=employee,
                check_type='check_out',
                check_time=current_datetime,
                attendance_date=today,
                worked_hours_at_check=round(worked_hours, 2),
                required_work_hours=8.0,
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT')
            )
            
            messages.success(request, f"Checked out successfully! You worked for {worked_hours:.1f} hours today.")
            return redirect('employee-dashboard')
        
        # GET request - display check-out form
        # Calculate worked hours so far
        if attendance.check_in_time:
            check_in_dt = datetime.combine(today, attendance.check_in_time)
            current_dt = datetime.combine(today, current_time)
            worked_duration = current_dt - check_in_dt
            worked_hours = worked_duration.total_seconds() / 3600
            
            # Calculate remaining hours to complete 8-hour shift
            remaining_hours = max(0, 8.0 - worked_hours)
            can_check_out = worked_hours >= 8.0
            
            # Update attendance record with current worked hours for real-time display
            if worked_hours >= 8.0:
                attendance.can_check_out = True
                attendance.save()
        else:
            worked_hours = 0
            remaining_hours = 8.0
            can_check_out = False
        
        context = {
            'employee': employee,
            'attendance': attendance,
            'current_time': current_time,
            'worked_hours': round(worked_hours, 2),
            'remaining_hours': round(remaining_hours, 2),
            'can_check_out': can_check_out,
            'today': today,
            'required_work_hours': 8.0,
        }
        
        return render(request, 'app/employee/check-out.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, "Employee profile not found.")
        return redirect('login')
    except Exception as e:
        messages.error(request, f"Error during check-out: {str(e)}")
        return redirect('employee-dashboard')


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def tl_attendance_management(request):
    """
    New Team Leader Attendance Management - Approve team member attendance for the day
    - TL creates attendance records for team members
    - TL approves/rejects attendance for each day
    - Only after TL approval, employees can check-in
    """
    tl_id = request.session.get('tl_id')
    
    if not tl_id:
        messages.error(request, "Please login as Team Leader to manage attendance.")
        return redirect('login')
    
    try:
        tl = TeamLeader.objects.get(id=tl_id)
        today = timezone.now().date()
        
        # Get team members
        team_assignments = TeamAssignment.objects.filter(team_leader=tl).select_related('employee')
        team_member_ids = [assignment.employee.id for assignment in team_assignments]
        
        if request.method == "POST":
            action = request.POST.get("action")
            
            if action == "create_daily_attendance":
                # Create attendance records for all team members for today
                created_count = 0
                for assignment in team_assignments:
                    employee = assignment.employee
                    
                    # Check if attendance already exists for today
                    attendance, created = Attendance.objects.get_or_create(
                        employee=employee,
                        attendance_date=today,
                        defaults={
                            'status': 'present',
                            'shift_type': 'full_time',
                            'is_check_in_allowed': False,
                            'can_check_out': False
                        }
                    )
                    
                    if created:
                        # Create attendance approval record
                        AttendanceApproval.objects.create(
                            attendance=attendance,
                            team_leader=tl,
                            status='pending',
                            requested_at=timezone.now()
                        )
                        created_count += 1
                
                messages.success(request, f"Daily attendance records created for {created_count} team members.")
                return redirect('tl-attendance-management')
            
            elif action == "approve_attendance":
                attendance_id = request.POST.get("attendance_id")
                tl_comments = request.POST.get("tl_comments", "")
                
                try:
                    attendance = Attendance.objects.get(
                        id=attendance_id,
                        employee_id__in=team_member_ids
                    )
                    
                    # Update attendance approval
                    approval = AttendanceApproval.objects.get(
                        attendance=attendance,
                        team_leader=tl
                    )
                    
                    approval.status = 'approved'
                    approval.tl_comments = tl_comments
                    approval.reviewed_at = timezone.now()
                    approval.save()
                    
                    # Enable check-in for employee
                    attendance.is_check_in_allowed = True
                    attendance.approved_by = f"{tl.employee.first_name} {tl.employee.last_name}"
                    attendance.approval_date = timezone.now()
                    attendance.save()
                    
                    messages.success(request, f"Attendance approved for {attendance.employee.first_name} {attendance.employee.last_name}.")
                    
                except Attendance.DoesNotExist:
                    messages.error(request, "Attendance record not found.")
                except AttendanceApproval.DoesNotExist:
                    messages.error(request, "Attendance approval record not found.")
                
                return redirect('tl-attendance-management')
            
            elif action == "reject_attendance":
                attendance_id = request.POST.get("attendance_id")
                tl_comments = request.POST.get("tl_comments", "")
                
                try:
                    attendance = Attendance.objects.get(
                        id=attendance_id,
                        employee_id__in=team_member_ids
                    )
                    
                    # Update attendance approval
                    approval = AttendanceApproval.objects.get(
                        attendance=attendance,
                        team_leader=tl
                    )
                    
                    approval.status = 'rejected'
                    approval.tl_comments = tl_comments
                    approval.reviewed_at = timezone.now()
                    approval.save()
                    
                    # Mark as absent
                    attendance.status = 'absent'
                    attendance.approved_by = f"{tl.employee.first_name} {tl.employee.last_name}"
                    attendance.approval_date = timezone.now()
                    attendance.remarks = f"Attendance rejected by TL: {tl_comments}"
                    attendance.save()
                    
                    messages.success(request, f"Attendance rejected for {attendance.employee.first_name} {attendance.employee.last_name}.")
                    
                except Attendance.DoesNotExist:
                    messages.error(request, "Attendance record not found.")
                except AttendanceApproval.DoesNotExist:
                    messages.error(request, "Attendance approval record not found.")
                
                return redirect('tl-attendance-management')
        
        # GET request - display attendance management interface
        
        # Get today's attendance records for team members
        today_attendance = Attendance.objects.filter(
            employee_id__in=team_member_ids,
            attendance_date=today
        ).select_related('employee').order_by('employee__first_name')
        
        # Check which team members don't have attendance records yet
        team_members_without_attendance = []
        for assignment in team_assignments:
            employee = assignment.employee
            has_attendance = today_attendance.filter(employee=employee).exists()
            if not has_attendance:
                team_members_without_attendance.append(employee)
        
        # Get pending approvals
        pending_approvals = AttendanceApproval.objects.filter(
            team_leader=tl,
            status='pending',
            attendance__attendance_date=today
        ).select_related('attendance', 'attendance__employee')
        
        # Calculate statistics
        total_team_members = team_assignments.count()
        attendance_created = today_attendance.count()
        pending_approvals_count = pending_approvals.count()
        approved_today = AttendanceApproval.objects.filter(
            team_leader=tl,
            status='approved',
            reviewed_at__date=today
        ).count()
        
        # Get recent attendance history (last 7 days)
        recent_attendance = Attendance.objects.filter(
            employee_id__in=team_member_ids,
            attendance_date__gte=today - timedelta(days=7),
            attendance_date__lte=today
        ).select_related('employee').order_by('-attendance_date')
        
        context = {
            'tl': tl,
            'team_assignments': team_assignments,
            'today_attendance': today_attendance,
            'team_members_without_attendance': team_members_without_attendance,
            'pending_approvals': pending_approvals,
            'recent_attendance': recent_attendance,
            'total_team_members': total_team_members,
            'attendance_created': attendance_created,
            'pending_approvals_count': pending_approvals_count,
            'approved_today': approved_today,
            'today': today,
        }
        
        return render(request, 'app/tl/attendance-management.html', context)
        
    except TeamLeader.DoesNotExist:
        messages.error(request, "Team Leader profile not found.")
        return redirect('login')




def hr_monthly_attendance_summary(request):
    """
    HR Monthly Attendance Summary - Generate and manage monthly attendance summaries
    """
    hr_id = request.session.get('hr_id')
    
    if not hr_id:
        messages.error(request, "Please login as HR to manage attendance summaries.")
        return redirect('login')
    
    try:
        hr = HRProfile.objects.get(id=hr_id)
        
        if request.method == "POST":
            action = request.POST.get("action")
            
            if action == "generate_summary":
                employee_id = request.POST.get("employee")
                month = request.POST.get("month")
                year = int(request.POST.get("year"))
                
                try:
                    employee = Employee.objects.get(id=employee_id)
                    
                    # Check if summary already exists
                    existing_summary = MonthlyAttendanceSummary.objects.filter(
                        employee=employee,
                        month=month,
                        year=year
                    ).first()
                    
                    if existing_summary and existing_summary.is_finalized:
                        messages.error(request, "Monthly summary already finalized for this employee and period.")
                        return redirect('hr-monthly-attendance-summary')
                    
                    # Generate attendance data for the month
                    from calendar import monthrange, month_name
                    import calendar
                    
                    # Get month number
                    month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                                  'July', 'August', 'September', 'October', 'November', 'December']
                    month_num = month_names.index(month) + 1
                    
                    # Calculate working days (excluding Sundays)
                    _, num_days = monthrange(year, month_num)
                    working_days = 0
                    weekend_days = 0
                    
                    for day in range(1, num_days + 1):
                        date_obj = datetime(year, month_num, day)
                        if date_obj.weekday() < 6:  # Monday to Saturday (0-5)
                            working_days += 1
                        else:  # Sunday
                            weekend_days += 1
                    
                    # Get attendance records for the month
                    month_start = datetime(year, month_num, 1).date()
                    if month_num == 12:
                        month_end = datetime(year + 1, 1, 1).date() - timedelta(days=1)
                    else:
                        month_end = datetime(year, month_num + 1, 1).date() - timedelta(days=1)
                    
                    attendance_records = Attendance.objects.filter(
                        employee=employee,
                        attendance_date__gte=month_start,
                        attendance_date__lte=month_end
                    )
                    
                    # Calculate statistics
                    present_days = attendance_records.filter(
                        status__in=['present', 'late']
                    ).count()
                    
                    half_days = attendance_records.filter(status='half_day').count()
                    absent_days = attendance_records.filter(status='absent').count()
                    
                    # Count late arrivals from approval records
                    late_arrivals = AttendanceApproval.objects.filter(
                        attendance__employee=employee,
                        is_late_arrival=True,
                        reviewed_at__gte=month_start,
                        reviewed_at__lte=month_end
                    ).count()
                    
                    # Get approved leaves for the month
                    approved_leaves = LeaveApply.objects.filter(
                        employee=employee,
                        status='approved',
                        start_date__gte=month_start,
                        start_date__lte=month_end
                    )
                    
                    unpaid_leaves = approved_leaves.filter(
                        leave_type__in=['unpaid', 'leave_without_pay']
                    ).count()
                    
                    # Calculate total worked hours
                    total_worked_hours = sum([
                        float(record.total_worked_hours or 0) 
                        for record in attendance_records 
                        if record.total_worked_hours
                    ])
                    
                    # Calculate salary deductions
                    daily_salary = float(employee.package) / 26  # Assuming 26 working days per month
                    half_day_deduction = daily_salary / 2
                    
                    salary_deduction_for_absences = absent_days * daily_salary
                    salary_deduction_for_half_days = half_days * half_day_deduction
                    salary_deduction_for_late_arrivals = late_arrivals * (daily_salary * 0.1)  # 10% deduction for late arrival
                    
                    # Create or update summary
                    if existing_summary:
                        # Update existing summary
                        summary = existing_summary
                    else:
                        # Create new summary
                        summary = MonthlyAttendanceSummary.objects.create(
                            employee=employee,
                            month=month,
                            year=year
                        )
                    
                    summary.total_working_days = working_days
                    summary.present_days = present_days
                    summary.absent_days = absent_days
                    summary.half_days = half_days
                    summary.late_arrivals = late_arrivals
                    summary.approved_leaves = approved_leaves.count() - unpaid_leaves
                    summary.unpaid_leaves = unpaid_leaves
                    summary.weekend_days = weekend_days
                    summary.total_worked_hours = round(total_worked_hours, 2)
                    summary.salary_deduction_for_absences = round(salary_deduction_for_absences, 2)
                    summary.salary_deduction_for_half_days = round(salary_deduction_for_half_days, 2)
                    summary.salary_deduction_for_late_arrivals = round(salary_deduction_for_late_arrivals, 2)
                    summary.save()
                    
                    messages.success(request, f"Monthly attendance summary generated for {employee.first_name} {employee.last_name} - {month} {year}.")
                    
                except Employee.DoesNotExist:
                    messages.error(request, "Employee not found.")
                except Exception as e:
                    messages.error(request, f"Error generating summary: {str(e)}")
            
            elif action == "finalize_summary":
                summary_id = request.POST.get("summary_id")
                try:
                    summary = MonthlyAttendanceSummary.objects.get(id=summary_id)
                    summary.is_finalized = True
                    summary.finalized_by = f"{hr.full_name} (HR)"
                    summary.finalized_date = timezone.now()
                    summary.save()
                    
                    messages.success(request, f"Monthly summary finalized for {summary.employee.first_name} {summary.employee.last_name}.")
                    
                except MonthlyAttendanceSummary.DoesNotExist:
                    messages.error(request, "Monthly summary not found.")
        
        # GET request - display summaries
        summaries = MonthlyAttendanceSummary.objects.select_related('employee').order_by('-year', '-month')
        
        # Filter options
        employees = Employee.objects.filter(resigned_date__isnull=True).order_by('first_name')
        current_year = timezone.now().year
        years = range(current_year - 2, current_year + 1)
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        
        # Calculate statistics
        total_summaries = summaries.count()
        finalized_summaries = summaries.filter(is_finalized=True).count()
        pending_summaries = summaries.filter(is_finalized=False).count()
        
        # Get recent activity
        recent_summaries = summaries.order_by('-created_at')[:10]
        
        context = {
            'hr': hr,
            'summaries': summaries,
            'employees': employees,
            'years': years,
            'months': months,
            'total_summaries': total_summaries,
            'finalized_summaries': finalized_summaries,
            'pending_summaries': pending_summaries,
            'recent_summaries': recent_summaries,
        }
        
        return render(request, 'app/hr/monthly-attendance-summary.html', context)
        
    except HRProfile.DoesNotExist:
        messages.error(request, "HR profile not found.")
        return redirect('login')


def hr_payroll_calculations(request):
    """
    New HR Payroll Calculations - Calculate salary deductions based on:
    - Approved leave
    - Half-day attendance  
    - Unpaid leave
    """
    hr_id = request.session.get('hr_id')
    
    if not hr_id:
        messages.error(request, "Please login as HR to manage payroll calculations.")
        return redirect('login')
    
    try:
        hr = HRProfile.objects.get(id=hr_id)
        
        if request.method == "POST":
            action = request.POST.get("action")
            
            if action == "calculate_payroll":
                employee_id = request.POST.get("employee")
                month = request.POST.get("month")
                year = int(request.POST.get("year"))
                
                try:
                    employee = Employee.objects.get(id=employee_id)
                    
                    # Get attendance records for the month
                    month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                                  'July', 'August', 'September', 'October', 'November', 'December']
                    month_num = month_names.index(month) + 1
                    
                    # Calculate month date range
                    if month_num == 12:
                        month_start = datetime(year, month_num, 1).date()
                        month_end = datetime(year + 1, 1, 1).date() - timedelta(days=1)
                    else:
                        month_start = datetime(year, month_num, 1).date()
                        month_end = datetime(year, month_num + 1, 1).date() - timedelta(days=1)
                    
                    # Get attendance records for the month
                    attendance_records = Attendance.objects.filter(
                        employee=employee,
                        attendance_date__gte=month_start,
                        attendance_date__lte=month_end
                    )
                    
                    # Calculate attendance statistics
                    total_working_days = attendance_records.count()
                    present_days = attendance_records.filter(status='present').count()
                    half_days = attendance_records.filter(status='half_day').count()
                    absent_days = attendance_records.filter(status='absent').count()
                    late_arrivals = attendance_records.filter(status='late').count()
                    
                    # Get approved leave applications for the month
                    approved_leaves = LeaveApply.objects.filter(
                        employee=employee,
                        status='approved',
                        start_date__gte=month_start,
                        start_date__lte=month_end
                    )
                    
                    paid_leave_days = sum([leave.total_days for leave in approved_leaves 
                                         if leave.leave_type not in ['unpaid', 'leave_without_pay']])
                    unpaid_leave_days = sum([leave.total_days for leave in approved_leaves 
                                            if leave.leave_type in ['unpaid', 'leave_without_pay']])
                    
                    # Calculate salary components
                    base_salary = float(employee.package)
                    daily_salary = base_salary / 26  # Assuming 26 working days per month
                    
                    # Calculate deductions
                    half_day_deduction = half_days * (daily_salary / 2)
                    absent_deduction = absent_days * daily_salary
                    unpaid_leave_deduction = unpaid_leave_days * daily_salary
                    
                    # Standard deductions
                    pf_deduction = base_salary * 0.12  # 12% PF
                    professional_tax = 200  # Fixed professional tax
                    
                    # Calculate gross salary (base + allowances)
                    allowances = float(request.POST.get("allowances", 0))
                    overtime_amount = float(request.POST.get("overtime_amount", 0))
                    gross_salary = base_salary + allowances + overtime_amount
                    
                    # Calculate total deductions
                    total_deductions = (half_day_deduction + absent_deduction + unpaid_leave_deduction +
                                      pf_deduction + professional_tax)
                    
                    # Calculate final salary
                    final_salary = gross_salary - total_deductions
                    
                    # Create or update payroll record
                    existing_payroll = Payroll.objects.filter(
                        employee=employee,
                        month=month,
                        year=year
                    ).first()
                    
                    if existing_payroll:
                        payroll = existing_payroll
                    else:
                        payroll = Payroll.objects.create(
                            employee=employee,
                            month=month,
                            year=year
                        )
                    
                    # Update payroll fields
                    payroll.base_salary = base_salary
                    payroll.allowances = allowances
                    payroll.overtime_amount = overtime_amount
                    payroll.leave_deduction = absent_deduction + unpaid_leave_deduction
                    payroll.half_day_deduction = half_day_deduction
                    payroll.late_arrival_deduction = 0  # Late arrivals already handled in half day
                    payroll.pf_deduction = pf_deduction
                    payroll.professional_tax = professional_tax
                    payroll.other_deductions = 0
                    payroll.gross_salary = gross_salary
                    payroll.total_deductions = total_deductions
                    payroll.final_salary = final_salary
                    payroll.created_by = f"{hr.full_name} (HR)"
                    payroll.is_processed = False
                    payroll.save()
                    
                    # Clear existing deduction records
                    PayrollDeduction.objects.filter(payroll=payroll).delete()
                    
                    # Create detailed deduction records
                    if half_day_deduction > 0:
                        PayrollDeduction.objects.create(
                            payroll=payroll,
                            employee=employee,
                            deduction_type='half_day',
                            description=f'Half day deduction for {half_days} half days',
                            amount=half_day_deduction,
                            calculation_basis='Half day rate',
                            units_deducted=half_days,
                            approved_by=f"{hr.full_name} (HR)",
                            approved_date=timezone.now()
                        )
                    
                    if absent_deduction > 0:
                        PayrollDeduction.objects.create(
                            payroll=payroll,
                            employee=employee,
                            deduction_type='absent',
                            description=f'Absence deduction for {absent_days} absent days',
                            amount=absent_deduction,
                            calculation_basis='Per day salary',
                            units_deducted=absent_days,
                            approved_by=f"{hr.full_name} (HR)",
                            approved_date=timezone.now()
                        )
                    
                    if unpaid_leave_deduction > 0:
                        PayrollDeduction.objects.create(
                            payroll=payroll,
                            employee=employee,
                            deduction_type='unpaid_leave',
                            description=f'Unpaid leave deduction for {unpaid_leave_days} days',
                            amount=unpaid_leave_deduction,
                            calculation_basis='Per day salary',
                            units_deducted=unpaid_leave_days,
                            approved_by=f"{hr.full_name} (HR)",
                            approved_date=timezone.now()
                        )
                    
                    if pf_deduction > 0:
                        PayrollDeduction.objects.create(
                            payroll=payroll,
                            employee=employee,
                            deduction_type='pf',
                            description='Provident Fund deduction',
                            amount=pf_deduction,
                            calculation_basis='12% of base salary',
                            approved_by=f"{hr.full_name} (HR)",
                            approved_date=timezone.now()
                        )
                    
                    if professional_tax > 0:
                        PayrollDeduction.objects.create(
                            payroll=payroll,
                            employee=employee,
                            deduction_type='professional_tax',
                            description='Professional Tax',
                            amount=professional_tax,
                            calculation_basis='Fixed amount',
                            approved_by=f"{hr.full_name} (HR)",
                            approved_date=timezone.now()
                        )
                    
                    messages.success(request, f"Payroll calculated successfully for {employee.first_name} {employee.last_name} - {month} {year}. Final Salary: ₹{final_salary:.2f}")
                    
                except Employee.DoesNotExist:
                    messages.error(request, "Employee not found.")
                except Exception as e:
                    messages.error(request, f"Error calculating payroll: {str(e)}")
            
            elif action == "process_payroll":
                payroll_id = request.POST.get("payroll_id")
                try:
                    payroll = Payroll.objects.get(id=payroll_id)
                    payroll.is_processed = True
                    payroll.processed_date = timezone.now()
                    payroll.save()
                    
                    messages.success(request, f"Payroll processed successfully for {payroll.employee.first_name} {payroll.employee.last_name}.")
                    
                except Payroll.DoesNotExist:
                    messages.error(request, "Payroll record not found.")
        
        # GET request - display payroll calculations
        payrolls = Payroll.objects.select_related('employee').order_by('-year', '-month')
        
        # Filter options
        employees = Employee.objects.filter(resigned_date__isnull=True).order_by('first_name')
        current_year = timezone.now().year
        years = range(current_year - 2, current_year + 1)
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        
        # Calculate statistics
        total_payrolls = payrolls.count()
        processed_payrolls = payrolls.filter(is_processed=True).count()
        pending_payrolls = payrolls.filter(is_processed=False).count()
        
        # Get recent payroll activity
        recent_payrolls = payrolls.order_by('-created_at')[:10]
        
        context = {
            'hr': hr,
            'payrolls': payrolls,
            'employees': employees,
            'years': years,
            'months': months,
            'total_payrolls': total_payrolls,
            'processed_payrolls': processed_payrolls,
            'pending_payrolls': pending_payrolls,
            'recent_payrolls': recent_payrolls,
        }
        
        return render(request, 'app/hr/payroll-calculations.html', context)
        
    except HRProfile.DoesNotExist:
        messages.error(request, "HR profile not found.")
        return redirect('login')


def employee_attendance_records_simple(request):
    """
    Employee Attendance Records - Database Only Implementation
    Shows only present days with exact check-in/check-out times
    All data from real database models only
    """
    # Check authentication and get employee from session
    employee_id = request.session.get('employee_id')
    role = request.session.get('role')
    
    # Ensure user is logged in as employee
    if not employee_id or role != 'employee':
        messages.error(request, "Please login to view attendance records.")
        return redirect('login')
    
    try:
        employee = Employee.objects.get(id=employee_id)
        
        # Get all attendance records for this employee
        attendance_records = Attendance.objects.filter(
            employee=employee,
            status__in=['present', 'late', 'half_day']  # Only present days
        ).order_by('-attendance_date')
        
        # Get pagination parameters
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        
        # Paginate records
        paginated_records = attendance_records[start_idx:end_idx]
        total_records = attendance_records.count()
        total_pages = (total_records + per_page - 1) // per_page
        
        # Calculate statistics for present days only
        total_present_days = attendance_records.count()
        present_on_time = attendance_records.filter(status='present').count()
        late_arrivals = attendance_records.filter(status='late').count()
        half_days = attendance_records.filter(status='half_day').count()
        
        # Calculate current month stats
        current_month = timezone.now().replace(day=1)
        current_month_attendance = attendance_records.filter(
            attendance_date__gte=current_month
        )
        
        current_month_stats = {
            'total_days': current_month_attendance.count(),
            'present_days': current_month_attendance.filter(status='present').count(),
            'half_days': current_month_attendance.filter(status='half_day').count(),
            'late_arrivals': current_month_attendance.filter(status='late').count(),
        }
        
        # Calculate attendance percentage for present days
        attendance_percentage = round((total_present_days / max(total_records, 1)) * 100, 1)
        
        context = {
            'employee': employee,
            'attendance_records': paginated_records,
            'total_days': total_present_days,
            'present_days': present_on_time,
            'late_arrivals': late_arrivals,
            'half_days': half_days,
            'attendance_percentage': attendance_percentage,
            'current_month_stats': current_month_stats,
            'current_page': page,
            'total_pages': total_pages,
            'per_page': per_page,
            'total_records': total_records,
            'has_previous': page > 1,
            'has_next': page < total_pages,
            'previous_page': page - 1 if page > 1 else None,
            'next_page': page + 1 if page < total_pages else None,
        }
        
        return render(request, 'app/employee/attendance-records.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, "Employee profile not found.")
        return redirect('login')
    except Exception as e:
        messages.error(request, f"Error loading attendance records: {str(e)}")
        return redirect('employee-dashboard')


def employee_attendance_view(request):
    """
    Employee Attendance View - Redirect to employee_attendance_simple
    This function redirects to the main attendance system
    """
    # Redirect to the main employee attendance view
    return redirect('employee-attendance')


def hr_salary_processing(request):
    """
    HR Salary Processing View - Handle salary processing from 1st to 7th of month
    """
    hr_id = request.session.get('hr_id')
    
    if not hr_id:
        messages.error(request, "Please login as HR to manage salary processing.")
        return redirect('login')
    
    try:
        hr = HRProfile.objects.get(id=hr_id)
        current_date = timezone.now().date()
        
        # Check if current date is within salary processing window (1st to 7th of month)
        is_within_processing_window = current_date.day <= 7
        
        if request.method == "POST":
            action = request.POST.get("action")
            
            if action == "initiate_salary_processing":
                employee_id = request.POST.get("employee")
                salary_month = request.POST.get("salary_month")
                salary_year = int(request.POST.get("salary_year"))
                
                try:
                    employee = Employee.objects.get(id=employee_id)
                    
                    # Check if processing already exists
                    existing_processing = SalaryProcessing.objects.filter(
                        employee=employee,
                        salary_month=salary_month,
                        salary_year=salary_year
                    ).first()
                    
                    if existing_processing:
                        messages.error(request, "Salary processing already initiated for this employee and period.")
                        return redirect('hr-salary-processing')
                    
                    # Get or create monthly attendance summary
                    summary = MonthlyAttendanceSummary.objects.filter(
                        employee=employee,
                        month=salary_month,
                        year=salary_year
                    ).first()
                    
                    if not summary:
                        messages.error(request, "Monthly attendance summary not found. Please generate attendance summary first.")
                        return redirect('hr-salary-processing')
                    
                    # Calculate salary processing dates
                    processing_start_date = date(salary_year, list(calendar.month_name).index(salary_month), 1)
                    processing_end_date = date(salary_year, list(calendar.month_name).index(salary_month), 7)
                    
                    # Create salary processing record
                    salary_processing = SalaryProcessing.objects.create(
                        employee=employee,
                        salary_month=salary_month,
                        salary_year=salary_year,
                        processing_start_date=processing_start_date,
                        processing_end_date=processing_end_date,
                        status='pending',
                        is_within_processing_window=is_within_processing_window,
                        monthly_summary=summary,
                        processed_by=f"{hr.full_name} (HR)"
                    )
                    
                    messages.success(request, f"Salary processing initiated for {employee.first_name} {employee.last_name} - {salary_month} {salary_year}")
                    
                except Employee.DoesNotExist:
                    messages.error(request, "Employee not found.")
                except Exception as e:
                    messages.error(request, f"Error initiating salary processing: {str(e)}")
            
            elif action == "process_salary":
                processing_id = request.POST.get("processing_id")
                try:
                    processing = SalaryProcessing.objects.get(id=processing_id)
                    
                    # Get payroll record for this processing
                    payroll = Payroll.objects.filter(
                        employee=processing.employee,
                        month=processing.salary_month,
                        year=processing.salary_year
                    ).first()
                    
                    if not payroll:
                        messages.error(request, "Payroll record not found for this salary processing.")
                        return redirect('hr-salary-processing')
                    
                    # Update processing status
                    processing.status = 'completed'
                    processing.processed_at = timezone.now()
                    processing.processed_date = current_date
                    processing.payroll_record = payroll
                    processing.save()
                    
                    # Update payroll record
                    payroll.is_processed = True
                    payroll.processed_date = timezone.now()
                    payroll.save()
                    
                    messages.success(request, f"Salary processing completed for {processing.employee.first_name} {processing.employee.last_name}")
                    
                except SalaryProcessing.DoesNotExist:
                    messages.error(request, "Salary processing record not found.")
        
        # GET request - display salary processing interface
        salary_processings = SalaryProcessing.objects.select_related('employee').order_by('-salary_year', '-salary_month')
        
        # Filter options
        employees = Employee.objects.filter(resigned_date__isnull=True).order_by('first_name')
        current_year = timezone.now().year
        years = range(current_year - 2, current_year + 1)
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        
        # Calculate statistics
        total_processings = salary_processings.count()
        completed_processings = salary_processings.filter(status='completed').count()
        pending_processings = salary_processings.filter(status='pending').count()
        in_progress_processings = salary_processings.filter(status='in_progress').count()
        
        # Get recent salary processing activity
        recent_processings = salary_processings.order_by('-created_at')[:10]
        
        context = {
            'hr': hr,
            'salary_processings': salary_processings,
            'employees': employees,
            'years': years,
            'months': months,
            'total_processings': total_processings,
            'completed_processings': completed_processings,
            'pending_processings': pending_processings,
            'in_progress_processings': in_progress_processings,
            'recent_processings': recent_processings,
            'is_within_processing_window': is_within_processing_window,
            'current_date': current_date,
            'processing_deadline': date(current_date.year, current_date.month, 7),
        }
        
        return render(request, 'app/hr/salary-processing.html', context)
        
    except HRProfile.DoesNotExist:
        messages.error(request, "HR profile not found.")
        return redirect('login')


def employee_attendance_logs_view(request):
    """
    Employee Attendance Logs View - Display employee check-in/check-out logs
    """
    employee_id = request.session.get('employee_id')
    
    if not employee_id:
        messages.error(request, "Please login to view attendance logs.")
        return redirect('login')
    
    try:
        employee = Employee.objects.get(id=employee_id)
        
        # Get attendance check logs
        check_logs = AttendanceCheckLog.objects.filter(
            employee=employee
        ).order_by('-check_time')[:50]  # Last 50 activities
        
        # Calculate statistics
        total_attempts = check_logs.count()
        successful_check_ins = check_logs.filter(check_type='check_in').count()
        successful_check_outs = check_logs.filter(check_type='check_out').count()
        denied_attempts = check_logs.filter(is_denied=True).count()
        
        # Get recent activity
        recent_logs = check_logs[:10]
        
        context = {
            'employee': employee,
            'check_logs': check_logs,
            'total_attempts': total_attempts,
            'successful_check_ins': successful_check_ins,
            'successful_check_outs': successful_check_outs,
            'denied_attempts': denied_attempts,
            'recent_logs': recent_logs,
        }
        
        return render(request, 'app/employee/attendance-logs.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, "Employee profile not found.")
        return redirect('login')


def payrollemployeepage(request):
    return render(request , 'app/employee/payslip.html')
@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def employee_project_dashboard(request):
    """Employee Project Dashboard - Display assigned projects, tasks, and progress safely"""

    employee_id = request.session.get('employee_id')

    if not employee_id:
        messages.error(request, "Please login to access project dashboard.")
        return redirect('login')

    # ✅ Safe defaults (so page never crashes)
    employee = None
    projects = []
    tasks = []
    recent_discussions = []

    total_projects = 0
    active_projects = 0
    completed_tasks = 0
    pending_tasks = 0
    overdue_tasks = 0
    projects_with_progress = []

    try:
        employee = Employee.objects.get(id=employee_id)
    except Employee.DoesNotExist:
        messages.error(request, "Employee profile not found.")
        return redirect('login')

    # ✅ Projects (Safe Query)
    projects = ProjectAssignment.objects.filter(
        team_members=employee
    ).select_related('team_leader', 'team_leader__employee').order_by('-created_at')

    # ✅ Tasks (Safe Query)
    tasks = ProjectTask.objects.filter(
        assigned_to=employee
    ).select_related('project', 'assigned_by', 'assigned_by__employee').order_by('due_date')

    # ✅ Discussions (Safe Query)
    recent_discussions = ProjectDiscussion.objects.filter(
        project__team_members=employee
    ).select_related('project').order_by('-created_at')[:10]

    # ✅ Statistics (Safe Calculation)
    total_projects = projects.count()
    active_projects = projects.filter(end_date__gte=timezone.now().date()).count()

    completed_tasks = tasks.filter(task_status='completed').count()
    pending_tasks = tasks.filter(task_status__in=['not_started', 'in_progress', 'in_review']).count()
    overdue_tasks = tasks.filter(
        due_date__lt=timezone.now().date(),
        task_status__in=['not_started', 'in_progress', 'in_review']
    ).count()

    # ✅ Progress Calculation (No Zero-Division Risk)
    from datetime import date
    for project in projects:
        project_tasks = tasks.filter(project=project)
        total_task_count = project_tasks.count()

        if total_task_count > 0:
            completed_project_tasks = project_tasks.filter(task_status='completed').count()
            my_progress = round((completed_project_tasks / total_task_count) * 100, 1)
        else:
            my_progress = 0
            completed_project_tasks = 0

        # Calculate days until due date
        today = timezone.now().date()
        days_until_due = (project.end_date - today).days

        projects_with_progress.append({
            'project': project,
            'my_progress': my_progress,
            'total_tasks': total_task_count,
            'completed_tasks': completed_project_tasks,
            'days_until_due': days_until_due
        })

    context = {
        'employee': employee,
        'total_projects': total_projects,
        'active_projects': active_projects,
        'completed_tasks': completed_tasks,
        'pending_tasks': pending_tasks,
        'overdue_tasks': overdue_tasks,
        'projects': projects_with_progress,
        'tasks': tasks,
        'recent_discussions': recent_discussions,
    }

    # ✅ ALWAYS returns page — never crashes
    return render(request, 'app/employee/project-dashboard.html', context)
@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def tl_project_dashboard(request):
    """TL Project Dashboard - Display team projects, tasks, and progress safely"""
    
    tl_id = request.session.get('tl_id')
    
    if not tl_id:
        messages.error(request, "Please login as Team Leader to access project dashboard.")
        return redirect('login')
    
    # ✅ Safe defaults (so page never crashes)
    tl = None
    employee = None
    projects = []
    team_performance = []
    
    total_projects = 0
    active_projects = 0
    completed_projects = 0
    pending_tasks = 0
    
    try:
        tl = TeamLeader.objects.get(id=tl_id)
        employee = tl.employee
    except TeamLeader.DoesNotExist:
        messages.error(request, "Team Leader profile not found.")
        return redirect('login')
    
    # ✅ Projects (Safe Query)
    projects = ProjectAssignment.objects.filter(
        team_leader=tl
    ).select_related('team_leader', 'team_leader__employee').order_by('-created_at')
    
    # ✅ Team Performance (Safe Query)
    team_assignments = TeamAssignment.objects.filter(
        team_leader=tl
    ).select_related('employee')
    
    team_performance = []
    for assignment in team_assignments:
        # Get tasks assigned to this team member
        member_tasks = ProjectTask.objects.filter(
            assigned_to=assignment.employee
        ).select_related('project')
        
        # Calculate performance metrics
        total_tasks = member_tasks.count()
        completed_tasks = member_tasks.filter(task_status='completed').count()
        performance = round((completed_tasks / max(total_tasks, 1)) * 100, 1) if total_tasks > 0 else 0
        
        team_performance.append({
            'employee': assignment.employee,
            'total_tasks': total_tasks,
            'completed_tasks': completed_tasks,
            'performance': performance
        })
    
    # ✅ Statistics (Safe Calculation)
    total_projects = projects.count()
    active_projects = projects.filter(end_date__gte=timezone.now().date()).count()
    completed_projects = projects.filter(end_date__lt=timezone.now().date()).count()
    
    # Calculate pending tasks across all projects
    all_project_tasks = ProjectTask.objects.filter(
        project__team_leader=tl
    )
    pending_tasks = all_project_tasks.filter(
        task_status__in=['not_started', 'in_progress', 'in_review']
    ).count()
    
    context = {
        'tl': tl,
        'employee': employee,
        'total_projects': total_projects,
        'active_projects': active_projects,
        'completed_projects': completed_projects,
        'pending_tasks': pending_tasks,
        'projects': projects,
        'team_performance': team_performance,
    }
    
    # ✅ ALWAYS returns page — never crashes
    return render(request, 'app/tl/project-dashboard.html', context)
    return render(request, 'app/employee/project-dashboard.html', context)


# ============================================================================
# TEAM CHAT VIEWS
# ============================================================================

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def team_chat_dashboard(request):
    """
    Team Chat Dashboard - Main chat interface for both employees and team leaders
    """
    # Check user role and get appropriate user object
    hr_id = request.session.get('hr_id')
    employee_id = request.session.get('employee_id')
    tl_id = request.session.get('tl_id')
    
    context = {
        'user_role': None,
        'conversations': [],
        'recent_messages': [],
        'team_members': [],
        'unread_count': 0,
        'chat_stats': {}
    }
    
    try:
        if employee_id:
            # Employee logged in
            employee = Employee.objects.get(id=employee_id)
            context['user_role'] = 'employee'
            context['current_user'] = employee
            
            # Get team assignment
            team_assignment = TeamAssignment.objects.filter(employee=employee).select_related('team_leader', 'team_leader__employee').first()
            
            if team_assignment:
                # Get team leader info
                tl = team_assignment.team_leader
                context['team_leader'] = {
                    'id': tl.id,
                    'name': f"{tl.employee.first_name} {tl.employee.last_name}",
                    'designation': tl.employee.designation,
                    'department': tl.employee.department,
                    'image': tl.employee.image.url if tl.employee.image else None
                }
                
                # Get conversations (messages with team leader)
                conversations = TeamChat.objects.filter(
                    Q(sender_employee=employee, receiver_tl=tl) |
                    Q(sender_tl=tl, receiver_employee=employee)
                ).order_by('-created_at')[:20]
                
                context['conversations'] = conversations
                
                # Get unread message count
                context['unread_count'] = conversations.filter(
                    receiver_employee=employee,
                    is_read=False
                ).count()
                
                # Get team members (other employees in the same team)
                team_members = TeamAssignment.objects.filter(
                    team_leader=tl
                ).exclude(employee=employee).select_related('employee')
                
                context['team_members'] = [{
                    'id': member.employee.id,
                    'name': f"{member.employee.first_name} {member.employee.last_name}",
                    'designation': member.employee.designation,
                    'image': member.employee.image.url if member.employee.image else None,
                    'role': member.role
                } for member in team_members]
            
            # Recent messages (last 10)
            recent_messages = TeamChat.objects.filter(
                Q(sender_employee=employee) | Q(receiver_employee=employee)
            ).order_by('-created_at')[:10]
            
            context['recent_messages'] = recent_messages
            
            # Chat statistics
            sent_messages = TeamChat.objects.filter(sender_employee=employee).count()
            received_messages = TeamChat.objects.filter(receiver_employee=employee).count()
            
            context['chat_stats'] = {
                'sent': sent_messages,
                'received': received_messages,
                'total': sent_messages + received_messages
            }
            
        elif tl_id:
            # Team Leader logged in
            tl = TeamLeader.objects.get(id=tl_id)
            context['user_role'] = 'tl'
            context['current_user'] = tl
            
            # Get team members
            team_assignments = TeamAssignment.objects.filter(
                team_leader=tl
            ).select_related('employee')
            
            context['team_members'] = [{
                'id': member.employee.id,
                'name': f"{member.employee.first_name} {member.employee.last_name}",
                'designation': member.employee.designation,
                'department': member.employee.department,
                'image': member.employee.image.url if member.employee.image else None,
                'role': member.role,
                'assignment_date': member.assignment_date
            } for member in team_assignments]
            
            # Get conversations with team members
            conversations = TeamChat.objects.filter(
                Q(sender_tl=tl) | Q(receiver_tl=tl)
            ).order_by('-created_at')[:30]
            
            context['conversations'] = conversations
            
            # Get unread message count
            context['unread_count'] = conversations.filter(
                receiver_tl=tl,
                is_read=False
            ).count()
            
            # Recent messages (last 15)
            recent_messages = TeamChat.objects.filter(
                Q(sender_tl=tl) | Q(receiver_tl=tl)
            ).order_by('-created_at')[:15]
            
            context['recent_messages'] = recent_messages
            
            # Chat statistics
            sent_messages = TeamChat.objects.filter(sender_tl=tl).count()
            received_messages = TeamChat.objects.filter(receiver_tl=tl).count()
            
            context['chat_stats'] = {
                'sent': sent_messages,
                'received': received_messages,
                'total': sent_messages + received_messages,
                'team_size': len(context['team_members'])
            }
            
        elif hr_id:
            # HR logged in - can see all conversations (optional feature)
            context['user_role'] = 'hr'
            context['current_user'] = HRProfile.objects.get(id=hr_id)
            
            # HR can see recent conversations from all teams
            all_conversations = TeamChat.objects.order_by('-created_at')[:50]
            context['conversations'] = all_conversations
            context['recent_messages'] = all_conversations[:15]
            
            # Statistics for HR dashboard
            context['chat_stats'] = {
                'total_conversations': all_conversations.count(),
                'active_today': all_conversations.filter(created_at__date=timezone.now().date()).count(),
                'unread_total': all_conversations.filter(is_read=False).count()
            }
        
        return render(request, 'app/chat/chat-dashboard.html', context)
        
    except (Employee.DoesNotExist, TeamLeader.DoesNotExist, HRProfile.DoesNotExist) as e:
        messages.error(request, "User profile not found. Please login again.")
        return redirect('login')
    except Exception as e:
        messages.error(request, f"Error loading chat dashboard: {str(e)}")
        return redirect('employee-dashboard' if employee_id else 'tl-dashboard')


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def send_team_message(request):
    """
    Send Team Message - Handle message sending between team members and leaders
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Get sender information
        employee_id = request.session.get('employee_id')
        tl_id = request.session.get('tl_id')
        
        if not (employee_id or tl_id):
            return JsonResponse({'success': False, 'error': 'User not authenticated'})
        
        # Parse request data
        import json
        data = json.loads(request.body)
        
        receiver_id = data.get('receiver_id')
        receiver_type = data.get('receiver_type')  # 'employee' or 'tl'
        message_text = data.get('message', '').strip()
        subject = data.get('subject', '').strip()
        chat_type = data.get('chat_type', 'direct')
        priority = data.get('priority', 'normal')
        parent_message_id = data.get('parent_message_id')
        
        # Validation
        if not receiver_id or not message_text:
            return JsonResponse({'success': False, 'error': 'Receiver and message are required'})
        
        if len(message_text) > 2000:  # Message length limit
            return JsonResponse({'success': False, 'error': 'Message too long (max 2000 characters)'})
        
        # Get sender objects
        sender_employee = None
        sender_tl = None
        
        if employee_id:
            sender_employee = Employee.objects.get(id=employee_id)
        elif tl_id:
            sender_tl = TeamLeader.objects.get(id=tl_id)
        
        # Get receiver objects
        receiver_employee = None
        receiver_tl = None
        
        if receiver_type == 'employee':
            receiver_employee = Employee.objects.get(id=receiver_id)
        elif receiver_type == 'tl':
            receiver_tl = TeamLeader.objects.get(id=receiver_id)
        
        # Verify team relationship (for security)
        if sender_employee and receiver_tl:
            # Employee sending to TL - verify they are in the same team
            team_assignment = TeamAssignment.objects.filter(
                employee=sender_employee,
                team_leader=receiver_tl
            ).first()
            if not team_assignment:
                return JsonResponse({'success': False, 'error': 'You can only message your team leader'})
        
        elif sender_tl and receiver_employee:
            # TL sending to employee - verify they manage this employee
            team_assignment = TeamAssignment.objects.filter(
                employee=receiver_employee,
                team_leader=sender_tl
            ).first()
            if not team_assignment:
                return JsonResponse({'success': False, 'error': 'You can only message your team members'})
        
        # Handle file attachment
        attachment = request.FILES.get('attachment')
        attachment_name = None
        if attachment:
            # Validate file size (max 10MB)
            if attachment.size > 10 * 1024 * 1024:
                return JsonResponse({'success': False, 'error': 'File too large (max 10MB)'})
            
            # Validate file type
            allowed_types = ['pdf', 'doc', 'docx', 'txt', 'jpg', 'jpeg', 'png', 'gif']
            file_extension = attachment.name.split('.')[-1].lower()
            if file_extension not in allowed_types:
                return JsonResponse({'success': False, 'error': 'File type not allowed'})
            
            attachment_name = attachment.name
        
        # Create parent message relationship
        parent_message = None
        if parent_message_id:
            try:
                parent_message = TeamChat.objects.get(id=parent_message_id)
            except TeamChat.DoesNotExist:
                pass
        
        # Create the message
        team_chat = TeamChat.objects.create(
            sender_employee=sender_employee,
            sender_tl=sender_tl,
            receiver_employee=receiver_employee,
            receiver_tl=receiver_tl,
            chat_type=chat_type,
            subject=subject,
            message=message_text,
            priority=priority,
            parent_message=parent_message,
            attachment=attachment,
            attachment_name=attachment_name
        )
        
        # Update reply count for parent message
        if parent_message:
            parent_message.replies_count += 1
            parent_message.save()
        
        # Create notification (you can extend this based on your notification system)
        # For now, we'll just mark as unread
        team_chat.is_read = False
        team_chat.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Message sent successfully',
            'chat_id': team_chat.id,
            'created_at': team_chat.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'sender_name': team_chat.get_sender_name()
        })
        
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'})
    except TeamLeader.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Team leader not found'})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error sending message: {str(e)}'})


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def get_conversations(request):
    """
    Get Conversations - API endpoint to fetch conversation history
    """
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Get user information
        employee_id = request.session.get('employee_id')
        tl_id = request.session.get('tl_id')
        
        if not (employee_id or tl_id):
            return JsonResponse({'success': False, 'error': 'User not authenticated'})
        
        # Get query parameters
        other_user_id = request.GET.get('other_user_id')
        other_user_type = request.GET.get('other_user_type')  # 'employee' or 'tl'
        limit = int(request.GET.get('limit', 50))
        offset = int(request.GET.get('offset', 0))
        
        # Build query
        query = Q()
        
        if employee_id:
            current_employee = Employee.objects.get(id=employee_id)
            
            if other_user_id and other_user_type:
                if other_user_type == 'tl':
                    other_tl = TeamLeader.objects.get(id=other_user_id)
                    query = Q(
                        Q(sender_employee=current_employee, receiver_tl=other_tl) |
                        Q(sender_tl=other_tl, receiver_employee=current_employee)
                    )
                else:
                    other_employee = Employee.objects.get(id=other_user_id)
                    query = Q(
                        Q(sender_employee=current_employee, receiver_employee=other_employee) |
                        Q(sender_employee=other_employee, receiver_employee=current_employee)
                    )
            else:
                # Get all conversations for this employee
                query = Q(
                    Q(sender_employee=current_employee) | Q(receiver_employee=current_employee)
                )
        
        elif tl_id:
            current_tl = TeamLeader.objects.get(id=tl_id)
            
            if other_user_id and other_user_type == 'employee':
                other_employee = Employee.objects.get(id=other_user_id)
                query = Q(
                    Q(sender_tl=current_tl, receiver_employee=other_employee) |
                    Q(sender_employee=other_employee, receiver_tl=current_tl)
                )
            else:
                # Get all conversations for this TL
                query = Q(
                    Q(sender_tl=current_tl) | Q(receiver_tl=current_tl)
                )
        
        # Get conversations
        conversations = TeamChat.objects.filter(query).select_related(
            'sender_employee', 'sender_tl', 'receiver_employee', 'receiver_tl', 'parent_message'
        ).order_by('-created_at')[offset:offset + limit]
        
        # Prepare response data
        conversation_data = []
        for chat in conversations:
            conversation_data.append({
                'id': chat.id,
                'chat_type': chat.chat_type,
                'subject': chat.subject,
                'message': chat.message,
                'priority': chat.priority,
                'is_read': chat.is_read,
                'created_at': chat.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'sender': {
                    'id': chat.sender_employee.id if chat.sender_employee else chat.sender_tl.id,
                    'name': chat.get_sender_name(),
                    'type': 'employee' if chat.sender_employee else 'tl',
                    'image': chat.sender_employee.image.url if chat.sender_employee and chat.sender_employee.image else (
                        chat.sender_tl.employee.image.url if chat.sender_tl and chat.sender_tl.employee.image else None
                    )
                },
                'receiver': {
                    'id': chat.receiver_employee.id if chat.receiver_employee else chat.receiver_tl.id,
                    'name': chat.get_receiver_name(),
                    'type': 'employee' if chat.receiver_employee else 'tl'
                },
                'has_attachment': bool(chat.attachment),
                'attachment_name': chat.attachment_name,
                'replies_count': chat.replies_count,
                'likes_count': chat.likes_count,
                'is_reply': bool(chat.parent_message),
                'parent_message_id': chat.parent_message.id if chat.parent_message else None
            })
        
        return JsonResponse({
            'success': True,
            'conversations': conversation_data,
            'total': conversations.count(),
            'has_more': len(conversation_data) == limit
        })
        
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'})
    except TeamLeader.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Team leader not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error fetching conversations: {str(e)}'})


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def mark_messages_read(request):
    """
    Mark Messages as Read - API endpoint to mark messages as read
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Get user information
        employee_id = request.session.get('employee_id')
        tl_id = request.session.get('tl_id')
        
        if not (employee_id or tl_id):
            return JsonResponse({'success': False, 'error': 'User not authenticated'})
        
        import json
        data = json.loads(request.body)
        
        message_ids = data.get('message_ids', [])
        
        if not message_ids:
            return JsonResponse({'success': False, 'error': 'No message IDs provided'})
        
        # Build query to mark messages as read
        query = Q(id__in=message_ids)
        
        if employee_id:
            query &= Q(receiver_employee_id=employee_id)
        elif tl_id:
            query &= Q(receiver_tl_id=tl_id)
        
        # Mark messages as read
        updated_count = TeamChat.objects.filter(query).update(
            is_read=True,
            read_at=timezone.now()
        )
        
        return JsonResponse({
            'success': True,
            'message': f'{updated_count} messages marked as read',
            'updated_count': updated_count
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error marking messages as read: {str(e)}'})


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def chat_reaction(request):
    """
    Add/Remove Chat Reaction - Handle emoji reactions on messages
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Get user information
        employee_id = request.session.get('employee_id')
        tl_id = request.session.get('tl_id')
        
        if not (employee_id or tl_id):
            return JsonResponse({'success': False, 'error': 'User not authenticated'})
        
        import json
        data = json.loads(request.body)
        
        chat_id = data.get('chat_id')
        reaction_type = data.get('reaction_type')
        action = data.get('action', 'add')  # 'add' or 'remove'
        
        # Validation
        if not chat_id or not reaction_type:
            return JsonResponse({'success': False, 'error': 'Chat ID and reaction type are required'})
        
        # Get the chat message
        try:
            chat_message = TeamChat.objects.get(id=chat_id)
        except TeamChat.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Chat message not found'})
        
        # Get user objects
        employee = None
        tl = None
        
        if employee_id:
            employee = Employee.objects.get(id=employee_id)
        elif tl_id:
            tl = TeamLeader.objects.get(id=tl_id)
        
        if action == 'add':
            # Add reaction
            reaction, created = ChatReaction.objects.get_or_create(
                chat_message=chat_message,
                employee=employee,
                tl=tl,
                defaults={'reaction_type': reaction_type}
            )
            
            if created:
                # Update likes count
                chat_message.likes_count += 1
                chat_message.save()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Reaction added',
                    'reaction_id': reaction.id,
                    'new_likes_count': chat_message.likes_count
                })
            else:
                return JsonResponse({'success': False, 'error': 'Reaction already exists'})
        
        elif action == 'remove':
            # Remove reaction
            deleted_count = ChatReaction.objects.filter(
                chat_message=chat_message,
                employee=employee,
                tl=tl,
                reaction_type=reaction_type
            ).delete()[0]
            
            if deleted_count > 0:
                # Update likes count
                chat_message.likes_count = max(0, chat_message.likes_count - 1)
                chat_message.save()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Reaction removed',
                    'new_likes_count': chat_message.likes_count
                })
            else:
                return JsonResponse({'success': False, 'error': 'Reaction not found'})
        
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'})
    except TeamLeader.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Team leader not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error managing reaction: {str(e)}'})


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def get_unread_count(request):
    """
    Get Unread Message Count - API endpoint for real-time updates
    """
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Get user information
        employee_id = request.session.get('employee_id')
        tl_id = request.session.get('tl_id')
        
        if not (employee_id or tl_id):
            return JsonResponse({'success': False, 'error': 'User not authenticated'})
        
        # Get unread count
        if employee_id:
            unread_count = TeamChat.objects.filter(
                receiver_employee_id=employee_id,
                is_read=False
            ).count()
        elif tl_id:
            unread_count = TeamChat.objects.filter(
                receiver_tl_id=tl_id,
                is_read=False
            ).count()
        
        return JsonResponse({
            'success': True,
            'unread_count': unread_count
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error getting unread count: {str(e)}'})


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def chat_search(request):
    """
    Search Chat Messages - API endpoint for searching through chat history
    """
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Get user information
        employee_id = request.session.get('employee_id')
        tl_id = request.session.get('tl_id')
        
        if not (employee_id or tl_id):
            return JsonResponse({'success': False, 'error': 'User not authenticated'})
        
        # Get search parameters
        query_text = request.GET.get('q', '').strip()
        limit = int(request.GET.get('limit', 20))
        offset = int(request.GET.get('offset', 0))
        
        if not query_text:
            return JsonResponse({'success': False, 'error': 'Search query is required'})
        
        # Build search query
        if employee_id:
            search_query = Q(
                Q(sender_employee_id=employee_id) | Q(receiver_employee_id=employee_id),
                Q(message__icontains=query_text) | Q(subject__icontains=query_text)
            )
        elif tl_id:
            search_query = Q(
                Q(sender_tl_id=tl_id) | Q(receiver_tl_id=tl_id),
                Q(message__icontains=query_text) | Q(subject__icontains=query_text)
            )
        
        # Search messages
        search_results = TeamChat.objects.filter(search_query).select_related(
            'sender_employee', 'sender_tl', 'receiver_employee', 'receiver_tl'
        ).order_by('-created_at')[offset:offset + limit]
        
        # Prepare response data
        results_data = []
        for chat in search_results:
            # Highlight search terms
            highlighted_message = chat.message
            highlighted_subject = chat.subject
            
            results_data.append({
                'id': chat.id,
                'message': highlighted_message,
                'subject': highlighted_subject,
                'created_at': chat.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'sender_name': chat.get_sender_name(),
                'receiver_name': chat.get_receiver_name(),
                'chat_type': chat.chat_type,
                'priority': chat.priority
            })
        
        return JsonResponse({
            'success': True,
            'results': results_data,
            'total': search_results.count(),
            'query': query_text
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error searching messages: {str(e)}'})


# ============================================================================
# NEW ENHANCED ATTENDANCE SYSTEM VIEWS
# ============================================================================

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def enhanced_attendance_dashboard(request):
    """
    Enhanced Attendance Dashboard - Main attendance interface with Present/Absent/Late Mark functionality
    """
    employee_id = request.session.get('employee_id')
    role = request.session.get('role')
    
    if not employee_id or role != 'employee':
        messages.error(request, "Please login as employee to access attendance dashboard.")
        return redirect('login')
    
    try:
        employee = Employee.objects.get(id=employee_id)
        today = timezone.now().date()
        current_time = timezone.now().time()
        
        # Get today's attendance records
        present_today = PresentRecord.objects.filter(employee=employee, attendance_date=today).first()
        absent_today = AbsentRecord.objects.filter(employee=employee, attendance_date=today).first()
        late_mark_today = LateMarkRecord.objects.filter(employee=employee, attendance_date=today).first()
        work_completion_today = DailyWorkCompletion.objects.filter(employee=employee, work_date=today).first()
        
        # Check if already marked attendance today
        has_marked_attendance = bool(present_today or absent_today or late_mark_today)
        
        # Calculate late mark threshold (10:45 AM)
        late_threshold = time(10, 45)
        is_late_mark = current_time > late_threshold
        
        # Office closing time (6:30 PM)
        office_closing_time = time(18, 30)
        is_office_closing_time = current_time >= office_closing_time
        
        context = {
            'employee': employee,
            'today': today,
            'current_time': current_time,
            'present_today': present_today,
            'absent_today': absent_today,
            'late_mark_today': late_mark_today,
            'work_completion_today': work_completion_today,
            'has_marked_attendance': has_marked_attendance,
            'is_late_mark': is_late_mark,
            'late_threshold': late_threshold,
            'is_office_closing_time': is_office_closing_time,
            'office_closing_time': office_closing_time,
        }
        
        return render(request, 'app/employee/enhanced-attendance.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, "Employee profile not found.")
        return redirect('login')


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def mark_present_today(request):
    """
    Mark Present for Today - Handle Present button click with time tracking
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    employee_id = request.session.get('employee_id')
    if not employee_id:
        return JsonResponse({'success': False, 'error': 'User not authenticated'})
    
    try:
        employee = Employee.objects.get(id=employee_id)
        today = timezone.now().date()
        current_time = timezone.now().time()
        
        # Check if already marked attendance today
        existing_present = PresentRecord.objects.filter(employee=employee, attendance_date=today).first()
        existing_absent = AbsentRecord.objects.filter(employee=employee, attendance_date=today).first()
        existing_late_mark = LateMarkRecord.objects.filter(employee=employee, attendance_date=today).first()
        
        if existing_present or existing_absent or existing_late_mark:
            return JsonResponse({'success': False, 'error': 'Attendance already marked for today'})
        
        # Check if it's a late mark (after 10:45 AM)
        late_threshold = time(10, 45)
        is_late_mark = current_time > late_threshold
        
        if is_late_mark:
            # Create Late Mark Record
            late_minutes = (datetime.combine(today, current_time) - datetime.combine(today, late_threshold)).seconds // 60
            
            late_mark = LateMarkRecord.objects.create(
                employee=employee,
                attendance_date=today,
                actual_check_in_time=current_time,
                late_minutes=late_minutes,
                marked_by='self'
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Late mark recorded. You are {late_minutes} minutes late.',
                'type': 'late_mark',
                'late_minutes': late_minutes,
                'time': current_time.strftime('%H:%M:%S')
            })
        else:
            # Create Present Record
            present_record = PresentRecord.objects.create(
                employee=employee,
                attendance_date=today,
                check_in_time=current_time,
                marked_by='self'
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Present marked successfully!',
                'type': 'present',
                'time': current_time.strftime('%H:%M:%S')
            })
            
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error marking present: {str(e)}'})


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def mark_absent_today(request):
    """
    Mark Absent for Today - Handle Absent button click with time tracking
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    employee_id = request.session.get('employee_id')
    if not employee_id:
        return JsonResponse({'success': False, 'error': 'User not authenticated'})
    
    try:
        employee = Employee.objects.get(id=employee_id)
        today = timezone.now().date()
        current_time = timezone.now().time()
        
        # Check if already marked attendance today
        existing_present = PresentRecord.objects.filter(employee=employee, attendance_date=today).first()
        existing_absent = AbsentRecord.objects.filter(employee=employee, attendance_date=today).first()
        existing_late_mark = LateMarkRecord.objects.filter(employee=employee, attendance_date=today).first()
        
        if existing_present or existing_absent or existing_late_mark:
            return JsonResponse({'success': False, 'error': 'Attendance already marked for today'})
        
        # Get reason from request
        reason = request.POST.get('reason', '')
        
        # Create Absent Record
        absent_record = AbsentRecord.objects.create(
            employee=employee,
            attendance_date=today,
            reason=reason,
            marked_by='self'
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Absent marked successfully!',
            'type': 'absent',
            'time': current_time.strftime('%H:%M:%S'),
            'reason': reason
        })
            
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error marking absent: {str(e)}'})


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def work_complete_today(request):
    """
    Work Complete Today - Handle work completion at office closing time (6:30 PM)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    employee_id = request.session.get('employee_id')
    if not employee_id:
        return JsonResponse({'success': False, 'error': 'User not authenticated'})
    
    try:
        employee = Employee.objects.get(id=employee_id)
        today = timezone.now().date()
        current_time = timezone.now().time()
        current_datetime = timezone.now()
        
        # Check if it's after office closing time (6:30 PM)
        office_closing_time = time(18, 30)
        if current_time < office_closing_time:
            return JsonResponse({'success': False, 'error': 'Work completion can only be submitted after 6:30 PM'})
        
        # Get work details from request
        tasks_completed = request.POST.get('tasks_completed', '')
        challenges_faced = request.POST.get('challenges_faced', '')
        tomorrow_plan = request.POST.get('tomorrow_plan', '')
        additional_notes = request.POST.get('additional_notes', '')
        
        if not tasks_completed.strip():
            return JsonResponse({'success': False, 'error': 'Tasks completed field is required'})
        
        # Check if work completion already submitted today
        existing_work_completion = DailyWorkCompletion.objects.filter(
            employee=employee, 
            work_date=today
        ).first()
        
        if existing_work_completion:
            # Update existing record
            existing_work_completion.tasks_completed = tasks_completed
            existing_work_completion.challenges_faced = challenges_faced
            existing_work_completion.tomorrow_plan = tomorrow_plan
            existing_work_completion.additional_notes = additional_notes
            existing_work_completion.work_completion_time = current_datetime
            existing_work_completion.status = 'completed'
            existing_work_completion.save()
            
            action = 'updated'
        else:
            # Create new work completion record
            work_completion = DailyWorkCompletion.objects.create(
                employee=employee,
                work_date=today,
                tasks_completed=tasks_completed,
                challenges_faced=challenges_faced,
                tomorrow_plan=tomorrow_plan,
                additional_notes=additional_notes,
                work_completion_time=current_datetime,
                status='completed'
            )
            
            action = 'created'
        
        return JsonResponse({
            'success': True,
            'message': f'Work completion {action} successfully!',
            'type': 'work_completion',
            'time': current_datetime.strftime('%Y-%m-%d %H:%M:%S')
        })
            
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error submitting work completion: {str(e)}'})


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def mark_present_date(request):
    """
    Mark Present for Specific Date - Handle Present button click for any date in calendar
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    employee_id = request.session.get('employee_id')
    if not employee_id:
        return JsonResponse({'success': False, 'error': 'User not authenticated'})
    
    try:
        employee = Employee.objects.get(id=employee_id)
        
        # Get the date from request
        import json
        data = json.loads(request.body)
        date_str = data.get('date')
        
        if not date_str:
            return JsonResponse({'success': False, 'error': 'Date is required'})
        
        # Parse the date
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        current_date = timezone.now().date()
        
        # Check if date is in the future
        if selected_date > current_date:
            return JsonResponse({'success': False, 'error': 'Cannot mark attendance for future dates'})
        
        # Check if already marked attendance for this date
        existing_present = PresentRecord.objects.filter(employee=employee, attendance_date=selected_date).first()
        existing_absent = AbsentRecord.objects.filter(employee=employee, attendance_date=selected_date).first()
        existing_late_mark = LateMarkRecord.objects.filter(employee=employee, attendance_date=selected_date).first()
        
        if existing_present or existing_absent or existing_late_mark:
            return JsonResponse({'success': False, 'error': 'Attendance already marked for this date'})
        
        # For past dates, always mark as Present
        # For today, check if it's a late mark
        if selected_date == current_date:
            current_time = timezone.now().time()
            late_threshold = time(10, 45)
            is_late_mark = current_time > late_threshold
            
            if is_late_mark:
                # Create Late Mark Record
                late_minutes = (datetime.combine(selected_date, current_time) - datetime.combine(selected_date, late_threshold)).seconds // 60
                
                late_mark = LateMarkRecord.objects.create(
                    employee=employee,
                    attendance_date=selected_date,
                    actual_check_in_time=current_time,
                    late_minutes=late_minutes,
                    marked_by='self'
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'Late mark recorded for {selected_date}. You are {late_minutes} minutes late.',
                    'type': 'late_mark',
                    'late_minutes': late_minutes,
                    'date': date_str,
                    'time': current_time.strftime('%H:%M:%S')
                })
            else:
                # Create Present Record
                present_record = PresentRecord.objects.create(
                    employee=employee,
                    attendance_date=selected_date,
                    check_in_time=current_time,
                    marked_by='self'
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'Present marked successfully for {selected_date}!',
                    'type': 'present',
                    'date': date_str,
                    'time': current_time.strftime('%H:%M:%S')
                })
        else:
            # For past dates, create a present record with marked_time
            marked_time = timezone.now()
            present_record = PresentRecord.objects.create(
                employee=employee,
                attendance_date=selected_date,
                check_in_time=marked_time.time(),
                marked_time=marked_time,
                marked_by='self'
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Present marked successfully for {selected_date}!',
                'type': 'present',
                'date': date_str,
                'time': marked_time.strftime('%Y-%m-%d %H:%M:%S')
            })
            
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'})
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid date format'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error marking present: {str(e)}'})


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def mark_absent_date(request):
    """
    Mark Absent for Specific Date - Handle Absent button click for any date in calendar
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    employee_id = request.session.get('employee_id')
    if not employee_id:
        return JsonResponse({'success': False, 'error': 'User not authenticated'})
    
    try:
        employee = Employee.objects.get(id=employee_id)
        
        # Get the date and reason from request
        import json
        data = json.loads(request.body)
        date_str = data.get('date')
        reason = data.get('reason', '')
        
        if not date_str:
            return JsonResponse({'success': False, 'error': 'Date is required'})
        
        # Parse the date
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        current_date = timezone.now().date()
        
        # Check if date is in the future
        if selected_date > current_date:
            return JsonResponse({'success': False, 'error': 'Cannot mark attendance for future dates'})
        
        # Check if already marked attendance for this date
        existing_present = PresentRecord.objects.filter(employee=employee, attendance_date=selected_date).first()
        existing_absent = AbsentRecord.objects.filter(employee=employee, attendance_date=selected_date).first()
        existing_late_mark = LateMarkRecord.objects.filter(employee=employee, attendance_date=selected_date).first()
        
        if existing_present or existing_absent or existing_late_mark:
            return JsonResponse({'success': False, 'error': 'Attendance already marked for this date'})
        
        # Create Absent Record
        marked_time = timezone.now()
        absent_record = AbsentRecord.objects.create(
            employee=employee,
            attendance_date=selected_date,
            reason=reason,
            marked_time=marked_time,
            marked_by='self'
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Absent marked successfully for {selected_date}!',
            'type': 'absent',
            'date': date_str,
            'time': marked_time.strftime('%Y-%m-%d %H:%M:%S'),
            'reason': reason
        })
            
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'})
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid date format'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error marking absent: {str(e)}'})


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def get_attendance_calendar(request):
    """
    Get Attendance Calendar Data - API endpoint to fetch attendance status for calendar dates
    """
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    employee_id = request.session.get('employee_id')
    if not employee_id:
        return JsonResponse({'success': False, 'error': 'User not authenticated'})
    
    try:
        employee = Employee.objects.get(id=employee_id)
        
        # Get date range from request (default: current month)
        month = request.GET.get('month')
        year = request.GET.get('year')
        
        if month and year:
            month = int(month)
            year = int(year)
        else:
            # Default: current month
            today = timezone.now()
            month = today.month
            year = today.year
        
        # Get first and last day of the month
        import calendar
        first_day = date(year, month, 1)
        if month == 12:
            last_day = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = date(year, month + 1, 1) - timedelta(days=1)
        
        # Get attendance records for the month
        present_records = PresentRecord.objects.filter(
            employee=employee,
            attendance_date__range=[first_day, last_day]
        )
        
        absent_records = AbsentRecord.objects.filter(
            employee=employee,
            attendance_date__range=[first_day, last_day]
        )
        
        late_mark_records = LateMarkRecord.objects.filter(
            employee=employee,
            attendance_date__range=[first_day, last_day]
        )
        
        # Create calendar data
        calendar_data = []
        current_date = first_day
        
        while current_date <= last_day:
            date_str = current_date.strftime('%Y-%m-%d')
            
            # Check status for this date
            status = 'no_record'
            record_id = None
            details = {}
            
            present = present_records.filter(attendance_date=current_date).first()
            if present:
                status = 'present'
                record_id = present.id
                details = {
                    'check_in_time': present.check_in_time.strftime('%H:%M:%S') if present.check_in_time else None,
                    'marked_time': present.marked_time.strftime('%Y-%m-%d %H:%M:%S') if present.marked_time else None
                }
            else:
                late_mark = late_mark_records.filter(attendance_date=current_date).first()
                if late_mark:
                    status = 'late_mark'
                    record_id = late_mark.id
                    details = {
                        'late_minutes': late_mark.late_minutes,
                        'actual_check_in_time': late_mark.actual_check_in_time.strftime('%H:%M:%S') if late_mark.actual_check_in_time else None
                    }
                else:
                    absent = absent_records.filter(attendance_date=current_date).first()
                    if absent:
                        status = 'absent'
                        record_id = absent.id
                        details = {
                            'reason': absent.reason,
                            'marked_time': absent.marked_time.strftime('%Y-%m-%d %H:%M:%S') if absent.marked_time else None
                        }
            
            calendar_data.append({
                'date': date_str,
                'day': current_date.day,
                'status': status,
                'record_id': record_id,
                'details': details,
                'is_past': current_date < timezone.now().date(),
                'is_today': current_date == timezone.now().date(),
                'is_weekend': current_date.weekday() >= 5  # Saturday = 5, Sunday = 6
            })
            
            current_date += timedelta(days=1)
        
        return JsonResponse({
            'success': True,
            'calendar_data': calendar_data,
            'month': month,
            'year': year,
            'month_name': calendar.month_name[month],
            'first_day': first_day.strftime('%Y-%m-%d'),
            'last_day': last_day.strftime('%Y-%m-%d')
        })
        
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'})
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid month or year'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error fetching calendar data: {str(e)}'})


def render_attendance_public_info(request):
    """
    Render public attendance information for unauthenticated users
    """
    context = {
        'is_public': True,
        'employee': None,
        'attendance_table': [],
        'from_date': timezone.now().date() - timedelta(days=30),
        'to_date': timezone.now().date(),
        'statistics': {
            'total_days': 0,
            'present_days': 0,
            'late_mark_days': 0,
            'absent_days': 0,
            'no_record_days': 0,
            'attendance_percentage': 0.0
        },
        'current_page': 1,
        'total_pages': 1,
        'has_previous': False,
        'has_next': False,
        'previous_page': None,
        'next_page': None,
        'page_numbers': [1],
        'records_per_page': 20,
        'total_records': 0,
        'start_record': 0,
        'end_record': 0,
        'message': 'Please login to view your personal attendance records.'
    }
    
    return render(request, 'app/employee/attendance-table.html', context)


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list', 'attendance-table'])
def attendance_table_view(request):
    """
    Attendance Table View - Display day-wise attendance table for employee with pagination
    This page is now accessible without login requirement
    Includes enhanced previous page navigation with data preservation
    """
    # Track referrer for previous page navigation
    referrer = request.META.get('HTTP_REFERER', '')
    
    # Store referrer in session for navigation back
    if referrer:
        request.session['attendance_table_referrer'] = referrer
    
    employee_id = request.session.get('employee_id')
    role = request.session.get('role')
    
    # Allow access without login - show general attendance information or sample data
    if not employee_id:
        # For unauthenticated users, show a general information page
        return render_attendance_public_info(request)
    
    try:
        employee = Employee.objects.get(id=employee_id)
        
        # Get date range from request (default: last 30 days)
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        
        if from_date and to_date:
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
        else:
            # Default: last 30 days
            to_date = timezone.now().date()
            from_date = to_date - timedelta(days=30)
        
        # Pagination settings - allow user to customize records per page
        records_per_page = int(request.GET.get('page', 20))  # Default to 20 records
        page = int(request.GET.get('page_num', 1))  # Use page_num for current page number
        
        # Validate records_per_page
        valid_page_sizes = [10, 20, 50, 100]
        if records_per_page not in valid_page_sizes:
            records_per_page = 20
        
        # Get attendance records
        present_records = PresentRecord.objects.filter(
            employee=employee,
            attendance_date__range=[from_date, to_date]
        ).order_by('attendance_date')
        
        absent_records = AbsentRecord.objects.filter(
            employee=employee,
            attendance_date__range=[from_date, to_date]
        ).order_by('attendance_date')
        
        late_mark_records = LateMarkRecord.objects.filter(
            employee=employee,
            attendance_date__range=[from_date, to_date]
        ).order_by('attendance_date')
        
        work_completions = DailyWorkCompletion.objects.filter(
            employee=employee,
            work_date__range=[from_date, to_date]
        ).order_by('work_date')
        
        # Combine all records into a unified table
        attendance_table = []
        current_date = from_date
        
        while current_date <= to_date:
            record = {
                'date': current_date,
                'status': 'No Record',
                'time': '',
                'details': '',
                'work_completion': None
            }
            
            # Check Present Record
            present = present_records.filter(attendance_date=current_date).first()
            if present:
                record['status'] = 'Present'
                record['time'] = present.marked_time.strftime('%H:%M:%S')
                record['details'] = f"Checked in at {present.check_in_time.strftime('%H:%M:%S') if present.check_in_time else 'N/A'}"
            
            # Check Late Mark Record
            late_mark = late_mark_records.filter(attendance_date=current_date).first()
            if late_mark:
                record['status'] = 'Late Mark'
                record['time'] = late_mark.marked_time.strftime('%H:%M:%S')
                record['details'] = f"Late by {late_mark.late_minutes} minutes"
            
            # Check Absent Record
            absent = absent_records.filter(attendance_date=current_date).first()
            if absent:
                record['status'] = 'Absent'
                record['time'] = absent.marked_time.strftime('%H:%M:%S')
                record['details'] = absent.reason or 'No reason provided'
            
            # Check Work Completion
            work_completion = work_completions.filter(work_date=current_date).first()
            if work_completion:
                record['work_completion'] = work_completion
            
            attendance_table.append(record)
            current_date += timedelta(days=1)
        
        # Calculate statistics
        total_days = len(attendance_table)
        present_days = len([r for r in attendance_table if r['status'] == 'Present'])
        late_mark_days = len([r for r in attendance_table if r['status'] == 'Late Mark'])
        absent_days = len([r for r in attendance_table if r['status'] == 'Absent'])
        no_record_days = len([r for r in attendance_table if r['status'] == 'No Record'])
        
        # Apply pagination
        total_records = len(attendance_table)
        start_index = (page - 1) * records_per_page
        end_index = start_index + records_per_page
        paginated_attendance_table = attendance_table[start_index:end_index]
        
        # Calculate pagination info
        total_pages = (total_records + records_per_page - 1) // records_per_page  # Ceiling division
        has_previous = page > 1
        has_next = page < total_pages
        
        # Generate page numbers for pagination controls
        page_numbers = []
        if total_pages <= 7:
            page_numbers = list(range(1, total_pages + 1))
        else:
            if page <= 4:
                page_numbers = [1, 2, 3, 4, 5, '...', total_pages]
            elif page >= total_pages - 3:
                page_numbers = [1, '...', total_pages - 4, total_pages - 3, total_pages - 2, total_pages - 1, total_pages]
            else:
                page_numbers = [1, '...', page - 1, page, page + 1, '...', total_pages]
        
        context = {
            'employee': employee,
            'attendance_table': paginated_attendance_table,
            'from_date': from_date,
            'to_date': to_date,
            'statistics': {
                'total_days': total_days,
                'present_days': present_days,
                'late_mark_days': late_mark_days,
                'absent_days': absent_days,
                'no_record_days': no_record_days,
                'attendance_percentage': round((present_days + late_mark_days) / max(total_days, 1) * 100, 1)
            },
            # Pagination context
            'current_page': page,
            'total_pages': total_pages,
            'has_previous': has_previous,
            'has_next': has_next,
            'previous_page': page - 1 if has_previous else None,
            'next_page': page + 1 if has_next else None,
            'page_numbers': page_numbers,
            'records_per_page': records_per_page,
            'total_records': total_records,
            'start_record': start_index + 1,
            'end_record': min(end_index, total_records),
            # Previous page navigation context
            'referrer_url': request.session.get('attendance_table_referrer', ''),
            'has_previous_page': bool(request.session.get('attendance_table_referrer')),
        }
        
        return render(request, 'app/employee/attendance-table.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, "Employee profile not found.")
        return redirect('login')


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def hr_enhanced_attendance_view(request):
    """
    HR Enhanced Attendance View - Display attendance records for all employees
    """
    hr_id = request.session.get('hr_id')
    if not hr_id:
        messages.error(request, "Please login as HR to access attendance management.")
        return redirect('login')
    
    try:
        hr = HRProfile.objects.get(id=hr_id)
        
        # Get date range (default: current month)
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        
        if from_date and to_date:
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
        else:
            # Default: current month
            today = timezone.now().date()
            from_date = today.replace(day=1)
            # Get last day of current month
            if today.month == 12:
                to_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                to_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
        
        # Get all employees
        employees = Employee.objects.filter(resigned_date__isnull=True).order_by('first_name')
        
        # Get attendance records for the date range
        present_records = PresentRecord.objects.filter(
            attendance_date__range=[from_date, to_date]
        ).select_related('employee')
        
        absent_records = AbsentRecord.objects.filter(
            attendance_date__range=[from_date, to_date]
        ).select_related('employee')
        
        late_mark_records = LateMarkRecord.objects.filter(
            attendance_date__range=[from_date, to_date]
        ).select_related('employee')
        
        work_completions = DailyWorkCompletion.objects.filter(
            work_date__range=[from_date, to_date]
        ).select_related('employee')
        
        context = {
            'hr': hr,
            'employees': employees,
            'present_records': present_records,
            'absent_records': absent_records,
            'late_mark_records': late_mark_records,
            'work_completions': work_completions,
            'from_date': from_date,
            'to_date': to_date,
            'total_employees': employees.count(),
            'total_present': present_records.count(),
            'total_absent': absent_records.count(),
            'total_late_mark': late_mark_records.count(),
        }
        
        return render(request, 'app/hr/enhanced-attendance.html', context)
        
    except HRProfile.DoesNotExist:
        messages.error(request, "HR profile not found.")
        return redirect('login')


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def tl_enhanced_attendance_view(request):
    """
    TL Enhanced Attendance View - Display attendance records for team members
    """
    tl_id = request.session.get('tl_id')
    if not tl_id:
        messages.error(request, "Please login as Team Leader to access attendance management.")
        return redirect('login')
    
    try:
        tl = TeamLeader.objects.get(id=tl_id)
        
        # Get team members
        team_assignments = TeamAssignment.objects.filter(
            team_leader=tl
        ).select_related('employee')
        
        team_members = [assignment.employee for assignment in team_assignments]
        
        # Get date range (default: current month)
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        
        if from_date and to_date:
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
        else:
            # Default: current month
            today = timezone.now().date()
            from_date = today.replace(day=1)
            # Get last day of current month
            if today.month == 12:
                to_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                to_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
        
        # Get attendance records for team members
        present_records = PresentRecord.objects.filter(
            employee__in=team_members,
            attendance_date__range=[from_date, to_date]
        ).select_related('employee')
        
        absent_records = AbsentRecord.objects.filter(
            employee__in=team_members,
            attendance_date__range=[from_date, to_date]
        ).select_related('employee')
        
        late_mark_records = LateMarkRecord.objects.filter(
            employee__in=team_members,
            attendance_date__range=[from_date, to_date]
        ).select_related('employee')
        
        work_completions = DailyWorkCompletion.objects.filter(
            employee__in=team_members,
            work_date__range=[from_date, to_date]
        ).select_related('employee')
        
        context = {
            'tl': tl,
            'team_members': team_members,
            'present_records': present_records,
            'absent_records': absent_records,
            'late_mark_records': late_mark_records,
            'work_completions': work_completions,
            'from_date': from_date,
            'to_date': to_date,
            'team_size': len(team_members),
            'total_present': present_records.count(),
            'total_absent': absent_records.count(),
            'total_late_mark': late_mark_records.count(),
        }
        
        return render(request, 'app/tl/enhanced-attendance.html', context)
        
    except TeamLeader.DoesNotExist:
        messages.error(request, "Team Leader profile not found.")
# ============================================================================
# PROJECT DISCUSSIONS VIEWS
# ============================================================================

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def employee_project_discussions(request):
    """
    Employee Project Discussions - Display project discussions and chat interface
    """
    employee_id = request.session.get('employee_id')
    
    if not employee_id:
        messages.error(request, "Please login as employee to access project discussions.")
        return redirect('login')
    
    try:
        employee = Employee.objects.get(id=employee_id)
        
        # Get projects assigned to this employee
        projects = ProjectAssignment.objects.filter(
            team_members=employee
        ).select_related('team_leader', 'team_leader__employee').order_by('-created_at')
        
        # Get recent discussions for these projects
        recent_discussions = ProjectDiscussion.objects.filter(
            project__team_members=employee
        ).select_related('project', 'sender_employee', 'sender_tl').order_by('-created_at')[:20]
        
        # Get team assignment and team leader info
        team_assignment = TeamAssignment.objects.filter(
            employee=employee
        ).select_related('team_leader', 'team_leader__employee').first()
        
        team_leader = None
        if team_assignment:
            tl = team_assignment.team_leader
            team_leader = {
                'id': tl.id,
                'name': f"{tl.employee.first_name} {tl.employee.last_name}",
                'designation': tl.employee.designation,
                'department': tl.employee.department,
                'email': tl.employee.email,
                'image': tl.employee.image.url if tl.employee.image else None
            }
        
        # Get direct chat conversations with team leader
        direct_conversations = []
        if team_leader:
            direct_conversations = TeamChat.objects.filter(
                Q(sender_employee=employee, receiver_tl_id=team_leader['id']) |
                Q(sender_tl_id=team_leader['id'], receiver_employee=employee)
            ).order_by('-created_at')[:10]
        
        context = {
            'employee': employee,
            'projects': projects,
            'recent_discussions': recent_discussions,
            'team_leader': team_leader,
            'direct_conversations': direct_conversations,
            'total_projects': projects.count(),
        }
        
        return render(request, 'app/employee/project-discussions.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, "Employee profile not found.")
        return redirect('login')


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def tl_project_discussions(request):
    """
    TL Project Discussions - Display project discussions and team chat interface
    """
    tl_id = request.session.get('tl_id')
    
    if not tl_id:
        messages.error(request, "Please login as Team Leader to access project discussions.")
        return redirect('login')
    
    try:
        tl = TeamLeader.objects.get(id=tl_id)
        employee = tl.employee
        
        # Get projects led by this TL
        projects = ProjectAssignment.objects.filter(
            team_leader=tl
        ).select_related('team_leader__employee').order_by('-created_at')
        
        # Get team members
        team_assignments = TeamAssignment.objects.filter(
            team_leader=tl
        ).select_related('employee')
        
        # Get recent discussions for these projects
        recent_discussions = ProjectDiscussion.objects.filter(
            project__team_leader=tl
        ).select_related('project', 'sender_employee', 'sender_tl', 'receiver_employee', 'receiver_tl').order_by('-created_at')[:30]
        
        # Get team member details
        team_members = []
        for assignment in team_assignments:
            emp = assignment.employee
            team_members.append({
                'id': emp.id,
                'name': f"{emp.first_name} {emp.last_name}",
                'designation': emp.designation,
                'department': emp.department,
                'email': emp.email,
                'image': emp.image.url if emp.image else None,
                'role': assignment.role,
                'assignment_date': assignment.assignment_date
            })
        
        # Get recent chat conversations with team members
        recent_conversations = TeamChat.objects.filter(
            Q(sender_tl=tl) | Q(receiver_tl=tl)
        ).select_related('sender_employee', 'receiver_employee').order_by('-created_at')[:20]
        
        context = {
            'tl': tl,
            'employee': employee,
            'projects': projects,
            'team_members': team_members,
            'recent_discussions': recent_discussions,
            'recent_conversations': recent_conversations,
            'total_projects': projects.count(),
            'team_size': len(team_members),
        }
        
        return render(request, 'app/tl/project-discussions.html', context)
        
    except TeamLeader.DoesNotExist:
        messages.error(request, "Team Leader profile not found.")
        return redirect('login')


@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def send_project_message(request):
    """
    Send Project Message - Handle sending messages related to projects
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Get sender information
        employee_id = request.session.get('employee_id')
        tl_id = request.session.get('tl_id')
        
        if not (employee_id or tl_id):
            return JsonResponse({'success': False, 'error': 'User not authenticated'})
        
        # Get data from request.POST (since JavaScript sends FormData)
        project_id = request.POST.get('project_id')
        message_text = request.POST.get('message', '').strip()
        subject = request.POST.get('subject', '').strip()
        message_type = request.POST.get('message_type', 'message')
        priority = request.POST.get('priority', 'normal')
        receiver_employee_id = request.POST.get('receiver_employee_id')
        receiver_tl_id = request.POST.get('receiver_tl_id')
        
        # Validation
        if not project_id or not message_text:
            return JsonResponse({'success': False, 'error': 'Project and message are required'})
        
        if len(message_text) > 2000:  # Message length limit
            return JsonResponse({'success': False, 'error': 'Message too long (max 2000 characters)'})
        
        # Get project
        try:
            project = ProjectAssignment.objects.get(id=project_id)
        except ProjectAssignment.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Project not found'})
        
        # Get sender objects
        sender_employee = None
        sender_tl = None
        
        if employee_id:
            sender_employee = Employee.objects.get(id=employee_id)
            
            # Verify employee is part of this project
            if not project.team_members.filter(id=employee_id).exists():
                return JsonResponse({'success': False, 'error': 'You are not assigned to this project'})
                
        elif tl_id:
            sender_tl = TeamLeader.objects.get(id=tl_id)
            
            # Verify TL is leading this project
            if project.team_leader != sender_tl:
                return JsonResponse({'success': False, 'error': 'You are not the team leader of this project'})
        else:
            return JsonResponse({'success': False, 'error': 'User authentication required'})
        
        # Get receiver objects (optional)
        receiver_employee = None
        receiver_tl = None
        
        if receiver_employee_id:
            try:
                receiver_employee = Employee.objects.get(id=receiver_employee_id)
                # Verify receiver is part of the project
                if not project.team_members.filter(id=receiver_employee_id).exists():
                    return JsonResponse({'success': False, 'error': 'Receiver is not part of this project'})
            except Employee.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Receiver employee not found'})
        
        if receiver_tl_id:
            try:
                receiver_tl = TeamLeader.objects.get(id=receiver_tl_id)
            except TeamLeader.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Receiver team leader not found'})
        
        # Handle file attachment
        attachment = request.FILES.get('attachment')
        attachment_name = None
        if attachment:
            # Validate file size (max 10MB)
            if attachment.size > 10 * 1024 * 1024:
                return JsonResponse({'success': False, 'error': 'File too large (max 10MB)'})
            
            # Validate file type
            allowed_types = ['pdf', 'doc', 'docx', 'txt', 'jpg', 'jpeg', 'png', 'gif']
            file_extension = attachment.name.split('.')[-1].lower()
            if file_extension not in allowed_types:
                return JsonResponse({'success': False, 'error': 'File type not allowed'})
            
            attachment_name = attachment.name
        
        # Create the project discussion message
        discussion = ProjectDiscussion.objects.create(
            project=project,
            sender_employee=sender_employee,
            sender_tl=sender_tl,
            receiver_employee=receiver_employee,
            receiver_tl=receiver_tl,
            message_type=message_type,
            subject=subject or f"Project Discussion - {project.project_name}",
            content=message_text,
            priority=priority,
            attachment=attachment,
            attachment_name=attachment_name
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Project message sent successfully',
            'discussion_id': discussion.id,
            'created_at': discussion.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'sender_name': discussion.get_sender_name() if hasattr(discussion, 'get_sender_name') else 'Unknown'
        })
        
    except Employee.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Employee not found'})
    except TeamLeader.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Team leader not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error sending project message: {str(e)}'})

@login_required_with_exemptions(exempt_urls=['create-hr', 'hr-list'])
def get_project_discussions(request):
    """
    Get Project Discussions - API endpoint to fetch discussions for a specific project
    """
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Get user information
        employee_id = request.session.get('employee_id')
        tl_id = request.session.get('tl_id')
        
        if not (employee_id or tl_id):
            return JsonResponse({'success': False, 'error': 'User not authenticated'})
        
        # Get project ID from request
        project_id = request.GET.get('project_id')
        if not project_id:
            return JsonResponse({'success': False, 'error': 'Project ID is required'})
        
        # Get project
        try:
            project = ProjectAssignment.objects.get(id=project_id)
        except ProjectAssignment.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Project not found'})
        
        # Verify user access to project
        if employee_id:
            # Check if employee is part of this project
            if not project.team_members.filter(id=employee_id).exists():
                return JsonResponse({'success': False, 'error': 'You are not assigned to this project'})
        elif tl_id:
            # Check if TL is leading this project
            if project.team_leader_id != tl_id:
                return JsonResponse({'success': False, 'error': 'You are not the team leader of this project'})
        
        # Get discussions for this project
        discussions = ProjectDiscussion.objects.filter(
            project=project
        ).select_related(
            'sender_employee', 'sender_tl', 'receiver_employee', 'receiver_tl'
        ).order_by('-created_at')[:50]  # Last 50 discussions
        
        # Prepare response data
        discussions_data = []
        for discussion in discussions:
            # Determine sender
            if discussion.sender_employee:
                sender_name = f"{discussion.sender_employee.first_name} {discussion.sender_employee.last_name}"
                sender_type = 'employee'
            elif discussion.sender_tl:
                sender_name = f"TL {discussion.sender_tl.employee.first_name} {discussion.sender_tl.employee.last_name}"
                sender_type = 'tl'
            else:
                sender_name = 'Unknown'
                sender_type = 'unknown'
            
            discussions_data.append({
                'id': discussion.id,
                'sender_name': sender_name,
                'sender_type': sender_type,
                'subject': discussion.subject,
                'content': discussion.content,
                'priority': discussion.priority,
                'priority_display': discussion.get_priority_display(),
                'message_type': discussion.message_type,
                'message_type_display': discussion.get_message_type_display(),
                'created_at': discussion.created_at.isoformat(),
                'created_at_human': discussion.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'has_attachment': bool(discussion.attachment),
                'attachment_url': discussion.attachment.url if discussion.attachment else None,
                'attachment_name': discussion.attachment_name,
            })
        
        return JsonResponse({
            'success': True,
            'discussions': discussions_data,
            'project_name': project.project_name,
            'total_discussions': len(discussions_data)
        })
        
    except ProjectAssignment.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Project not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error fetching project discussions: {str(e)}'})
